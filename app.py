import os
import io
import math
import zipfile
import re
import time
import socket
import ssl
import smtplib
import datetime
import html
import logging
import json
import hashlib
from datetime import timedelta
from logging.handlers import RotatingFileHandler
from urllib.parse import urlparse
from urllib.request import urlopen
from functools import wraps
from email.message import EmailMessage
from openpyxl import Workbook, load_workbook

from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
import jwt
from ldap3 import Connection, Server, SUBTREE, Tls
from ldap3.core.exceptions import LDAPExceptionError, LDAPInvalidFilterError
from ldap3.utils.conv import escape_filter_chars
from sqlalchemy import func, text, cast, String, or_
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL", "sqlite:///inventory.db"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-key")
app.config["LOG_DIR"] = os.environ.get("LOG_DIR", "/data/logs")
app.config["LOG_FILE"] = os.environ.get("LOG_FILE", "app.log")
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=14)
LOG_PAGE_EXCLUDE_PREFIXES = (
    "/static",
    "/branding/logo",
    "/favicon.ico",
)
LOG_PAGE_EXCLUDE_ENDPOINTS = {
    "logs_tail",
    "view_logs",
}
JWT_ACCESS_SECONDS = 15 * 60
JWT_REFRESH_SECONDS = 14 * 24 * 60 * 60
JWT_ALGORITHM = "HS256"
APP_START_TIME = datetime.datetime.utcnow()
DOCKERHUB_REPO = "tayyabtahir/assetmanager"
APP_VERSION = os.environ.get("APP_VERSION", "1.0.0").strip()
UPDATE_CHECK_CACHE = {
    "timestamp": 0.0,
    "available": False,
    "last_updated": None,
    "latest_digest": None,
    "current_digest": None,
    "current_tag": APP_VERSION or None,
    "error": None,
}

SECTION_ENDPOINTS = {
    "index",
    "free_inventory",
    "user_assets",
    "list_assets",
    "list_custom_assets",
    "list_users",
    "list_groups",
    "list_roles",
    "list_asset_types",
    "manage_departments",
    "audit_log",
    "view_logs",
    "ldap_settings",
    "smtp_settings",
    "branding_settings",
}

db = SQLAlchemy(app)

_LDAP_USER_CACHE = {"users": [], "timestamp": 0.0}
_LDAP_GROUP_CACHE = {"groups": [], "timestamp": 0.0}
_LDAP_USER_RECORDS_CACHE = {"records": [], "timestamp": 0.0}
_DEPT_CACHE = {"items": [], "timestamp": 0.0}
_DB_INIT_DONE = False
_REPORT_CACHE = {"items": {}}
REPORT_CACHE_SECONDS = 30
DEFAULT_PAGE_SIZE = 25
_USER_DISPLAY_CACHE = {"timestamp": 0, "map": {}}
ASSET_TITLE_OVERRIDES = {
    "laptops": "Laptops",
    "computers": "Computers",
    "screens": "Screens",
    "keyboards": "Keyboards",
    "mice": "Mice",
    "headsets": "Headsets",
    "ram": "RAM",
}
STATUS_OPTIONS = ["In Stock", "Assigned", "Broken", "Write Off"]
STATUS_LABELS = {
    "in stock": "In Stock",
    "in_stock": "In Stock",
    "assigned": "Assigned",
    "broken": "Broken",
    "write off": "Write Off",
    "write_off": "Write Off",
}


def setup_logging():
    log_dir = app.config["LOG_DIR"]
    os.makedirs(log_dir, exist_ok=True)
    log_path = os.path.join(log_dir, app.config["LOG_FILE"])
    handler = RotatingFileHandler(log_path, maxBytes=5_000_000, backupCount=5)
    formatter = logging.Formatter("%(asctime)s %(levelname)s %(name)s %(message)s")
    handler.setFormatter(formatter)
    handler.setLevel(logging.INFO)
    app.logger.setLevel(logging.INFO)
    if not any(isinstance(h, RotatingFileHandler) for h in app.logger.handlers):
        app.logger.addHandler(handler)
    logging.getLogger("werkzeug").addHandler(handler)


setup_logging()


@app.teardown_request
def log_unhandled_exception(exc):
    if exc is not None:
        app.logger.exception("Unhandled exception", exc_info=exc)


@app.before_request
def log_page_views():
    if not request:
        return
    if request.path.startswith("/api"):
        return
    if request.method != "GET":
        return
    if request.endpoint in LOG_PAGE_EXCLUDE_ENDPOINTS:
        return
    path = request.path or ""
    for prefix in LOG_PAGE_EXCLUDE_PREFIXES:
        if path.startswith(prefix):
            return
    user = get_current_user()
    username = user.username if user else "-"
    ip_address = request.remote_addr or "-"
    query = request.query_string.decode("utf-8", errors="ignore")
    url = f"{path}?{query}" if query else path
    app.logger.info("page_view user=%s ip=%s url=%s", username, ip_address, url)


@app.before_request
def handle_api_preflight():
    if request.method == "OPTIONS" and request.path.startswith("/api"):
        resp = app.response_class("", status=204)
        resp.headers["Access-Control-Allow-Origin"] = "*"
        resp.headers["Access-Control-Allow-Headers"] = "Authorization, Content-Type"
        resp.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
        return resp


@app.after_request
def add_api_headers(response):
    if request.path.startswith("/api"):
        response.headers["Access-Control-Allow-Origin"] = "*"
        response.headers["Access-Control-Allow-Headers"] = "Authorization, Content-Type"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
    return response


def _normalize_asset_payload(definition, payload, existing=None):
    data = {}
    for field_name, _label, field_type in definition["fields"]:
        value = payload.get(field_name)
        if field_type == "checkbox":
            data[field_name] = bool(value)
        elif field_type == "number":
            data[field_name] = _parse_int_value(value, 0)
        else:
            data[field_name] = str(value).strip() if value is not None else ""
    if definition["model"] is Mouse:
        apply_mouse_connection_to_data(data)
    if "assigned_to" in data and data["assigned_to"] == "":
        data["assigned_to"] = "free"
    if "status" in data:
        status_norm = normalize_status(data["status"])
        assigned_norm = normalize_assignee(data.get("assigned_to"))
        if status_norm in {"broken", "write off"}:
            data["status"] = "Broken" if status_norm == "broken" else "Write Off"
            data["assigned_to"] = "free"
            if "dept" in data:
                data["dept"] = ""
        elif status_norm in {"in stock", "in_stock"}:
            data["status"] = "In Stock"
            data["assigned_to"] = "free"
            if "dept" in data:
                data["dept"] = ""
        elif assigned_norm and assigned_norm != "free":
            data["status"] = "Assigned"
        else:
            data["status"] = "In Stock"
    if "asset_tag" in data:
        data["asset_tag"] = data["asset_tag"].strip()
        if data["asset_tag"]:
            existing_query = definition["model"].query.filter(
                func.lower(definition["model"].asset_tag) == data["asset_tag"].lower()
            )
            if existing is not None:
                existing_query = existing_query.filter(definition["model"].id != existing.id)
            if existing_query.first():
                return None, "Asset tag must be unique."
    if "total_quantity" in data and "assigned_quantity" in data:
        if data["assigned_quantity"] < 0 or data["total_quantity"] < 0:
            return None, "Quantities cannot be negative."
        if data["assigned_quantity"] > data["total_quantity"]:
            return None, "Assigned quantity cannot exceed total quantity."
    return data, None


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(255), nullable=True)


class Role(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    can_add = db.Column(db.Boolean, nullable=False, default=False)
    can_delete = db.Column(db.Boolean, nullable=False, default=False)
    can_read = db.Column(db.Boolean, nullable=False, default=True)
    is_app_admin = db.Column(db.Boolean, nullable=False, default=False)
    can_bulk_delete = db.Column(db.Boolean, nullable=False, default=False)
    can_manage_depts = db.Column(db.Boolean, nullable=False, default=False)


class RolePermission(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    role_id = db.Column(db.Integer, db.ForeignKey("role.id"), nullable=False)
    asset_type = db.Column(db.String(50), nullable=False)
    can_add = db.Column(db.Boolean, nullable=False, default=False)
    can_delete = db.Column(db.Boolean, nullable=False, default=False)
    can_read = db.Column(db.Boolean, nullable=False, default=False)
    can_bulk_delete = db.Column(db.Boolean, nullable=False, default=False)

    __table_args__ = (db.UniqueConstraint("role_id", "asset_type", name="uq_role_asset"),)


class LdapConfig(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    server = db.Column(db.String(255))
    base_dn = db.Column(db.String(255))
    bind_dn = db.Column(db.String(255))
    bind_password = db.Column(db.String(255))
    user_filter = db.Column(db.String(255))
    list_filter = db.Column(db.String(255))
    user_attribute = db.Column(db.String(80))
    email_attribute = db.Column(db.String(80))
    group_filter = db.Column(db.String(255))
    group_attribute = db.Column(db.String(80))
    group_member_attribute = db.Column(db.String(80))
    user_dn_template = db.Column(db.String(255))
    use_ssl = db.Column(db.Boolean, nullable=False, default=False)
    start_tls = db.Column(db.Boolean, nullable=False, default=False)
    cache_seconds = db.Column(db.Integer)
    list_limit = db.Column(db.Integer)
    default_role = db.Column(db.String(50))
    admin_users = db.Column(db.String(255))


class Group(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(80), unique=True, nullable=False)
    role = db.Column(db.String(50), nullable=False, default="unassigned")


class GroupMember(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    group_id = db.Column(db.Integer, db.ForeignKey("group.id"), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)

    __table_args__ = (
        db.UniqueConstraint("group_id", "user_id", name="uq_group_member"),
    )


class UserRole(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    role_id = db.Column(db.Integer, db.ForeignKey("role.id"), nullable=False)

    __table_args__ = (db.UniqueConstraint("user_id", "role_id", name="uq_user_role"),)


class GroupRole(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    group_id = db.Column(db.Integer, db.ForeignKey("group.id"), nullable=False)
    role_id = db.Column(db.Integer, db.ForeignKey("role.id"), nullable=False)

    __table_args__ = (db.UniqueConstraint("group_id", "role_id", name="uq_group_role"),)


class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime, nullable=False, server_default=func.now())
    user_id = db.Column(db.Integer, nullable=True)
    username = db.Column(db.String(80), nullable=True)
    action = db.Column(db.String(50), nullable=False)
    entity_type = db.Column(db.String(50), nullable=False)
    entity_id = db.Column(db.String(80), nullable=True)
    success = db.Column(db.Boolean, nullable=False, default=True)
    ip_address = db.Column(db.String(45), nullable=True)
    details = db.Column(db.Text, nullable=True)


class SMTPConfig(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    host = db.Column(db.String(255))
    port = db.Column(db.Integer)
    encryption = db.Column(db.String(20), nullable=False, default="none")
    username = db.Column(db.String(255))
    password = db.Column(db.String(255))
    skip_auth = db.Column(db.Boolean, nullable=False, default=False)
    sender_email = db.Column(db.String(255))
    enabled = db.Column(db.Boolean, nullable=False, default=False)
    monthly_report_enabled = db.Column(db.Boolean, nullable=False, default=False)
    monthly_report_day = db.Column(db.Integer, nullable=False, default=1)
    low_stock_enabled = db.Column(db.Boolean, nullable=False, default=False)
    low_stock_threshold = db.Column(db.Integer, nullable=False, default=5)
    low_stock_frequency_days = db.Column(db.Integer, nullable=False, default=1)


class SMTPRecipient(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(255), unique=True, nullable=False)
    notify_create = db.Column(db.Boolean, nullable=False, default=False)
    notify_update = db.Column(db.Boolean, nullable=False, default=False)
    notify_delete = db.Column(db.Boolean, nullable=False, default=False)
    notify_bulk_delete = db.Column(db.Boolean, nullable=False, default=False)
    notify_monthly = db.Column(db.Boolean, nullable=False, default=False)
    notify_low_stock = db.Column(db.Boolean, nullable=False, default=False)


class RefreshToken(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, nullable=False, index=True)
    token_hash = db.Column(db.String(255), nullable=False, unique=True)
    issued_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, nullable=False)
    expires_at = db.Column(db.DateTime, nullable=False)
    revoked = db.Column(db.Boolean, nullable=False, default=False)


class AssetAssignmentHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    asset_type = db.Column(db.String(100), nullable=False)
    asset_id = db.Column(db.Integer, nullable=False)
    from_user = db.Column(db.String(255), nullable=True)
    to_user = db.Column(db.String(255), nullable=True)
    assigned_by = db.Column(db.String(80), nullable=True)
    assigned_by_id = db.Column(db.Integer, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, server_default=func.now())


class AssetComment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    asset_type = db.Column(db.String(100), nullable=False)
    asset_id = db.Column(db.Integer, nullable=False)
    body = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, nullable=False, server_default=func.now())
    user_id = db.Column(db.Integer, nullable=True)
    username = db.Column(db.String(80), nullable=True)

class BrandingConfig(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    company_name = db.Column(db.String(120), nullable=True)
    logo_filename = db.Column(db.String(255), nullable=True)


class Department(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(80), unique=True, nullable=False)


class LowStockState(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    asset_key = db.Column(db.String(80), nullable=False)
    entity_id = db.Column(db.String(80), nullable=True)
    last_notified_at = db.Column(db.DateTime, nullable=True)

    __table_args__ = (
        db.UniqueConstraint("asset_key", "entity_id", name="uq_low_stock_state"),
    )


class AssetType(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(50), unique=True, nullable=False)
    label = db.Column(db.String(80), nullable=False)
    fields = db.relationship("AssetField", backref="asset_type", cascade="all, delete-orphan")
    items = db.relationship("AssetItem", backref="asset_type", cascade="all, delete-orphan")


class BuiltinAssetTypeSetting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(50), unique=True, nullable=False)
    label = db.Column(db.String(80), nullable=False)
    fields = db.relationship(
        "BuiltinAssetFieldSetting", backref="asset_type", cascade="all, delete-orphan"
    )


class BuiltinAssetFieldSetting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    asset_key = db.Column(
        db.String(50), db.ForeignKey("builtin_asset_type_setting.key"), nullable=False
    )
    field_name = db.Column(db.String(50), nullable=False)
    label = db.Column(db.String(80), nullable=False)
    options = db.Column(db.Text, nullable=True)


class AssetField(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    asset_type_id = db.Column(db.Integer, db.ForeignKey("asset_type.id"), nullable=False)
    name = db.Column(db.String(50), nullable=False)
    label = db.Column(db.String(80), nullable=False)
    field_type = db.Column(db.String(20), nullable=False, default="text")
    options = db.Column(db.JSON, nullable=False, default=list)


class AssetItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    asset_type_id = db.Column(db.Integer, db.ForeignKey("asset_type.id"), nullable=False)
    data = db.Column(db.JSON, nullable=False, default=dict)


class Laptop(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    screen_size = db.Column(db.String(50), nullable=False)
    ram = db.Column(db.String(50), nullable=False)
    processor = db.Column(db.String(50), nullable=False)
    hard_disk = db.Column(db.String(50), nullable=False)
    asset_tag = db.Column(db.String(80), nullable=True)
    model = db.Column(db.String(80), nullable=False)
    vendor = db.Column(db.String(80), nullable=False)
    dept = db.Column(db.String(80), nullable=True)
    assigned_to = db.Column(db.String(80), nullable=True)
    status = db.Column(db.String(20), nullable=True)


class Computer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ram = db.Column(db.String(50), nullable=False)
    processor = db.Column(db.String(50), nullable=False)
    hard_disk = db.Column(db.String(50), nullable=False)
    asset_tag = db.Column(db.String(80), nullable=True)
    model = db.Column(db.String(80), nullable=False)
    vendor = db.Column(db.String(80), nullable=False)
    dept = db.Column(db.String(80), nullable=True)
    assigned_to = db.Column(db.String(80), nullable=True)
    status = db.Column(db.String(20), nullable=True)


class Screen(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    size = db.Column(db.String(50), nullable=False)
    asset_tag = db.Column(db.String(80), nullable=True)
    model = db.Column(db.String(80), nullable=False)
    vendor = db.Column(db.String(80), nullable=False)
    dept = db.Column(db.String(80), nullable=True)
    assigned_to = db.Column(db.String(80), nullable=True)
    status = db.Column(db.String(20), nullable=True)


class Keyboard(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    wired = db.Column(db.Boolean, nullable=False, default=False)
    wireless = db.Column(db.Boolean, nullable=False, default=False)
    model = db.Column(db.String(80), nullable=False)
    dept = db.Column(db.String(80), nullable=True)
    assigned_to = db.Column(db.String(80), nullable=True)
    status = db.Column(db.String(20), nullable=True)


class Mouse(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    wired = db.Column(db.Boolean, nullable=False, default=False)
    wireless = db.Column(db.Boolean, nullable=False, default=False)
    model = db.Column(db.String(80), nullable=False)
    dept = db.Column(db.String(80), nullable=True)
    assigned_to = db.Column(db.String(80), nullable=True)
    status = db.Column(db.String(20), nullable=True)


class Headset(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    model = db.Column(db.String(80), nullable=False)
    dept = db.Column(db.String(80), nullable=True)
    assigned_to = db.Column(db.String(80), nullable=True)
    status = db.Column(db.String(20), nullable=True)


class Ram(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ram_type = db.Column(db.String(20), nullable=True)
    size = db.Column(db.String(50), nullable=False)
    speed = db.Column(db.String(50), nullable=False)
    vendor = db.Column(db.String(80), nullable=False)
    total_quantity = db.Column(db.Integer, nullable=False, default=0)
    assigned_quantity = db.Column(db.Integer, nullable=False, default=0)
    dept = db.Column(db.String(80), nullable=True)
    assigned_to = db.Column(db.String(80), nullable=True)
    status = db.Column(db.String(20), nullable=True)


ASSET_DEFS = {
    "laptops": {
        "label": "Laptop",
        "model": Laptop,
        "bulk_add": False,
        "fields": [
            ("asset_tag", "Asset Tag", "text"),
            ("vendor", "Vendor", "text"),
            ("model", "Model", "text"),
            ("processor", "Processor", "text"),
            ("ram", "RAM", "text"),
            ("hard_disk", "Hard Disk", "text"),
            ("screen_size", "Screen size", "text"),
            ("dept", "Dept", "text"),
            ("assigned_to", "User", "text"),
            ("status", "Status", "select"),
        ],
        "field_options": {"status": STATUS_OPTIONS},
    },
    "computers": {
        "label": "Computer",
        "model": Computer,
        "bulk_add": False,
        "fields": [
            ("asset_tag", "Asset Tag", "text"),
            ("vendor", "Vendor", "text"),
            ("model", "Model", "text"),
            ("processor", "Processor", "text"),
            ("ram", "RAM", "text"),
            ("hard_disk", "Hard Disk", "text"),
            ("dept", "Dept", "text"),
            ("assigned_to", "User", "text"),
            ("status", "Status", "select"),
        ],
        "field_options": {"status": STATUS_OPTIONS},
    },
    "screens": {
        "label": "Screen",
        "model": Screen,
        "bulk_add": False,
        "fields": [
            ("asset_tag", "Asset Tag", "text"),
            ("vendor", "Vendor", "text"),
            ("model", "Model", "text"),
            ("size", "Screen size", "text"),
            ("dept", "Dept", "text"),
            ("assigned_to", "User", "text"),
            ("status", "Status", "select"),
        ],
        "field_options": {"status": STATUS_OPTIONS},
    },
    "keyboards": {
        "label": "Keyboard",
        "model": Keyboard,
        "bulk_add": True,
        "fields": [
            ("wired", "Wired", "checkbox"),
            ("wireless", "Wireless", "checkbox"),
            ("model", "Model", "text"),
            ("dept", "Dept", "text"),
            ("assigned_to", "User", "text"),
            ("status", "Status", "select"),
        ],
        "field_options": {"status": STATUS_OPTIONS},
    },
    "mice": {
        "label": "Mouse",
        "model": Mouse,
        "bulk_add": True,
        "fields": [
            ("connection", "Connection", "select"),
            ("model", "Model", "text"),
            ("dept", "Dept", "text"),
            ("assigned_to", "User", "text"),
            ("status", "Status", "select"),
        ],
        "field_options": {
            "status": STATUS_OPTIONS,
            "connection": ["Wired", "Wireless"],
        },
    },
    "headsets": {
        "label": "Headset",
        "model": Headset,
        "bulk_add": True,
        "fields": [
            ("model", "Model", "text"),
            ("dept", "Dept", "text"),
            ("assigned_to", "User", "text"),
            ("status", "Status", "select"),
        ],
        "field_options": {"status": STATUS_OPTIONS},
    },
    "ram": {
        "label": "RAM",
        "model": Ram,
        "bulk_add": True,
        "fields": [
            ("ram_type", "RAM Type", "select"),
            ("size", "Size", "text"),
            ("speed", "Speed", "text"),
            ("vendor", "Vendor", "text"),
            ("dept", "Dept", "text"),
            ("assigned_to", "User", "text"),
            ("status", "Status", "select"),
        ],
        "field_options": {"status": STATUS_OPTIONS, "ram_type": ["DDR3", "DDR4", "DDR5"]},
    },
}


def slugify_key(value):
    cleaned = re.sub(r"[^a-zA-Z0-9_]+", "_", value.strip().lower())
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    return cleaned or "asset"


def normalize_search(value):
    return (value or "").strip().lower()


def matches_query(query, *values):
    if not query:
        return True
    for value in values:
        if value is None:
            continue
        if query in str(value).lower():
            return True
    return False


def asset_audit_details(asset_type, item):
    asset_tag = getattr(item, "asset_tag", None)
    model = getattr(item, "model", None)
    parts = [f"type={asset_type}"]
    if asset_tag:
        parts.append(f"asset_tag={asset_tag}")
    if model:
        parts.append(f"model={model}")
    return " ".join(parts)


def format_changes(old_values, new_values):
    changes = []
    keys = set(old_values.keys()) | set(new_values.keys())
    for key in sorted(keys):
        old = old_values.get(key)
        new = new_values.get(key)
        if old != new:
            changes.append(f"{key}: {old} -> {new}")
    return "; ".join(changes)


def get_smtp_config():
    return SMTPConfig.query.first()


def send_smtp_notification(action, entity_type, entity_id, success, details):
    if not success:
        return
    config = get_smtp_config()
    if not config or not config.enabled:
        return
    action_map = {
        "create": "notify_create",
        "update": "notify_update",
        "delete": "notify_delete",
        "bulk_delete": "notify_bulk_delete",
    }
    notify_field = action_map.get(action)
    if not notify_field:
        return
    recipients = SMTPRecipient.query.filter_by(**{notify_field: True}).all()
    if not recipients:
        return
    subject = f"[{get_branding_name()}] {action.upper()} {entity_type}"
    user = get_current_user()
    if not user and request:
        user = getattr(request, "api_user", None)
    user_name = user.username if user else "-"
    body_lines = [
        f"Action: {action}",
        f"Entity: {entity_type}",
        f"Entity ID: {entity_id or '-'}",
        f"User: {user_name}",
        f"Success: {success}",
        f"Details: {details or '-'}",
    ]
    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = config.sender_email or config.username or "no-reply@inventory.local"
    message["To"] = ", ".join([recipient.email for recipient in recipients])
    message.set_content("\n".join(body_lines))
    meta_lines = [
        ("Action", action),
        ("Entity", entity_type),
        ("Entity ID", entity_id or "-"),
        ("User", user_name),
        ("Status", "Success" if success else "Failed"),
        ("Timestamp", _format_timestamp(datetime.datetime.utcnow())),
    ]
    parsed = _parse_audit_details(details or "")
    def display_value(value):
        if value is None:
            return "-"
        if isinstance(value, str) and value.strip().lower() == "free":
            return "-"
        return value

    detail_rows = []
    for label, value in (
        ("Type", parsed.get("type")),
        ("Asset Tag", parsed.get("asset_tag")),
        ("Model", parsed.get("model")),
    ):
        if value:
            detail_rows.append((label, display_value(value)))
    if parsed.get("asset_tags"):
        detail_rows.append(("Asset Tags", ", ".join(parsed["asset_tags"])))
    if action == "bulk_delete":
        deleted_count = 0
        if parsed.get("asset_tags"):
            deleted_count = len(parsed["asset_tags"])
        elif parsed.get("ids"):
            deleted_count = len(parsed["ids"])
        if deleted_count:
            detail_rows.append(("Deleted Assets", str(deleted_count)))
    if parsed.get("changes"):
        for field_name, change in parsed["changes"].items():
            old_value = display_value(change.get("old", "-"))
            new_value = display_value(change.get("new", "-"))
            detail_rows.append(
                (
                    f"{field_name}",
                    f"{old_value} → {new_value}",
                )
            )
    if not detail_rows:
        detail_rows.append(("Details", display_value(details or "-")))
    detail_html = "".join(
        "<tr>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\">{label}</td>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\">{value}</td>"
        "</tr>".format(
            label=html.escape(str(label)),
            value=html.escape(str(value)),
        )
        for label, value in detail_rows
    )
    detail_html = (
        "<div style=\"color:#7dd3fc; font-weight:600; margin-bottom:8px;\">Details</div>"
        "<table style=\"width:100%; border-collapse:collapse; font-size:14px;\">"
        "<tbody>"
        f"{detail_html}"
        "</tbody>"
        "</table>"
    )
    html_body = _render_report_html(
        "Asset Notification",
        f"{action.upper()} event",
        meta_lines,
        detail_html,
    )
    message.add_alternative(html_body, subtype="html")
    try:
        if config.encryption == "ssl":
            server = smtplib.SMTP_SSL(config.host, config.port or 465, timeout=10)
        else:
            server = smtplib.SMTP(config.host, config.port or 25, timeout=10)
        with server:
            if config.encryption == "starttls":
                server.starttls()
            if not config.skip_auth and config.username and config.password:
                server.login(config.username, config.password)
            server.send_message(message)
    except Exception:
        pass


def send_email(subject, body, recipients, html_body=None):
    config = get_smtp_config()
    if not config or not config.enabled:
        return False
    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = config.sender_email or config.username or "no-reply@inventory.local"
    message["To"] = ", ".join(recipients)
    message.set_content(body)
    if html_body:
        message.add_alternative(html_body, subtype="html")
    try:
        if config.encryption == "ssl":
            server = smtplib.SMTP_SSL(config.host, config.port or 465, timeout=10)
        else:
            server = smtplib.SMTP(config.host, config.port or 25, timeout=10)
        with server:
            if config.encryption == "starttls":
                server.starttls()
            if not config.skip_auth and config.username and config.password:
                server.login(config.username, config.password)
            server.send_message(message)
        return True
    except Exception:
        return False


def send_assignment_email(username, asset_label, specs):
    config = get_smtp_config()
    if not config or not config.enabled:
        return False
    email = resolve_user_email(username)
    if not email:
        log_audit(
            "notify_failed",
            "assignment",
            success=False,
            details=f"user={username} asset={asset_label}",
        )
        return False
    subject = f"[{get_branding_name()}] Asset assigned: {asset_label}"
    lines = [
        f"Hello {username},",
        "",
        f"A {asset_label} has been assigned to you.",
        "Details:",
    ]
    for label, value in specs:
        lines.append(f"- {label}: {value}")
    lines.append("")
    lines.append("If this is unexpected or incorrect, please contact IT.")
    body_rows = "".join(
        "<tr>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\">{label}</td>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\">{value}</td>"
        "</tr>".format(
            label=html.escape(str(label)),
            value=html.escape(str(value)),
        )
        for label, value in specs
    )
    intro_html = (
        f"<div style=\"font-size:16px; margin-bottom:8px;\">Hello {html.escape(username)},</div>"
        f"<div style=\"margin-bottom:14px; color:#cbd5f5;\">"
        f"A {html.escape(asset_label)} has been assigned to you.</div>"
    )
    body_html = (
        f"{intro_html}"
        "<div style=\"color:#7dd3fc; font-weight:600; margin-bottom:8px;\">Asset details</div>"
        "<table style=\"width:100%; border-collapse:collapse; font-size:14px;\">"
        "<tbody>"
        f"{body_rows}"
        "</tbody>"
        "</table>"
        "<div style=\"margin-top:16px; color:#cbd5f5;\">"
        "If this is unexpected or incorrect, please contact IT."
        "</div>"
    )
    meta_lines = [
        ("User", username),
        ("Asset", asset_label),
        ("Assigned by", get_current_user().username if get_current_user() else "-"),
        ("Timestamp", _format_timestamp(datetime.datetime.utcnow())),
    ]
    html_body = _render_report_html(
        "Asset Assignment",
        "Your assigned asset summary",
        meta_lines,
        body_html,
    )
    return send_email(subject, "\n".join(lines), [email], html_body=html_body)


def build_assignment_specs(definition, item):
    fields = definition.get("fields", [])
    label_map = {name: label for name, label, _ in fields}
    type_map = {name: field_type for name, _, field_type in fields}
    values = {}
    for field_name, _, _ in fields:
        if field_name in {"assigned_to", "screen_size"}:
            continue
        value = getattr(item, field_name, None)
        if type_map.get(field_name) == "checkbox":
            value = "Yes" if value else "No"
        if value not in (None, ""):
            values[field_name] = value
    priority = ["vendor", "model", "processor", "hard_disk", "ram", "asset_tag"]
    specs = []
    for field_name in priority:
        if field_name in values:
            specs.append((label_map.get(field_name, field_name.replace("_", " ").title()), values[field_name]))
    for field_name, value in values.items():
        if field_name in priority:
            continue
        specs.append((label_map.get(field_name, field_name.replace("_", " ").title()), value))
    return specs


def _format_timestamp(value):
    if not value:
        return "-"
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def _render_report_html(title, subtitle, meta_lines, body_html):
    meta_items = "".join(
        f"<div style=\"margin-bottom:6px;\"><span style=\"color:#7dd3fc;\">{html.escape(str(label))}:</span> {html.escape(str(value))}</div>"
        for label, value in meta_lines
    )
    brand_name = get_branding_name()
    return f"""\
<!DOCTYPE html>
<html>
  <head>
    <meta charset="utf-8">
    <title>{html.escape(title)}</title>
  </head>
  <body style="margin:0; padding:0; background:#0b1220; color:#e2e8f0; font-family:Segoe UI,Tahoma,Arial,sans-serif;">
    <div style="max-width:720px; margin:24px auto; background:linear-gradient(135deg,#0f172a 0%,#0b1220 50%,#0a1f2c 100%); border:1px solid #1f2a44; border-radius:16px; padding:24px; box-shadow:0 12px 32px rgba(0,0,0,0.45);">
      <div style="font-size:20px; letter-spacing:1px; text-transform:uppercase; color:#7dd3fc;">{html.escape(title)}</div>
      <div style="font-size:14px; color:#94a3b8; margin-top:4px;">{html.escape(subtitle)}</div>
      <div style="margin-top:16px; padding:12px; border-radius:12px; background:rgba(15,23,42,0.7); border:1px solid #24324a;">
        {meta_items}
      </div>
      <div style="margin-top:18px;">
        {body_html}
      </div>
      <div style="margin-top:24px; font-size:12px; color:#64748b;">Generated by {html.escape(brand_name)}</div>
    </div>
  </body>
</html>
"""


def _parse_audit_details(details):
    parsed = {
        "type": None,
        "asset_tag": None,
        "asset_tags": [],
        "ids": [],
        "model": None,
        "changes": {},
    }
    if not details:
        return parsed
    parsed["type"] = re.search(r"\\btype=([^\\s]+)", details)
    parsed["type"] = parsed["type"].group(1) if parsed["type"] else None
    parsed["asset_tag"] = re.search(r"\\basset_tag=([^\\s]+)", details)
    parsed["asset_tag"] = parsed["asset_tag"].group(1) if parsed["asset_tag"] else None
    tags_match = re.search(r"\basset_tags=\[([^\]]*)\]", details)
    if tags_match:
        tags_raw = tags_match.group(1)
        parsed["asset_tags"] = [
            tag.strip().strip("'").strip('"')
            for tag in tags_raw.split(",")
            if tag.strip()
        ]
    ids_match = re.search(r"\bids=\[([^\]]*)\]", details)
    if ids_match:
        ids_raw = ids_match.group(1)
        parsed["ids"] = [
            part.strip()
            for part in ids_raw.split(",")
            if part.strip()
        ]
    parsed["model"] = re.search(r"\\bmodel=([^\\s]+)", details)
    parsed["model"] = parsed["model"].group(1) if parsed["model"] else None
    _, _, changes_blob = details.partition(" changes=")
    if changes_blob:
        for part in changes_blob.split(";"):
            part = part.strip()
            if not part or ":" not in part or "->" not in part:
                continue
            field, rest = part.split(":", 1)
            old_new = rest.split("->")
            if len(old_new) != 2:
                continue
            parsed["changes"][field.strip()] = {
                "old": old_new[0].strip(),
                "new": old_new[1].strip(),
            }
    return parsed


def _get_stock_snapshot():
    rows = []
    for asset_key, definition in ASSET_DEFS.items():
        label = definition["label"]
        model = definition["model"]
        if definition.get("consumable"):
            items = model.query.all()
            total = sum(item.total_quantity or 0 for item in items)
            assigned = sum(item.assigned_quantity or 0 for item in items)
            available = max(total - assigned, 0)
        else:
            total = model.query.count()
            available = is_free_filter(model).count()
            assigned = max(total - available, 0)
        rows.append(
            {
                "label": label,
                "total": total,
                "assigned": assigned,
                "available": available,
            }
        )
    for asset_type in get_custom_asset_types():
        fields = AssetField.query.filter_by(asset_type_id=asset_type.id).all()
        assigned_fields, _ = get_custom_special_fields(fields)
        items = AssetItem.query.filter_by(asset_type_id=asset_type.id).all()
        total = len(items)
        available = 0
        assigned = 0
        if assigned_fields:
            for item in items:
                assigned_to = None
                for field_name in assigned_fields:
                    assigned_to = (item.data or {}).get(field_name)
                    if assigned_to is not None:
                        break
                assigned_text = str(assigned_to or "").strip().lower()
                if assigned_text in {"", "free"}:
                    available += 1
                else:
                    assigned += 1
        else:
            available = total
        rows.append(
            {
                "label": asset_type.label,
                "total": total,
                "assigned": assigned,
                "available": available,
            }
        )
    return rows


def get_low_stock_items(threshold):
    low_stock = []
    for asset_key, definition in ASSET_DEFS.items():
        if definition.get("consumable"):
            items = definition["model"].query.all()
            for item in items:
                available = max((item.total_quantity or 0) - (item.assigned_quantity or 0), 0)
                if available < threshold:
                    low_stock.append(
                        {
                            "asset_key": asset_key,
                            "entity_id": str(item.id),
                            "label": definition["label"],
                            "available": available,
                        }
                    )
        else:
            available = is_free_filter(definition["model"]).count()
            if available < threshold:
                low_stock.append(
                    {
                        "asset_key": asset_key,
                        "entity_id": None,
                        "label": definition["label"],
                        "available": available,
                    }
                )
    for asset_type in get_custom_asset_types():
        fields = AssetField.query.filter_by(asset_type_id=asset_type.id).all()
        assigned_fields, _ = get_custom_special_fields(fields)
        if not assigned_fields:
            continue
        items = AssetItem.query.filter_by(asset_type_id=asset_type.id).all()
        available = 0
        for item in items:
            assigned_to = None
            for field_name in assigned_fields:
                assigned_to = (item.data or {}).get(field_name)
                if assigned_to is not None:
                    break
            assigned_text = str(assigned_to or "").strip().lower()
            if assigned_text in {"", "free"}:
                available += 1
        if available < threshold:
            low_stock.append(
                {
                    "asset_key": f"custom:{asset_type.key}",
                    "entity_id": None,
                    "label": asset_type.label,
                    "available": available,
                }
            )
    return low_stock


def send_low_stock_report(force=False):
    config = get_smtp_config()
    if not config or not config.enabled or not config.low_stock_enabled:
        return False
    threshold = max(config.low_stock_threshold or 1, 1)
    frequency_days = max(config.low_stock_frequency_days or 1, 1)
    items = get_low_stock_items(threshold)
    if not items:
        return False
    recipients = [recipient.email for recipient in SMTPRecipient.query.filter_by(notify_low_stock=True).all()]
    if not recipients:
        return False
    lines = ["Low stock items:"]
    send_items = []
    for item in items:
        key = item["asset_key"]
        entity_id = item["entity_id"]
        state = LowStockState.query.filter_by(asset_key=key, entity_id=entity_id).first()
        if state and state.last_notified_at and not force:
            delta = (func.julianday(func.now()) - func.julianday(state.last_notified_at))
            if delta is not None and delta < frequency_days:
                continue
        send_items.append(item)
        lines.append(f"- {item['label']} (available: {item['available']})")
        if not state:
            state = LowStockState(asset_key=key, entity_id=entity_id)
            db.session.add(state)
        state.last_notified_at = func.now()
    if not send_items:
        db.session.commit()
        return False
    db.session.commit()
    subject = f"[{get_branding_name()}] Low stock report"
    body_rows = "".join(
        "<tr>"
        "<td style=\"padding:10px 12px; border-bottom:1px solid #1f2a44;\">{label}</td>"
        "<td style=\"padding:10px 12px; border-bottom:1px solid #1f2a44; text-align:right;\">{available}</td>"
        "</tr>".format(
            label=html.escape(str(item["label"])),
            available=html.escape(str(item["available"])),
        )
        for item in send_items
    )
    body_html = (
        "<table style=\"width:100%; border-collapse:collapse; font-size:14px;\">"
        "<thead>"
        "<tr>"
        "<th style=\"text-align:left; padding:10px 12px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">Asset</th>"
        "<th style=\"text-align:right; padding:10px 12px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">Available</th>"
        "</tr>"
        "</thead>"
        "<tbody>"
        f"{body_rows}"
        "</tbody>"
        "</table>"
    )
    meta_lines = [
        ("Threshold", str(threshold)),
        ("Frequency (days)", str(frequency_days)),
        ("Generated", _format_timestamp(datetime.datetime.utcnow())),
    ]
    html_body = _render_report_html(
        "Low Stock Report",
        "Assets below the configured threshold",
        meta_lines,
        body_html,
    )
    return send_email(subject, "\n".join(lines), recipients, html_body=html_body)


def send_monthly_report(force=False):
    config = get_smtp_config()
    if not config or not config.enabled or not config.monthly_report_enabled:
        return False
    recipients = [recipient.email for recipient in SMTPRecipient.query.filter_by(notify_monthly=True).all()]
    if not recipients:
        return False
    day = max(min(config.monthly_report_day or 1, 28), 1)
    if not force:
        today = datetime.datetime.utcnow().day
        if today != day:
            return False
    end = datetime.datetime.utcnow()
    start = end - datetime.timedelta(days=30)
    logs = AuditLog.query.filter(AuditLog.created_at >= start).order_by(AuditLog.created_at.desc()).limit(200).all()
    counts = {}
    created_assets = []
    assigned_assets = []
    for entry in logs:
        counts[entry.action] = counts.get(entry.action, 0) + 1
        if entry.entity_type == "asset" and entry.action == "create":
            details = _parse_audit_details(entry.details or "")
            created_assets.append(
                {
                    "time": entry.created_at,
                    "asset_type": details.get("type") or "-",
                    "asset_tag": details.get("asset_tag") or "-",
                    "user": entry.username or "-",
                }
            )
        if entry.entity_type == "asset" and entry.action == "update":
            details = _parse_audit_details(entry.details or "")
            change = details.get("changes", {}).get("assigned_to")
            if change:
                new_value = change.get("new", "")
                if str(new_value).strip().lower() not in {"", "free"}:
                    assigned_assets.append(
                        {
                            "time": entry.created_at,
                            "asset_type": details.get("type") or "-",
                            "asset_tag": details.get("asset_tag") or "-",
                            "assigned_to": new_value,
                            "user": entry.username or "-",
                        }
                    )
    lines = ["Monthly report (last 30 days):", ""]
    lines.append("Counts by action:")
    for action, count in sorted(counts.items()):
        lines.append(f"- {action}: {count}")
    lines.append("")
    lines.append(f"Assets added: {len(created_assets)}")
    lines.append(f"Assets assigned: {len(assigned_assets)}")
    lines.append("")
    lines.append("Recent activity:")
    for entry in logs:
        lines.append(f"- {entry.created_at} {entry.username or '-'} {entry.action} {entry.entity_type} {entry.entity_id or '-'}")
    subject = f"[{get_branding_name()}] Monthly report"
    counts_rows = "".join(
        "<tr>"
        "<td style=\"padding:10px 12px; border-bottom:1px solid #1f2a44;\">{action}</td>"
        "<td style=\"padding:10px 12px; border-bottom:1px solid #1f2a44; text-align:right;\">{count}</td>"
        "</tr>".format(
            action=html.escape(str(action)),
            count=html.escape(str(count)),
        )
        for action, count in sorted(counts.items())
    )
    if not counts_rows:
        counts_rows = (
            "<tr><td style=\"padding:10px 12px; border-bottom:1px solid #1f2a44;\" colspan=\"2\">No activity recorded.</td></tr>"
        )
    activity_rows = "".join(
        "<tr>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\">{when}</td>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\">{user}</td>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\">{action}</td>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\">{entity}</td>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\">{entity_id}</td>"
        "</tr>".format(
            when=html.escape(_format_timestamp(entry.created_at)),
            user=html.escape(entry.username or "-"),
            action=html.escape(entry.action),
            entity=html.escape(entry.entity_type),
            entity_id=html.escape(entry.entity_id or "-"),
        )
        for entry in logs[:50]
    )
    added_rows = "".join(
        "<tr>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\">{when}</td>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\">{asset_type}</td>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\">{asset_tag}</td>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\">{user}</td>"
        "</tr>".format(
            when=html.escape(_format_timestamp(item["time"])),
            asset_type=html.escape(str(item["asset_type"])),
            asset_tag=html.escape(str(item["asset_tag"])),
            user=html.escape(str(item["user"])),
        )
        for item in created_assets[:50]
    )
    if not added_rows:
        added_rows = (
            "<tr><td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\" colspan=\"4\">No assets added.</td></tr>"
        )
    assigned_rows = "".join(
        "<tr>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\">{when}</td>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\">{asset_type}</td>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\">{asset_tag}</td>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\">{assigned_to}</td>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\">{user}</td>"
        "</tr>".format(
            when=html.escape(_format_timestamp(item["time"])),
            asset_type=html.escape(str(item["asset_type"])),
            asset_tag=html.escape(str(item["asset_tag"])),
            assigned_to=html.escape(str(item["assigned_to"])),
            user=html.escape(str(item["user"])),
        )
        for item in assigned_assets[:50]
    )
    if not assigned_rows:
        assigned_rows = (
            "<tr><td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\" colspan=\"5\">No assignments recorded.</td></tr>"
        )
    stock_rows = "".join(
        "<tr>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44;\">{label}</td>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44; text-align:right;\">{total}</td>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44; text-align:right;\">{assigned}</td>"
        "<td style=\"padding:8px 10px; border-bottom:1px solid #1f2a44; text-align:right;\">{available}</td>"
        "</tr>".format(
            label=html.escape(str(item["label"])),
            total=html.escape(str(item["total"])),
            assigned=html.escape(str(item["assigned"])),
            available=html.escape(str(item["available"])),
        )
        for item in _get_stock_snapshot()
    )
    body_html = (
        "<div style=\"margin-bottom:18px;\">"
        "<div style=\"color:#7dd3fc; font-weight:600; margin-bottom:8px;\">Counts by action</div>"
        "<table style=\"width:100%; border-collapse:collapse; font-size:14px;\">"
        "<thead>"
        "<tr>"
        "<th style=\"text-align:left; padding:10px 12px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">Action</th>"
        "<th style=\"text-align:right; padding:10px 12px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">Count</th>"
        "</tr>"
        "</thead>"
        "<tbody>"
        f"{counts_rows}"
        "</tbody>"
        "</table>"
        "</div>"
        "<div style=\"margin-bottom:18px;\">"
        "<div style=\"color:#7dd3fc; font-weight:600; margin-bottom:8px;\">Assets added (last 30 days)</div>"
        "<table style=\"width:100%; border-collapse:collapse; font-size:13px;\">"
        "<thead>"
        "<tr>"
        "<th style=\"text-align:left; padding:8px 10px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">Time</th>"
        "<th style=\"text-align:left; padding:8px 10px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">Type</th>"
        "<th style=\"text-align:left; padding:8px 10px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">Asset Tag</th>"
        "<th style=\"text-align:left; padding:8px 10px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">User</th>"
        "</tr>"
        "</thead>"
        "<tbody>"
        f"{added_rows}"
        "</tbody>"
        "</table>"
        "</div>"
        "<div style=\"margin-bottom:18px;\">"
        "<div style=\"color:#7dd3fc; font-weight:600; margin-bottom:8px;\">Assignments (last 30 days)</div>"
        "<table style=\"width:100%; border-collapse:collapse; font-size:13px;\">"
        "<thead>"
        "<tr>"
        "<th style=\"text-align:left; padding:8px 10px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">Time</th>"
        "<th style=\"text-align:left; padding:8px 10px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">Type</th>"
        "<th style=\"text-align:left; padding:8px 10px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">Asset Tag</th>"
        "<th style=\"text-align:left; padding:8px 10px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">User</th>"
        "<th style=\"text-align:left; padding:8px 10px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">By</th>"
        "</tr>"
        "</thead>"
        "<tbody>"
        f"{assigned_rows}"
        "</tbody>"
        "</table>"
        "</div>"
        "<div style=\"margin-bottom:18px;\">"
        "<div style=\"color:#7dd3fc; font-weight:600; margin-bottom:8px;\">Current stock snapshot</div>"
        "<table style=\"width:100%; border-collapse:collapse; font-size:13px;\">"
        "<thead>"
        "<tr>"
        "<th style=\"text-align:left; padding:8px 10px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">Asset</th>"
        "<th style=\"text-align:right; padding:8px 10px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">Total</th>"
        "<th style=\"text-align:right; padding:8px 10px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">Assigned</th>"
        "<th style=\"text-align:right; padding:8px 10px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">Available</th>"
        "</tr>"
        "</thead>"
        "<tbody>"
        f"{stock_rows}"
        "</tbody>"
        "</table>"
        "</div>"
        "<div>"
        "<div style=\"color:#7dd3fc; font-weight:600; margin-bottom:8px;\">Recent activity</div>"
        "<table style=\"width:100%; border-collapse:collapse; font-size:13px;\">"
        "<thead>"
        "<tr>"
        "<th style=\"text-align:left; padding:8px 10px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">Time</th>"
        "<th style=\"text-align:left; padding:8px 10px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">User</th>"
        "<th style=\"text-align:left; padding:8px 10px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">Action</th>"
        "<th style=\"text-align:left; padding:8px 10px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">Entity</th>"
        "<th style=\"text-align:left; padding:8px 10px; border-bottom:1px solid #1f2a44; color:#7dd3fc;\">Entity ID</th>"
        "</tr>"
        "</thead>"
        "<tbody>"
        f"{activity_rows}"
        "</tbody>"
        "</table>"
        "</div>"
    )
    meta_lines = [
        ("Report window", f"{_format_timestamp(start)} to {_format_timestamp(end)}"),
        ("Total events", str(len(logs))),
        ("Assets added", str(len(created_assets))),
        ("Assets assigned", str(len(assigned_assets))),
        ("Generated", _format_timestamp(datetime.datetime.utcnow())),
    ]
    html_body = _render_report_html(
        "Monthly Activity Report",
        "Asset actions and changes in the last 30 days",
        meta_lines,
        body_html,
    )
    return send_email(subject, "\n".join(lines), recipients, html_body=html_body)


def _parse_bool(value, default=False):
    if value is None:
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def _parse_int(value, default):
    if value is None or value == "":
        return default
    try:
        return int(value)
    except ValueError:
        return default


def normalize_username(value):
    return (value or "").strip().lower()


def get_user_by_username_ci(username):
    if not username:
        return None
    return User.query.filter(func.lower(User.username) == username.lower()).first()


def _hash_token(token):
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def _jwt_encode(payload, expires_in):
    now = datetime.datetime.utcnow()
    data = dict(payload)
    data.update({"iat": now, "exp": now + datetime.timedelta(seconds=expires_in)})
    return jwt.encode(data, app.config["SECRET_KEY"], algorithm=JWT_ALGORITHM)


def _jwt_decode(token):
    return jwt.decode(token, app.config["SECRET_KEY"], algorithms=[JWT_ALGORITHM])


def _issue_tokens(user_id):
    access_token = _jwt_encode({"sub": str(user_id), "type": "access"}, JWT_ACCESS_SECONDS)
    refresh_token = _jwt_encode({"sub": str(user_id), "type": "refresh"}, JWT_REFRESH_SECONDS)
    refresh_entry = RefreshToken(
        user_id=user_id,
        token_hash=_hash_token(refresh_token),
        expires_at=datetime.datetime.utcnow() + datetime.timedelta(seconds=JWT_REFRESH_SECONDS),
    )
    db.session.add(refresh_entry)
    db.session.commit()
    return access_token, refresh_token


def _rotate_refresh_token(user_id, token_value):
    token_hash = _hash_token(token_value)
    entry = RefreshToken.query.filter_by(token_hash=token_hash, revoked=False).first()
    if not entry:
        return None, None
    if entry.expires_at < datetime.datetime.utcnow():
        entry.revoked = True
        db.session.commit()
        return None, None
    entry.revoked = True
    db.session.commit()
    return _issue_tokens(user_id)


def log_audit(action, entity_type, entity_id=None, success=True, details=None):
    user = get_current_user()
    if not user and request:
        user = getattr(request, "api_user", None)
    username = user.username if user else None
    user_id = user.id if user else None
    ip_address = request.remote_addr if request else None
    detail_text = None
    if details is not None:
        detail_text = str(details)
    try:
        app.logger.info(
            "audit action=%s entity=%s entity_id=%s user=%s ip=%s success=%s details=%s",
            action,
            entity_type,
            entity_id or "-",
            username or "-",
            ip_address or "-",
            "yes" if success else "no",
            detail_text or "-",
        )
        db.session.add(
            AuditLog(
                user_id=user_id,
                username=username,
                action=action,
                entity_type=entity_type,
                entity_id=str(entity_id) if entity_id is not None else None,
                success=success,
                ip_address=ip_address,
                details=detail_text,
            )
        )
        db.session.commit()
    except Exception:
        db.session.rollback()
        return
    send_smtp_notification(action, entity_type, entity_id, success, detail_text)


def get_ldap_config_from_db():
    config = LdapConfig.query.first()
    if not config:
        return None
    return {
        "server": (config.server or "").strip(),
        "base_dn": (config.base_dn or "").strip(),
        "bind_dn": (config.bind_dn or "").strip(),
        "bind_password": config.bind_password or "",
        "user_filter": config.user_filter or "(uid={username})",
        "list_filter": config.list_filter or "(&(objectClass=user)(!(objectClass=computer))(sAMAccountName=*))",
        "user_attribute": config.user_attribute or "uid",
        "email_attribute": config.email_attribute or "mail",
        "user_dn_template": (config.user_dn_template or "").strip(),
        "use_ssl": bool(config.use_ssl),
        "start_tls": bool(config.start_tls),
        "cache_seconds": config.cache_seconds,
        "list_limit": config.list_limit,
        "default_role": (config.default_role or "unassigned").strip() or "unassigned",
        "admin_users": (config.admin_users or "").strip(),
        "group_filter": config.group_filter or "(&(objectClass=group)(cn=*))",
        "group_attribute": config.group_attribute or "cn",
        "group_member_attribute": config.group_member_attribute or "member",
    }


def _ldap_config():
    config = {
        "server": os.environ.get("LDAP_SERVER", "").strip(),
        "base_dn": os.environ.get("LDAP_BASE_DN", "").strip(),
        "bind_dn": os.environ.get("LDAP_BIND_DN", "").strip(),
        "bind_password": os.environ.get("LDAP_BIND_PASSWORD", ""),
        "user_filter": os.environ.get("LDAP_USER_FILTER", "(uid={username})"),
        "list_filter": os.environ.get(
            "LDAP_USER_LIST_FILTER",
            "(&(objectClass=user)(!(objectClass=computer))(sAMAccountName=*))",
        ),
        "user_attribute": os.environ.get("LDAP_USER_ATTRIBUTE", "uid"),
        "email_attribute": os.environ.get("LDAP_EMAIL_ATTRIBUTE", "mail"),
        "user_dn_template": os.environ.get("LDAP_USER_DN_TEMPLATE", "").strip(),
        "use_ssl": _parse_bool(os.environ.get("LDAP_USE_SSL")),
        "start_tls": _parse_bool(os.environ.get("LDAP_START_TLS")),
        "cache_seconds": _parse_int(os.environ.get("LDAP_CACHE_SECONDS"), 300),
        "list_limit": _parse_int(os.environ.get("LDAP_USER_LIST_LIMIT"), 500),
        "default_role": os.environ.get("LDAP_DEFAULT_ROLE", "unassigned").strip()
        or "unassigned",
        "admin_users": os.environ.get("LDAP_ADMIN_USERS", "").strip(),
        "group_filter": os.environ.get(
            "LDAP_GROUP_LIST_FILTER", "(&(objectClass=group)(cn=*))"
        ),
        "group_attribute": os.environ.get("LDAP_GROUP_ATTRIBUTE", "cn"),
        "group_member_attribute": os.environ.get("LDAP_GROUP_MEMBER_ATTRIBUTE", "member"),
    }
    db_config = get_ldap_config_from_db()
    if db_config:
        for key, value in db_config.items():
            if value is not None:
                config[key] = value
    if config["cache_seconds"] is None:
        config["cache_seconds"] = 300
    if config["list_limit"] is None:
        config["list_limit"] = 500
    if not config["default_role"]:
        config["default_role"] = "unassigned"
    return config


def ldap_enabled():
    config = _ldap_config()
    return bool(config["server"] and config["base_dn"])


def _resolve_ldap_server(config):
    raw = (config.get("server") or "").strip()
    use_ssl = bool(config.get("use_ssl"))
    host = raw
    port = None
    if raw.startswith("ldap://") or raw.startswith("ldaps://"):
        parsed = urlparse(raw)
        host = parsed.hostname or ""
        port = parsed.port
        if parsed.scheme == "ldaps":
            use_ssl = True
    else:
        if ":" in raw:
            host, port_text = raw.rsplit(":", 1)
            try:
                port = int(port_text)
            except ValueError:
                port = None
    return host, port, use_ssl


def _ldap_connect(bind_dn=None, bind_password=None, config=None):
    config = config or _ldap_config()
    host, port, use_ssl = _resolve_ldap_server(config)
    if not host:
        return None
    tls_config = Tls(validate=ssl.CERT_NONE) if use_ssl else None
    server = Server(host, port=port, use_ssl=use_ssl, get_info=None, tls=tls_config)
    conn = Connection(server, user=bind_dn, password=bind_password, auto_bind=False)
    opened = conn.open()
    if opened is False or (opened is None and conn.closed):
        return None
    if config["start_tls"]:
        if not conn.start_tls():
            conn.unbind()
            return None
    if not conn.bind():
        conn.unbind()
        return None
    return conn


def _ldap_service_connection(config=None):
    config = config or _ldap_config()
    bind_dn = config["bind_dn"] or None
    bind_password = config["bind_password"] or None
    return _ldap_connect(bind_dn=bind_dn, bind_password=bind_password, config=config)


def _ldap_is_computer_entry(entry, user_attribute):
    try:
        classes = entry["objectClass"].value
    except (KeyError, AttributeError):
        classes = []
    if isinstance(classes, str):
        classes = [classes]
    if any(str(value).lower() == "computer" for value in classes or []):
        return True
    try:
        name_value = entry[user_attribute].value
    except (KeyError, AttributeError):
        return False
    if isinstance(name_value, list):
        name_value = next((item for item in name_value if item), "")
    return str(name_value or "").endswith("$")


def ldap_authenticate(username, password):
    if not ldap_enabled() or not password:
        return False
    config = _ldap_config()
    username_norm = normalize_username(username)
    if config["user_dn_template"]:
        user_dn = config["user_dn_template"].format(username=username_norm)
        return _ldap_bind(user_dn, password, config=config)
    conn = _ldap_service_connection(config=config)
    if not conn:
        return False
    search_filter = config["user_filter"].format(
        username=escape_filter_chars(username_norm)
    )
    conn.search(
        config["base_dn"],
        search_filter,
        search_scope=SUBTREE,
        attributes=["dn"],
        size_limit=1,
    )
    if not conn.entries:
        conn.unbind()
        return False
    user_dn = conn.entries[0].entry_dn
    conn.unbind()
    return _ldap_bind(user_dn, password, config=config)


def _ldap_bind(user_dn, password, config=None):
    conn = _ldap_connect(bind_dn=user_dn, bind_password=password, config=config)
    if not conn:
        return False
    conn.unbind()
    return True


def get_ldap_users(force_refresh=False):
    if not ldap_enabled():
        return []
    config = _ldap_config()
    cache_seconds = max(config["cache_seconds"], 0)
    now = time.time()
    if not force_refresh and now - _LDAP_USER_CACHE["timestamp"] < cache_seconds:
        return list(_LDAP_USER_CACHE["users"])
    conn = _ldap_service_connection(config=config)
    if not conn:
        _LDAP_USER_CACHE["users"] = []
        _LDAP_USER_CACHE["timestamp"] = now
        log_audit("sync_failed", "ldap_users", success=False, details="Bind failed")
        return []
    conn.search(
        config["base_dn"],
        config["list_filter"],
        search_scope=SUBTREE,
        attributes=[config["user_attribute"], "objectClass"],
        size_limit=config["list_limit"],
    )
    usernames = set()
    for entry in conn.entries:
        if _ldap_is_computer_entry(entry, config["user_attribute"]):
            continue
        try:
            value = entry[config["user_attribute"]].value
        except (KeyError, AttributeError):
            continue
        if isinstance(value, list):
            usernames.update(item for item in value if item)
        elif value:
            usernames.add(value)
    conn.unbind()
    users_sorted = sorted((normalize_username(item) for item in usernames if item), key=str)
    _LDAP_USER_CACHE["users"] = users_sorted
    _LDAP_USER_CACHE["timestamp"] = now
    return list(users_sorted)


def get_ldap_user_records(force_refresh=False):
    if not ldap_enabled():
        return []
    config = _ldap_config()
    cache_seconds = max(config["cache_seconds"], 0)
    now = time.time()
    if not force_refresh and now - _LDAP_USER_RECORDS_CACHE["timestamp"] < cache_seconds:
        return list(_LDAP_USER_RECORDS_CACHE["records"])
    conn = _ldap_service_connection(config=config)
    if not conn:
        _LDAP_USER_RECORDS_CACHE["records"] = []
        _LDAP_USER_RECORDS_CACHE["timestamp"] = now
        log_audit("sync_failed", "ldap_users", success=False, details="Bind failed")
        return []
    email_attr = config.get("email_attribute") or "mail"
    conn.search(
        config["base_dn"],
        config["list_filter"],
        search_scope=SUBTREE,
        attributes=[config["user_attribute"], email_attr, "objectClass"],
        size_limit=config["list_limit"],
    )
    records = []
    for entry in conn.entries:
        if _ldap_is_computer_entry(entry, config["user_attribute"]):
            continue
        try:
            username_value = entry[config["user_attribute"]].value
        except (KeyError, AttributeError):
            continue
        if isinstance(username_value, list):
            username_value = next((item for item in username_value if item), None)
        if not username_value:
            continue
        email_value = None
        try:
            email_value = entry[email_attr].value
        except (KeyError, AttributeError):
            email_value = None
        if isinstance(email_value, list):
            email_value = next((item for item in email_value if item), None)
        records.append(
            {
                "username": normalize_username(str(username_value)),
                "email": (email_value or "").strip() or None,
            }
        )
    conn.unbind()
    records_sorted = sorted(records, key=lambda item: str(item["username"]))
    _LDAP_USER_RECORDS_CACHE["records"] = records_sorted
    _LDAP_USER_RECORDS_CACHE["timestamp"] = now
    return list(records_sorted)


def get_ldap_user_display_list(force_refresh=False):
    try:
        records = get_ldap_user_records(force_refresh=force_refresh)
    except Exception as exc:
        app.logger.warning("LDAP user list unavailable: %s", exc)
        return []
    display = []
    for record in records:
        value = record.get("email") or record.get("username")
        if value:
            display.append(value)
    return sorted(set(display), key=str)


def get_dept_options_cached():
    now = time.time()
    if now - _DEPT_CACHE["timestamp"] < 30:
        return list(_DEPT_CACHE["items"])
    items = [dept.name for dept in Department.query.order_by(Department.name.asc()).all()]
    _DEPT_CACHE["items"] = items
    _DEPT_CACHE["timestamp"] = now
    return list(items)


def get_ldap_groups(force_refresh=False):
    if not ldap_enabled():
        return []
    config = _ldap_config()
    cache_seconds = max(config["cache_seconds"], 0)
    now = time.time()
    if not force_refresh and now - _LDAP_GROUP_CACHE["timestamp"] < cache_seconds:
        return list(_LDAP_GROUP_CACHE["groups"])
    conn = _ldap_service_connection(config=config)
    if not conn:
        _LDAP_GROUP_CACHE["groups"] = []
        _LDAP_GROUP_CACHE["timestamp"] = now
        log_audit("sync_failed", "ldap_groups", success=False, details="Bind failed")
        return []
    conn.search(
        config["base_dn"],
        config["group_filter"],
        search_scope=SUBTREE,
        attributes=[config["group_attribute"], config["group_member_attribute"]],
        size_limit=config["list_limit"],
    )
    groups = []
    for entry in conn.entries:
        try:
            name_value = entry[config["group_attribute"]].value
        except (KeyError, AttributeError):
            continue
        if not name_value:
            continue
        members_value = None
        try:
            members_value = entry[config["group_member_attribute"]].value
        except (KeyError, AttributeError):
            members_value = None
        if isinstance(members_value, list):
            members = [item for item in members_value if item]
        elif members_value:
            members = [members_value]
        else:
            members = []
        groups.append({"name": name_value, "members": members})
    conn.unbind()
    groups_sorted = sorted(groups, key=lambda item: str(item["name"]).lower())
    _LDAP_GROUP_CACHE["groups"] = groups_sorted
    _LDAP_GROUP_CACHE["timestamp"] = now
    return list(groups_sorted)


def ensure_ldap_user(username):
    username_norm = normalize_username(username)
    user = get_user_by_username_ci(username_norm)
    if user:
        return user
    config = _ldap_config()
    admin_users = {
        item.strip().lower()
        for item in (config.get("admin_users") or "").split(",")
        if item.strip()
    }
    default_role = config.get("default_role") or "unassigned"
    role = "admin" if username.lower() in admin_users else default_role
    if not Role.query.filter_by(name=role).first():
        role = "reader"
    random_password = os.urandom(24).hex()
    email = ldap_lookup_user_email(username_norm, config=config)
    user = User(
        username=username_norm,
        password_hash=generate_password_hash(random_password),
        role=role,
        email=email,
    )
    db.session.add(user)
    db.session.commit()
    return user


def _ldap_config_from_form(form, existing=None):
    def get_text(name):
        return form.get(name, "").strip()

    def get_int_or_none(name):
        value = form.get(name, "").strip()
        if value == "":
            return None
        try:
            return int(value)
        except ValueError:
            return None

    bind_password = form.get("bind_password", "")
    if bind_password == "" and existing:
        bind_password = existing.bind_password or ""

    config = {
        "server": get_text("server"),
        "base_dn": get_text("base_dn"),
        "bind_dn": get_text("bind_dn"),
        "bind_password": bind_password,
        "user_filter": get_text("user_filter") or "(sAMAccountName={username})",
        "list_filter": get_text("list_filter") or "(&(objectClass=user)(!(objectClass=computer))(sAMAccountName=*))",
        "user_attribute": get_text("user_attribute") or "sAMAccountName",
        "email_attribute": get_text("email_attribute") or "mail",
        "user_dn_template": get_text("user_dn_template"),
        "use_ssl": "use_ssl" in form,
        "start_tls": "start_tls" in form,
        "cache_seconds": get_int_or_none("cache_seconds"),
        "list_limit": get_int_or_none("list_limit"),
        "default_role": get_text("default_role") or "unassigned",
        "admin_users": get_text("admin_users"),
    }
    config["group_filter"] = get_text("group_filter") or "(&(objectClass=group)(cn=*))"
    config["group_attribute"] = get_text("group_attribute") or "cn"
    config["group_member_attribute"] = (
        get_text("group_member_attribute") or "member"
    )
    return config


def _ldap_test_connection(config):
    if not config.get("server") or not config.get("base_dn"):
        return False, "LDAP server and base DN are required to test the connection."
    try:
        host, port, use_ssl = _resolve_ldap_server(config)
        if not host:
            return False, "Unable to connect to LDAP server."
        if port is None:
            port = 636 if use_ssl else 389
        tcp_ok = False
        tcp_error = ""
        try:
            sock = socket.create_connection((host, port), timeout=3)
            sock.close()
            tcp_ok = True
        except OSError as exc:
            tcp_error = f" Socket error: {exc}"
        tls_config = Tls(validate=ssl.CERT_NONE) if use_ssl else None
        server = Server(host, port=port, use_ssl=use_ssl, get_info=None, tls=tls_config)
        bind_dn = config.get("bind_dn") or None
        bind_password = config.get("bind_password") or None
        conn = Connection(server, user=bind_dn, password=bind_password, auto_bind=False)
        def get_message():
            result = conn.result or {}
            return result.get("message") or conn.last_error or "Unable to connect."
        opened = conn.open()
        if opened is False or (opened is None and conn.closed):
            message = get_message()
            hint = " TCP check OK; try StartTLS or LDAPS." if tcp_ok else ""
            return False, f"Unable to connect to LDAP server ({host}:{port}). {message}.{tcp_error}{hint}"
        if config["start_tls"]:
            if not conn.start_tls():
                message = get_message()
                conn.unbind()
                return False, f"StartTLS failed: {message}"
        if not conn.bind():
            message = get_message()
            conn.unbind()
            return False, f"Unable to bind to LDAP with the provided credentials. {message}"
        searched = conn.search(
            config["base_dn"],
            config["list_filter"],
            search_scope=SUBTREE,
            attributes=[config["user_attribute"], config.get("email_attribute") or "mail"],
            size_limit=1,
        )
        diagnostics = ""
        if conn.last_error:
            diagnostics = f" Last error: {conn.last_error}"
        conn.unbind()
        if not searched:
            return False, f"Bind succeeded but the test search failed.{diagnostics}"
        return True, f"LDAP connection successful.{diagnostics}"
    except Exception as exc:
        return False, f"LDAP test failed: {exc}"


def _ldap_form_values(config_row):
    if config_row:
        return {
            "server": config_row.server or "",
            "base_dn": config_row.base_dn or "",
            "bind_dn": config_row.bind_dn or "",
            "bind_password": "",
            "user_filter": config_row.user_filter or "(uid={username})",
            "list_filter": config_row.list_filter
            or "(&(objectClass=user)(!(objectClass=computer))(sAMAccountName=*))",
            "user_attribute": config_row.user_attribute or "uid",
            "email_attribute": config_row.email_attribute or "mail",
            "user_dn_template": config_row.user_dn_template or "",
            "use_ssl": bool(config_row.use_ssl),
            "start_tls": bool(config_row.start_tls),
            "cache_seconds": config_row.cache_seconds if config_row.cache_seconds is not None else "",
            "list_limit": config_row.list_limit if config_row.list_limit is not None else "",
            "default_role": config_row.default_role or "unassigned",
            "admin_users": config_row.admin_users or "",
            "group_filter": config_row.group_filter or "(&(objectClass=group)(cn=*))",
            "group_attribute": config_row.group_attribute or "cn",
            "group_member_attribute": config_row.group_member_attribute or "member",
        }
    config = _ldap_config()
    config["bind_password"] = ""
    return config


def ldap_lookup_user_email(username, config=None):
    if not username or not ldap_enabled():
        return None
    config = config or _ldap_config()
    email_attr = config.get("email_attribute") or "mail"
    conn = _ldap_service_connection(config=config)
    if not conn:
        return None
    safe_username = escape_filter_chars(username)
    search_filter = (config.get("user_filter") or "(uid={username})").format(
        username=safe_username
    )
    try:
        conn.search(
            config["base_dn"],
            search_filter,
            search_scope=SUBTREE,
            attributes=[email_attr],
            size_limit=1,
        )
    except LDAPInvalidFilterError:
        conn.unbind()
        raise
    except LDAPExceptionError:
        conn.unbind()
        raise
    email_value = None
    if conn.entries:
        try:
            email_value = conn.entries[0][email_attr].value
        except (KeyError, AttributeError):
            email_value = None
    conn.unbind()
    if isinstance(email_value, list):
        email_value = next((item for item in email_value if item), None)
    if email_value:
        return str(email_value).strip() or None
    return None


def resolve_user_email(username):
    if not username:
        return None
    username_norm = normalize_username(username)
    if "@" in username_norm:
        user = User.query.filter(func.lower(User.email) == username_norm.lower()).first()
        if user and user.email:
            return user.email
        return username_norm
    user = get_user_by_username_ci(username_norm)
    if user and user.email:
        return user.email
    email = ldap_lookup_user_email(username_norm)
    if email and user:
        user.email = email
        db.session.commit()
    return email


def get_branding():
    return BrandingConfig.query.first()


def get_branding_name():
    branding = get_branding()
    return branding.company_name or "Asset" if branding else "Asset"


def get_branding_logo_url():
    branding = get_branding()
    if branding and branding.logo_filename:
        return url_for("branding_logo")
    return None


def get_update_status():
    now = time.time()
    if now - UPDATE_CHECK_CACHE["timestamp"] < 300:
        return UPDATE_CHECK_CACHE
    UPDATE_CHECK_CACHE["timestamp"] = now
    UPDATE_CHECK_CACHE["available"] = False
    UPDATE_CHECK_CACHE["last_updated"] = None
    UPDATE_CHECK_CACHE["latest_digest"] = None
    UPDATE_CHECK_CACHE["current_digest"] = None
    UPDATE_CHECK_CACHE["current_tag"] = APP_VERSION or None
    UPDATE_CHECK_CACHE["error"] = None
    try:
        latest_url = f"https://hub.docker.com/v2/repositories/{DOCKERHUB_REPO}/tags/latest"
        with urlopen(latest_url, timeout=4) as response:
            payload = json.loads(response.read().decode("utf-8"))
        last_updated = payload.get("last_updated")
        latest_digest = None
        images = payload.get("images") or []
        if images:
            latest_digest = images[0].get("digest")
        if latest_digest:
            UPDATE_CHECK_CACHE["latest_digest"] = latest_digest
        current_digest = None
        if APP_VERSION:
            current_url = f"https://hub.docker.com/v2/repositories/{DOCKERHUB_REPO}/tags/{APP_VERSION}"
            with urlopen(current_url, timeout=4) as response:
                current_payload = json.loads(response.read().decode("utf-8"))
            current_images = current_payload.get("images") or []
            if current_images:
                current_digest = current_images[0].get("digest")
        UPDATE_CHECK_CACHE["current_digest"] = current_digest
        if latest_digest and current_digest:
            UPDATE_CHECK_CACHE["available"] = latest_digest != current_digest
        if not UPDATE_CHECK_CACHE["available"] and last_updated:
            ts = last_updated.replace("Z", "+00:00")
            updated_at = datetime.datetime.fromisoformat(ts)
            UPDATE_CHECK_CACHE["last_updated"] = updated_at
            if not APP_VERSION:
                app.logger.info("Update check: APP_VERSION not set; falling back to last_updated.")
                UPDATE_CHECK_CACHE["available"] = updated_at > (APP_START_TIME + timedelta(minutes=2))
    except Exception as exc:
        UPDATE_CHECK_CACHE["error"] = str(exc)
        app.logger.warning("Update check failed: %s", exc)
    return UPDATE_CHECK_CACHE


def read_log_tail(lines=200):
    log_path = os.path.join(app.config["LOG_DIR"], app.config["LOG_FILE"])
    if not os.path.exists(log_path):
        return ""
    with open(log_path, "r", encoding="utf-8", errors="ignore") as handle:
        data = handle.readlines()
    return "".join(data[-lines:])


def get_custom_asset_types():
    return AssetType.query.order_by(AssetType.label.asc()).all()


def get_all_asset_keys():
    keys = list(ASSET_DEFS.keys())
    for asset_type in get_custom_asset_types():
        keys.append(f"custom:{asset_type.key}")
    return keys


def get_asset_display_list():
    assets = []
    for key, definition in ASSET_DEFS.items():
        assets.append({"key": key, "label": definition["label"], "kind": "static"})
    for asset_type in get_custom_asset_types():
        assets.append(
            {
                "key": f"custom:{asset_type.key}",
                "label": asset_type.label,
                "kind": "custom",
            }
        )
    return assets


def normalize_assignee(value):
    return (value or "").strip().lower()


def get_user_display_map():
    now = time.time()
    if now - _USER_DISPLAY_CACHE["timestamp"] < 30:
        return _USER_DISPLAY_CACHE["map"]
    mapping = {}
    for user in User.query.all():
        if user.username:
            mapping[user.username.strip().lower()] = user.username
        if user.email:
            mapping[user.email.strip().lower()] = user.username
    _USER_DISPLAY_CACHE["map"] = mapping
    _USER_DISPLAY_CACHE["timestamp"] = now
    return mapping


def display_assignee(value):
    if not value or normalize_assignee(value) == "free":
        return "-"
    value_text = str(value).strip()
    mapping = get_user_display_map()
    return mapping.get(value_text.lower(), value_text)


def log_assignment_change(asset_type, asset_id, old_assigned, new_assigned, actor):
    old_value = (old_assigned or "").strip()
    new_value = (new_assigned or "").strip()
    old_norm = normalize_assignee(old_value)
    new_norm = normalize_assignee(new_value)
    if old_norm == new_norm:
        return
    history = AssetAssignmentHistory(
        asset_type=asset_type,
        asset_id=asset_id,
        from_user=old_value or None,
        to_user=new_value or None,
        assigned_by=actor.username if actor else None,
        assigned_by_id=actor.id if actor else None,
    )
    db.session.add(history)
    db.session.commit()


def normalize_status(value):
    return (value or "").strip().lower().replace("_", " ")


def format_status_label(value):
    return STATUS_LABELS.get(normalize_status(value), value or "-")


def derive_status(status_value, assigned_to):
    status_norm = normalize_status(status_value)
    if status_norm in {"broken", "write off"}:
        return status_norm
    assigned_norm = normalize_assignee(assigned_to)
    if status_norm == "assigned" or (assigned_norm and assigned_norm != "free"):
        return "assigned"
    return "in stock"


def get_item_status(item):
    if hasattr(item, "status"):
        return derive_status(getattr(item, "status", None), getattr(item, "assigned_to", None))
    return derive_status(None, getattr(item, "assigned_to", None))


def get_custom_status(data, assigned_to):
    return derive_status((data or {}).get("status"), assigned_to)


def _normalize_header(value):
    return re.sub(r"\\s+", "_", str(value or "").strip().lower())


def get_static_import_headers(asset_type):
    definition = ASSET_DEFS[asset_type]
    headers = ["ID"]
    headers.extend([label for _, label, _ in definition["fields"]])
    return headers


def get_static_header_map(asset_type):
    definition = ASSET_DEFS[asset_type]
    header_map = {"id": "id"}
    for field_name, label, _field_type in definition["fields"]:
        header_map[_normalize_header(field_name)] = field_name
        header_map[_normalize_header(label)] = field_name
    header_map["user"] = "assigned_to"
    return header_map


def get_custom_import_headers(fields):
    headers = ["ID"]
    headers.extend([field.label for field in fields])
    return headers


def get_custom_header_map(fields):
    header_map = {"id": "id"}
    for field in fields:
        header_map[_normalize_header(field.name)] = field.name
        header_map[_normalize_header(field.label)] = field.name
    header_map["user"] = "assigned_to"
    return header_map


def _parse_bool(value):
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "on"}


def _parse_int_value(value, default=0):
    if value is None or value == "":
        return default
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default


def normalize_connection(value):
    text = str(value or "").strip().lower()
    if "wireless" in text:
        return "wireless"
    if "wired" in text:
        return "wired"
    return ""


def get_mouse_connection(item):
    if getattr(item, "wired", False):
        return "Wired"
    if getattr(item, "wireless", False):
        return "Wireless"
    return "-"


def apply_mouse_connection_to_data(data):
    connection = normalize_connection(data.get("connection"))
    if connection:
        data["wired"] = connection == "wired"
        data["wireless"] = connection == "wireless"
    elif "connection" in data:
        data["wired"] = False
        data["wireless"] = False
    data.pop("connection", None)


def format_static_field_value(field_name, field_type, value):
    if field_type == "checkbox":
        return "Yes" if value else "No"
    if field_name == "status":
        return format_status_label(value)
    if field_name == "assigned_to":
        return display_assignee(value)
    return value if value not in {None, ""} else "-"


def format_custom_field_value(field, value, assigned_fields):
    if field.field_type == "checkbox":
        if field.options:
            return ", ".join(value) if isinstance(value, list) and value else "-"
        return "Yes" if value else "No"
    if field.name == "status":
        return format_status_label(value)
    if field.name in assigned_fields:
        return display_assignee(value)
    return value if value not in {None, ""} else "-"


def format_asset_title(asset_key, label):
    if asset_key in ASSET_TITLE_OVERRIDES:
        return ASSET_TITLE_OVERRIDES[asset_key]
    if label:
        lower = label.lower()
        if lower.endswith("assets"):
            return label
        if lower.endswith("s"):
            return label
        return f"{label}s"
    return "Assets"


def parse_bulk_tags(raw):
    if not raw:
        return []
    parts = re.split(r"[,\n]+", raw)
    seen = set()
    tags = []
    for part in parts:
        value = part.strip()
        if not value or value in seen:
            continue
        seen.add(value)
        tags.append(value)
    return tags


def serialize_model_list(model):
    rows = []
    for item in model.query.all():
        data = {}
        for column in model.__table__.columns:
            data[column.name] = getattr(item, column.name)
        rows.append(data)
    return rows


def _filter_model_fields(model, data):
    allowed = {column.name for column in model.__table__.columns}
    return {key: value for key, value in data.items() if key in allowed}


def _restore_model_rows(model, rows):
    if not rows:
        return
    for row in rows:
        filtered = _filter_model_fields(model, row)
        db.session.add(model(**filtered))


def _reset_sequences(table_names):
    if db.engine.dialect.name != "postgresql":
        return
    for table in table_names:
        try:
            db.session.execute(
                text(
                    "SELECT setval(pg_get_serial_sequence(:table, 'id'), "
                    "COALESCE((SELECT MAX(id) FROM " + table + "), 1), true)"
                ),
                {"table": table},
            )
        except Exception:
            db.session.rollback()


def build_asset_summary():
    summary = []
    for asset_key, definition in ASSET_DEFS.items():
        label = definition["label"]
        model = definition["model"]
        if definition.get("consumable"):
            total = 0
            assigned = 0
            for item in model.query.all():
                total += max(item.total_quantity or 0, 0)
                assigned += max(item.assigned_quantity or 0, 0)
            available = max(total - assigned, 0)
        else:
            total = model.query.count()
            available = is_free_filter(model).count()
            assigned = max(total - available, 0)
        summary.append(
            {
                "label": label,
                "total": total,
                "assigned": assigned,
                "available": available,
            }
        )

    for asset_type in get_custom_asset_types():
        fields = AssetField.query.filter_by(asset_type_id=asset_type.id).all()
        assigned_fields, _ = get_custom_special_fields(fields)
        items = AssetItem.query.filter_by(asset_type_id=asset_type.id).all()
        total = len(items)
        assigned = 0
        available = 0
        for item in items:
            assigned_to = ""
            for field_name in assigned_fields:
                assigned_to = (item.data or {}).get(field_name)
                if assigned_to is not None:
                    break
            status = get_custom_status(item.data or {}, assigned_to)
            if status in {"broken", "write off"}:
                continue
            if status == "assigned":
                assigned += 1
            else:
                available += 1
        summary.append(
            {
                "label": asset_type.label,
                "total": total,
                "assigned": assigned,
                "available": available,
            }
        )
    summary.sort(key=lambda item: item["label"].lower())
    return summary


def _normalize_dept(value):
    return (value or "").strip()


def _dept_filter_match(dept_filter, status, dept_value):
    if dept_filter == "all":
        return True
    if dept_filter == "in_stock":
        return status == "in stock"
    if status != "assigned":
        return False
    return _normalize_dept(dept_value).lower() == dept_filter.lower()


def build_report_rows(dept_filter="all", asset_filter="all", status_filter="all"):
    rows = []
    status_filter_norm = normalize_status(status_filter)
    if status_filter_norm == "available":
        status_filter_norm = "in stock"
    def include_status(status):
        if status_filter_norm in {"all", ""}:
            return status in {"assigned", "in stock"}
        if status_filter_norm == "all status":
            return True
        return status_filter_norm == status

    def append_row(label, asset_tag, vendor, model, processor, ram, hard_disk, user, dept, status, quantity=None):
        rows.append(
            {
                "asset": label,
                "asset_tag": asset_tag or "-",
                "vendor": vendor or "-",
                "model": model or "-",
                "processor": processor or "-",
                "ram": ram or "-",
                "hard_disk": hard_disk or "-",
                "user": user or "-",
                "dept": dept or "-",
                "status": format_status_label(status),
                "status_key": status,
                "quantity": quantity,
            }
        )

    for asset_key, definition in ASSET_DEFS.items():
        if asset_filter not in {"all", asset_key}:
            continue
        label = definition["label"]
        model = definition["model"]
        is_consumable = definition.get("consumable", False)
        if is_consumable:
            for item in model.query.all():
                total = max(item.total_quantity or 0, 0)
                assigned = max(item.assigned_quantity or 0, 0)
                available = max(total - assigned, 0)
                model_label = " ".join(
                    part for part in [getattr(item, "size", None), getattr(item, "speed", None), getattr(item, "vendor", None)] if part
                )
                dept_value = getattr(item, "dept", None)
                if include_status("assigned") and assigned > 0 and _dept_filter_match(dept_filter, "assigned", dept_value):
                    append_row(label, None, getattr(item, "vendor", None), model_label, None, None, None, "-", dept_value or "-", "assigned", assigned)
                if include_status("in stock") and available > 0 and _dept_filter_match(dept_filter, "in stock", None):
                    append_row(label, None, getattr(item, "vendor", None), model_label, None, None, None, "-", "In Stock", "in stock", available)
            continue

        for item in model.query.all():
            assigned_to = getattr(item, "assigned_to", None)
            status = get_item_status(item)
            if status in {"broken", "write off"}:
                dept_value = getattr(item, "dept", None)
            else:
                dept_value = getattr(item, "dept", None) if status == "assigned" else "In Stock"
            if not include_status(status):
                continue
            if not _dept_filter_match(dept_filter, status, dept_value):
                continue
            append_row(
                label,
                getattr(item, "asset_tag", None),
                getattr(item, "vendor", None),
                getattr(item, "model", None) or getattr(item, "size", None),
                getattr(item, "processor", None),
                getattr(item, "ram", None),
                getattr(item, "hard_disk", None),
                display_assignee(assigned_to) if status == "assigned" else ("In Stock" if status == "in stock" else "-"),
                dept_value,
                status,
            )

    for asset_type in get_custom_asset_types():
        if asset_filter not in {"all", asset_type.key}:
            continue
        fields = AssetField.query.filter_by(asset_type_id=asset_type.id).all()
        assigned_fields, model_fields = get_custom_special_fields(fields)
        dept_fields = [field.name for field in fields if field.name.strip().lower() == "dept"]
        asset_tag_fields = [field.name for field in fields if field.name.strip().lower() == "asset_tag"]
        items = AssetItem.query.filter_by(asset_type_id=asset_type.id).all()
        for item in items:
            assigned_to = ""
            for field_name in assigned_fields:
                assigned_to = (item.data or {}).get(field_name)
                if assigned_to is not None:
                    break
            status = get_custom_status(item.data or {}, assigned_to)
            if not include_status(status):
                continue
            dept_value = None
            for field_name in dept_fields:
                dept_value = (item.data or {}).get(field_name)
                if dept_value:
                    break
            dept_value = dept_value if status == "assigned" else ("In Stock" if status == "in stock" else "-")
            if not _dept_filter_match(dept_filter, status, dept_value):
                continue
            model_value = None
            for field_name in model_fields:
                model_value = (item.data or {}).get(field_name)
                if model_value:
                    break
            asset_tag_value = None
            for field_name in asset_tag_fields:
                asset_tag_value = (item.data or {}).get(field_name)
                if asset_tag_value:
                    break
            append_row(
                asset_type.label,
                asset_tag_value,
                None,
                model_value,
                None,
                None,
                None,
                display_assignee(assigned_to) if status == "assigned" else ("In Stock" if status == "in stock" else "-"),
                dept_value,
                status,
            )

    return rows


def build_report_summary(rows):
    summary = {}
    for row in rows:
        entry = summary.setdefault(
            row["asset"], {"label": row["asset"], "assigned": 0, "available": 0}
        )
        quantity = row.get("quantity") or 1
        status_key = row.get("status_key", row.get("status", "")).strip().lower()
        if status_key == "assigned":
            entry["assigned"] += quantity
        elif status_key in {"in stock", "available"}:
            entry["available"] += quantity
    result = list(summary.values())
    result.sort(key=lambda item: item["label"].lower())
    return result


def _report_cache_key(dept_filter, asset_filter, status_filter):
    return f"{dept_filter}|{asset_filter}|{status_filter}"


def get_cached_report_rows(key):
    entry = _REPORT_CACHE["items"].get(key)
    if not entry:
        return None
    if time.time() - entry["timestamp"] > REPORT_CACHE_SECONDS:
        return None
    return entry["rows"]


def set_cached_report_rows(key, rows):
    _REPORT_CACHE["items"][key] = {"rows": rows, "timestamp": time.time()}


def build_dept_summary():
    dept_map = {}
    status_labels = {"broken": "Broken", "write off": "Write Off", "write_off": "Write Off"}

    def ensure_dept(name):
        dept = name or "In Stock"
        if dept not in dept_map:
            dept_map[dept] = {
                "dept": dept,
                "total": 0,
                "assigned": 0,
                "available": 0,
                "types": {},
            }
        return dept_map[dept]

    def bump(dept_name, label, total, assigned, available):
        dept_entry = ensure_dept(dept_name)
        dept_entry["total"] += total
        dept_entry["assigned"] += assigned
        dept_entry["available"] += available
        type_entry = dept_entry["types"].setdefault(
            label, {"label": label, "total": 0, "assigned": 0, "available": 0}
        )
        type_entry["total"] += total
        type_entry["assigned"] += assigned
        type_entry["available"] += available

    for asset_key, definition in ASSET_DEFS.items():
        label = definition["label"]
        model = definition["model"]
        if definition.get("consumable"):
            for item in model.query.all():
                total = max(item.total_quantity or 0, 0)
                assigned = max(item.assigned_quantity or 0, 0)
                available = max(total - assigned, 0)
                status_norm = normalize_status(getattr(item, "status", None))
                if status_norm in {"broken", "write off", "write_off"}:
                    dept = status_labels.get(status_norm, "Broken")
                    bump(dept, label, available, 0, available)
                else:
                    dept = getattr(item, "dept", None)
                    bump(dept, label, total, assigned, available)
        else:
            for item in model.query.all():
                status = get_item_status(item)
                if status in {"broken", "write off"}:
                    dept = status_labels.get(status, "Broken")
                    bump(dept, label, 1, 0, 1)
                    continue
                is_assigned = status == "assigned"
                dept = getattr(item, "dept", None) if is_assigned else None
                bump(dept, label, 1, 1 if is_assigned else 0, 0 if is_assigned else 1)

    for asset_type in get_custom_asset_types():
        fields = AssetField.query.filter_by(asset_type_id=asset_type.id).all()
        assigned_fields, _ = get_custom_special_fields(fields)
        for item in AssetItem.query.filter_by(asset_type_id=asset_type.id).all():
            assigned_to = ""
            for field_name in assigned_fields:
                assigned_to = (item.data or {}).get(field_name)
                if assigned_to is not None:
                    break
            status = get_custom_status(item.data or {}, assigned_to)
            if status in {"broken", "write off"}:
                dept = status_labels.get(status, "Broken")
                bump(dept, asset_type.label, 1, 0, 1)
                continue
            is_assigned = status == "assigned"
            bump(dept, asset_type.label, 1, 1 if is_assigned else 0, 0 if is_assigned else 1)

    summary = list(dept_map.values())
    priority = {"In Stock": 0, "Broken": 1, "Write Off": 2}
    summary.sort(key=lambda item: (priority.get(item["dept"], 99), item["dept"].lower()))
    for dept_entry in summary:
        if dept_entry["dept"] in {"In Stock", "Broken", "Write Off"}:
            dept_entry["assigned"] = 0
            dept_entry["total"] = dept_entry["available"]
        else:
            dept_entry["available"] = 0
            dept_entry["total"] = dept_entry["assigned"]
        dept_entry["types"] = sorted(
            dept_entry["types"].values(), key=lambda item: item["label"].lower()
        )
        for type_entry in dept_entry["types"]:
            if dept_entry["dept"] in {"In Stock", "Broken", "Write Off"}:
                type_entry["assigned"] = 0
                type_entry["total"] = type_entry["available"]
            else:
                type_entry["available"] = 0
                type_entry["total"] = type_entry["assigned"]
    return summary


def get_builtin_asset_types():
    return [
        {
            "key": key,
            "label": definition["label"],
            "fields_count": len(definition.get("fields", [])),
        }
        for key, definition in ASSET_DEFS.items()
    ]


def get_custom_asset_by_key(asset_key):
    return AssetType.query.filter_by(key=asset_key).first()


def is_assigned_field(field):
    name_norm = field.name.strip().lower().replace(" ", "_")
    label_norm = (field.label or "").strip().lower().replace(" ", "_")
    return name_norm == "assigned_to" or label_norm == "assigned_to"


def is_model_field(field):
    name_norm = field.name.strip().lower().replace(" ", "_")
    label_norm = (field.label or "").strip().lower().replace(" ", "_")
    return name_norm == "model" or label_norm == "model"


def get_custom_special_fields(fields):
    assigned_fields = [field.name for field in fields if is_assigned_field(field)]
    model_fields = [field.name for field in fields if is_model_field(field)]
    return assigned_fields, model_fields


def ensure_default_roles():
    existing = {role.name for role in Role.query.all()}
    roles = [
        ("admin", True, True, True, True, True, True),
        ("operator", True, False, True, False, False, False),
        ("reader", False, False, True, False, False, False),
        ("unassigned", False, False, False, False, False, False),
        ("app_admin", False, False, False, True, False, True),
    ]
    for name, can_add, can_delete, can_read, is_app_admin, can_bulk_delete, can_manage_depts in roles:
        if name in existing:
            role = Role.query.filter_by(name=name).first()
            if role:
                if role.is_app_admin != is_app_admin:
                    role.is_app_admin = is_app_admin
                if role.can_bulk_delete != can_bulk_delete:
                    role.can_bulk_delete = can_bulk_delete
                if role.can_manage_depts != can_manage_depts:
                    role.can_manage_depts = can_manage_depts
            continue
        db.session.add(
            Role(
                name=name,
                can_add=can_add,
                can_delete=can_delete,
                can_read=can_read,
                is_app_admin=is_app_admin,
                can_bulk_delete=can_bulk_delete,
                can_manage_depts=can_manage_depts,
            )
        )
    db.session.commit()


def ensure_default_users():
    if User.query.first():
        return
    users = [
        ("admin", "admin", "admin"),
        ("operator", "operator", "operator"),
        ("reader", "reader", "reader"),
    ]
    for username, password, role in users:
        db.session.add(
            User(
                username=username,
                password_hash=generate_password_hash(password),
                role=role,
            )
        )
    db.session.commit()


def ensure_role_permissions():
    roles = Role.query.all()
    for role in roles:
        for asset_key in get_all_asset_keys():
            existing = RolePermission.query.filter_by(
                role_id=role.id, asset_type=asset_key
            ).first()
            if existing:
                if role.name in {"admin", "operator", "reader"}:
                    if existing.can_bulk_delete != role.can_bulk_delete:
                        existing.can_bulk_delete = role.can_bulk_delete
                continue
            if role.name in {"admin", "operator", "reader"}:
                can_add = role.can_add
                can_delete = role.can_delete
                can_read = role.can_read
                can_bulk_delete = role.can_bulk_delete
            else:
                can_add = False
                can_delete = False
                can_read = False
                can_bulk_delete = False
            db.session.add(
                RolePermission(
                    role_id=role.id,
                    asset_type=asset_key,
                    can_add=can_add,
                    can_delete=can_delete,
                    can_read=can_read,
                    can_bulk_delete=can_bulk_delete,
                )
            )
    db.session.commit()


def ensure_asset_field_options_column():
    if db.engine.dialect.name != "sqlite":
        return
    result = db.session.execute(text("PRAGMA table_info(asset_field)")).fetchall()
    if not result:
        return
    columns = {row[1] for row in result}
    if "options" not in columns:
        db.session.execute(text("ALTER TABLE asset_field ADD COLUMN options TEXT"))
        db.session.commit()


def ensure_asset_tag_columns():
    if db.engine.dialect.name != "sqlite":
        return
    for table_name in ("laptop", "computer", "screen"):
        result = db.session.execute(text(f"PRAGMA table_info({table_name})")).fetchall()
        if not result:
            continue
        columns = {row[1] for row in result}
        if "asset_tag" not in columns:
            db.session.execute(
                text(f"ALTER TABLE {table_name} ADD COLUMN asset_tag VARCHAR(80)")
            )
    db.session.commit()


def ensure_dept_columns():
    if db.engine.dialect.name != "sqlite":
        return
    for table_name in ("laptop", "computer", "screen", "keyboard", "mouse", "headset", "ram"):
        result = db.session.execute(text(f"PRAGMA table_info({table_name})")).fetchall()
        if not result:
            continue
        columns = {row[1] for row in result}
        if "dept" not in columns:
            db.session.execute(text(f"ALTER TABLE {table_name} ADD COLUMN dept VARCHAR(80)"))
    db.session.commit()


def ensure_status_columns():
    tables = ("laptop", "computer", "screen", "keyboard", "mouse", "headset", "ram")
    if db.engine.dialect.name == "sqlite":
        for table_name in tables:
            result = db.session.execute(text(f"PRAGMA table_info({table_name})")).fetchall()
            if not result:
                continue
            columns = {row[1] for row in result}
            if "status" not in columns:
                db.session.execute(
                    text(f"ALTER TABLE {table_name} ADD COLUMN status VARCHAR(20)")
                )
        db.session.commit()
        return
    for table_name in tables:
        try:
            db.session.execute(
                text(f"ALTER TABLE {table_name} ADD COLUMN IF NOT EXISTS status VARCHAR(20)")
            )
        except Exception:
            db.session.rollback()
    db.session.commit()


def ensure_ram_type_column():
    tables = ("ram",)
    if db.engine.dialect.name == "sqlite":
        for table_name in tables:
            result = db.session.execute(text(f"PRAGMA table_info({table_name})")).fetchall()
            if not result:
                continue
            columns = {row[1] for row in result}
            if "ram_type" not in columns:
                db.session.execute(
                    text(f"ALTER TABLE {table_name} ADD COLUMN ram_type VARCHAR(20)")
                )
        db.session.commit()
        return
    for table_name in tables:
        try:
            db.session.execute(
                text(f"ALTER TABLE {table_name} ADD COLUMN IF NOT EXISTS ram_type VARCHAR(20)")
            )
        except Exception:
            db.session.rollback()
    db.session.commit()


def ensure_role_admin_column():
    if db.engine.dialect.name != "sqlite":
        return
    result = db.session.execute(text("PRAGMA table_info(role)")).fetchall()
    if not result:
        return
    columns = {row[1] for row in result}
    if "is_app_admin" not in columns:
        db.session.execute(
            text("ALTER TABLE role ADD COLUMN is_app_admin BOOLEAN DEFAULT 0")
        )
    if "can_bulk_delete" not in columns:
        db.session.execute(
            text("ALTER TABLE role ADD COLUMN can_bulk_delete BOOLEAN DEFAULT 0")
        )
    if "can_manage_depts" not in columns:
        db.session.execute(
            text("ALTER TABLE role ADD COLUMN can_manage_depts BOOLEAN DEFAULT 0")
        )
    db.session.commit()


def ensure_user_email_column():
    if db.engine.dialect.name != "sqlite":
        return
    result = db.session.execute(text("PRAGMA table_info(user)")).fetchall()
    if not result:
        return
    columns = {row[1] for row in result}
    if "email" not in columns:
        db.session.execute(text("ALTER TABLE user ADD COLUMN email TEXT"))
    db.session.commit()


def ensure_role_permission_bulk_column():
    if db.engine.dialect.name != "sqlite":
        return
    result = db.session.execute(text("PRAGMA table_info(role_permission)")).fetchall()
    if not result:
        return
    columns = {row[1] for row in result}
    if "can_bulk_delete" not in columns:
        db.session.execute(
            text("ALTER TABLE role_permission ADD COLUMN can_bulk_delete BOOLEAN DEFAULT 0")
        )
    db.session.commit()


def ensure_ldap_group_columns():
    if db.engine.dialect.name != "sqlite":
        return
    result = db.session.execute(text("PRAGMA table_info(ldap_config)")).fetchall()
    if not result:
        return
    columns = {row[1] for row in result}
    if "group_filter" not in columns:
        db.session.execute(text("ALTER TABLE ldap_config ADD COLUMN group_filter TEXT"))
    if "group_attribute" not in columns:
        db.session.execute(text("ALTER TABLE ldap_config ADD COLUMN group_attribute TEXT"))
    if "group_member_attribute" not in columns:
        db.session.execute(
            text("ALTER TABLE ldap_config ADD COLUMN group_member_attribute TEXT")
        )
    if "email_attribute" not in columns:
        db.session.execute(text("ALTER TABLE ldap_config ADD COLUMN email_attribute TEXT"))
    db.session.commit()


def ensure_role_assignment_tables():
    db.session.execute(
        text(
            "CREATE TABLE IF NOT EXISTS user_role ("
            "id INTEGER PRIMARY KEY,"
            "user_id INTEGER NOT NULL,"
            "role_id INTEGER NOT NULL,"
            "UNIQUE(user_id, role_id)"
            ")"
        )
    )
    db.session.execute(
        text(
            "CREATE TABLE IF NOT EXISTS group_role ("
            "id INTEGER PRIMARY KEY,"
            "group_id INTEGER NOT NULL,"
            "role_id INTEGER NOT NULL,"
            "UNIQUE(group_id, role_id)"
            ")"
        )
    )
    db.session.commit()


def ensure_role_assignments():
    roles = {role.name: role.id for role in Role.query.all()}
    if not roles:
        return
    for user in User.query.all():
        role_id = roles.get(user.role)
        if not role_id:
            continue
        exists = UserRole.query.filter_by(user_id=user.id, role_id=role_id).first()
        if not exists:
            db.session.add(UserRole(user_id=user.id, role_id=role_id))
    for group in Group.query.all():
        role_id = roles.get(group.role)
        if not role_id:
            continue
        exists = GroupRole.query.filter_by(group_id=group.id, role_id=role_id).first()
        if not exists:
            db.session.add(GroupRole(group_id=group.id, role_id=role_id))
    db.session.commit()


def ensure_smtp_columns():
    if db.engine.dialect.name != "sqlite":
        return
    result = db.session.execute(text("PRAGMA table_info(smtp_config)")).fetchall()
    if result:
        columns = {row[1] for row in result}
        if "monthly_report_enabled" not in columns:
            db.session.execute(
                text("ALTER TABLE smtp_config ADD COLUMN monthly_report_enabled BOOLEAN DEFAULT 0")
            )
        if "monthly_report_day" not in columns:
            db.session.execute(
                text("ALTER TABLE smtp_config ADD COLUMN monthly_report_day INTEGER DEFAULT 1")
            )
        if "low_stock_enabled" not in columns:
            db.session.execute(
                text("ALTER TABLE smtp_config ADD COLUMN low_stock_enabled BOOLEAN DEFAULT 0")
            )
        if "low_stock_threshold" not in columns:
            db.session.execute(
                text("ALTER TABLE smtp_config ADD COLUMN low_stock_threshold INTEGER DEFAULT 5")
            )
        if "low_stock_frequency_days" not in columns:
            db.session.execute(
                text("ALTER TABLE smtp_config ADD COLUMN low_stock_frequency_days INTEGER DEFAULT 1")
            )
    recipient_result = db.session.execute(text("PRAGMA table_info(smtp_recipient)")).fetchall()
    if recipient_result:
        recipient_columns = {row[1] for row in recipient_result}
        if "notify_monthly" not in recipient_columns:
            db.session.execute(
                text("ALTER TABLE smtp_recipient ADD COLUMN notify_monthly BOOLEAN DEFAULT 0")
            )
        if "notify_low_stock" not in recipient_columns:
            db.session.execute(
                text("ALTER TABLE smtp_recipient ADD COLUMN notify_low_stock BOOLEAN DEFAULT 0")
            )
    db.session.commit()


def ensure_low_stock_table():
    db.session.execute(
        text(
            "CREATE TABLE IF NOT EXISTS low_stock_state ("
            "id INTEGER PRIMARY KEY,"
            "asset_key TEXT NOT NULL,"
            "entity_id TEXT,"
            "last_notified_at DATETIME,"
            "UNIQUE(asset_key, entity_id)"
            ")"
        )
    )
    db.session.commit()


def ensure_custom_dept_field():
    for asset_type in get_custom_asset_types():
        exists = AssetField.query.filter_by(
            asset_type_id=asset_type.id, name="dept"
        ).first()
        if exists:
            continue
        db.session.add(
            AssetField(
                asset_type_id=asset_type.id,
                name="dept",
                label="Dept",
                field_type="text",
                options=[],
            )
        )
    db.session.commit()


def ensure_custom_status_field():
    for asset_type in get_custom_asset_types():
        exists = AssetField.query.filter_by(
            asset_type_id=asset_type.id, name="status"
        ).first()
        if exists:
            continue
        db.session.add(
            AssetField(
                asset_type_id=asset_type.id,
                name="status",
                label="Status",
                field_type="select",
                options=STATUS_OPTIONS,
            )
        )
    db.session.commit()


def backfill_status_values():
    for asset_key, definition in ASSET_DEFS.items():
        model = definition["model"]
        if not hasattr(model, "status"):
            continue
        items = model.query.filter(
            (model.status.is_(None)) | (func.trim(model.status) == "")
        ).all()
        if not items:
            continue
        for item in items:
            status_key = derive_status(None, getattr(item, "assigned_to", None))
            if status_key == "assigned":
                item.status = "Assigned"
            else:
                item.status = "In Stock"
    for asset_type in get_custom_asset_types():
        fields = AssetField.query.filter_by(asset_type_id=asset_type.id).all()
        assigned_fields, _ = get_custom_special_fields(fields)
        items = AssetItem.query.filter_by(asset_type_id=asset_type.id).all()
        for item in items:
            data = dict(item.data or {})
            if data.get("status"):
                continue
            assigned_to = ""
            for field_name in assigned_fields:
                assigned_to = data.get(field_name)
                if assigned_to is not None:
                    break
            status_key = derive_status(None, assigned_to)
            data["status"] = "Assigned" if status_key == "assigned" else "In Stock"
            item.data = data
    db.session.commit()


def ensure_branding_table():
    db.session.execute(
        text(
            "CREATE TABLE IF NOT EXISTS branding_config ("
            "id INTEGER PRIMARY KEY,"
            "company_name TEXT,"
            "logo_filename TEXT"
            ")"
        )
    )
    db.session.commit()


def ensure_department_table():
    db.session.execute(
        text(
            "CREATE TABLE IF NOT EXISTS department ("
            "id INTEGER PRIMARY KEY,"
            "name TEXT NOT NULL UNIQUE"
            ")"
        )
    )
    db.session.commit()


def ensure_builtin_asset_settings():
    if db.engine.dialect.name == "sqlite":
        columns = db.session.execute(text("PRAGMA table_info(builtin_asset_field_setting)")).fetchall()
        if columns:
            column_names = {row[1] for row in columns}
            if "options" not in column_names:
                db.session.execute(
                    text("ALTER TABLE builtin_asset_field_setting ADD COLUMN options TEXT")
                )
                db.session.commit()
    for key, definition in ASSET_DEFS.items():
        setting = BuiltinAssetTypeSetting.query.filter_by(key=key).first()
        if not setting:
            setting = BuiltinAssetTypeSetting(key=key, label=definition["label"])
            db.session.add(setting)
        for field_name, label, _ in definition.get("fields", []):
            field_setting = BuiltinAssetFieldSetting.query.filter_by(
                asset_key=key, field_name=field_name
            ).first()
            if not field_setting:
                field_setting = BuiltinAssetFieldSetting(
                    asset_key=key, field_name=field_name, label=label
                )
                db.session.add(field_setting)
    db.session.commit()


def ensure_indexes():
    statements = [
        "CREATE INDEX IF NOT EXISTS idx_laptop_asset_tag ON laptop(asset_tag)",
        "CREATE INDEX IF NOT EXISTS idx_laptop_model ON laptop(model)",
        "CREATE INDEX IF NOT EXISTS idx_laptop_assigned_to ON laptop(assigned_to)",
        "CREATE INDEX IF NOT EXISTS idx_laptop_dept ON laptop(dept)",
        "CREATE INDEX IF NOT EXISTS idx_computer_asset_tag ON computer(asset_tag)",
        "CREATE INDEX IF NOT EXISTS idx_computer_model ON computer(model)",
        "CREATE INDEX IF NOT EXISTS idx_computer_assigned_to ON computer(assigned_to)",
        "CREATE INDEX IF NOT EXISTS idx_computer_dept ON computer(dept)",
        "CREATE INDEX IF NOT EXISTS idx_screen_asset_tag ON screen(asset_tag)",
        "CREATE INDEX IF NOT EXISTS idx_screen_model ON screen(model)",
        "CREATE INDEX IF NOT EXISTS idx_screen_assigned_to ON screen(assigned_to)",
        "CREATE INDEX IF NOT EXISTS idx_screen_dept ON screen(dept)",
        "CREATE INDEX IF NOT EXISTS idx_keyboard_model ON keyboard(model)",
        "CREATE INDEX IF NOT EXISTS idx_keyboard_assigned_to ON keyboard(assigned_to)",
        "CREATE INDEX IF NOT EXISTS idx_keyboard_dept ON keyboard(dept)",
        "CREATE INDEX IF NOT EXISTS idx_mouse_model ON mouse(model)",
        "CREATE INDEX IF NOT EXISTS idx_mouse_assigned_to ON mouse(assigned_to)",
        "CREATE INDEX IF NOT EXISTS idx_mouse_dept ON mouse(dept)",
        "CREATE INDEX IF NOT EXISTS idx_headset_model ON headset(model)",
        "CREATE INDEX IF NOT EXISTS idx_headset_assigned_to ON headset(assigned_to)",
        "CREATE INDEX IF NOT EXISTS idx_headset_dept ON headset(dept)",
        "CREATE INDEX IF NOT EXISTS idx_ram_vendor ON ram(vendor)",
        "CREATE INDEX IF NOT EXISTS idx_ram_assigned_to ON ram(assigned_to)",
        "CREATE INDEX IF NOT EXISTS idx_ram_dept ON ram(dept)",
    ]
    for statement in statements:
        try:
            db.session.execute(text(statement))
        except Exception:
            db.session.rollback()
    db.session.commit()


def normalize_ram_quantities():
    rows = (
        Ram.query.filter(Ram.total_quantity > 1)
        .filter((Ram.assigned_quantity.is_(None)) | (Ram.assigned_quantity <= 0))
        .all()
    )
    if not rows:
        return
    for row in rows:
        assigned_to = normalize_assignee(row.assigned_to)
        if assigned_to and assigned_to != "free":
            continue
        total = max(row.total_quantity or 0, 0)
        if total <= 1:
            continue
        for _ in range(total):
            item = Ram(
                size=row.size,
                speed=row.speed,
                vendor=row.vendor,
                dept=row.dept,
                assigned_to="free",
                status="In Stock",
                total_quantity=1,
                assigned_quantity=0,
            )
            db.session.add(item)
        db.session.delete(row)
    db.session.commit()


def apply_builtin_overrides():
    settings = {setting.key: setting for setting in BuiltinAssetTypeSetting.query.all()}
    field_settings = BuiltinAssetFieldSetting.query.all()
    option_map = {}
    for field_setting in field_settings:
        if field_setting.options:
            items = [item.strip() for item in field_setting.options.split(",") if item.strip()]
            if items:
                option_map[(field_setting.asset_key, field_setting.field_name)] = items
    for key, definition in ASSET_DEFS.items():
        if "base_fields" not in definition:
            definition["base_fields"] = list(definition.get("fields", []))
        setting = settings.get(key)
        if setting:
            definition["label"] = setting.label
        new_fields = []
        field_options = dict(definition.get("field_options") or {})
        for name, label, field_type in definition.get("base_fields", []):
            override = next(
                (
                    setting
                    for setting in field_settings
                    if setting.asset_key == key and setting.field_name == name
                ),
                None,
            )
            new_label = override.label if override else label
            options = option_map.get((key, name), [])
            new_type = "select" if options else field_type
            new_fields.append((name, new_label, new_type))
            if options:
                field_options[name] = options
        definition["fields"] = new_fields
        definition["field_options"] = field_options


@app.before_request
def init_db():
    global _DB_INIT_DONE
    if not _DB_INIT_DONE:
        db.create_all()
        ensure_asset_field_options_column()
        ensure_asset_tag_columns()
        ensure_dept_columns()
        ensure_status_columns()
        ensure_ram_type_column()
        ensure_role_admin_column()
        ensure_user_email_column()
        ensure_role_permission_bulk_column()
        ensure_ldap_group_columns()
        ensure_role_assignment_tables()
        ensure_default_roles()
        ensure_role_assignments()
        ensure_smtp_columns()
        ensure_low_stock_table()
        ensure_branding_table()
        ensure_department_table()
        ensure_builtin_asset_settings()
        apply_builtin_overrides()
        ensure_custom_dept_field()
        ensure_custom_status_field()
        backfill_status_values()
        ensure_default_users()
        ensure_role_permissions()
        ensure_indexes()
        normalize_ram_quantities()
        _DB_INIT_DONE = True
    if request.method == "GET":
        endpoint = request.endpoint
        if endpoint in SECTION_ENDPOINTS:
            path = request.full_path
            if path.endswith("?"):
                path = path[:-1]
            session["last_section_url"] = path


def get_current_user():
    user_id = session.get("user_id")
    if not user_id:
        return None
    return User.query.get(user_id)


def get_user_role_names(user):
    if not user:
        return []
    names = []
    if user.role and user.role != "unassigned":
        names.append(user.role)
    role_ids = [row.role_id for row in UserRole.query.filter_by(user_id=user.id).all()]
    if role_ids:
        roles = Role.query.filter(Role.id.in_(role_ids)).all()
        for role in roles:
            if role.name and role.name != "unassigned":
                names.append(role.name)
    group_ids = [
        member.group_id
        for member in GroupMember.query.filter_by(user_id=user.id).all()
    ]
    if group_ids:
        groups = Group.query.filter(Group.id.in_(group_ids)).all()
        for group in groups:
            if group.role and group.role != "unassigned":
                names.append(group.role)
    if group_ids:
        group_role_ids = [
            row.role_id
            for row in GroupRole.query.filter(GroupRole.group_id.in_(group_ids)).all()
        ]
    else:
        group_role_ids = []
    if group_role_ids:
        roles = Role.query.filter(Role.id.in_(group_role_ids)).all()
        for role in roles:
            if role.name and role.name != "unassigned":
                names.append(role.name)
    return list(dict.fromkeys(names))


def user_has_app_admin(user):
    if not user:
        return False
    role_names = get_user_role_names(user)
    if not role_names:
        return False
    roles = Role.query.filter(Role.name.in_(role_names)).all()
    return any(role.is_app_admin for role in roles)


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not get_current_user():
            return redirect(url_for("login"))
        return view(*args, **kwargs)

    return wrapped


def require_roles(*roles):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            user = get_current_user()
            if not user:
                return redirect(url_for("login"))
            allowed = user.role in roles
            if not allowed:
                group_roles = set(get_user_role_names(user))
                allowed = bool(group_roles.intersection(set(roles)))
            if not allowed:
                flash("You do not have permission to perform this action.", "error")
                return redirect(request.referrer or url_for("index"))
            return view(*args, **kwargs)

        return wrapped

    return decorator


def get_role_permissions(user):
    if not user:
        return {
            "can_add": False,
            "can_delete": False,
            "can_read": False,
            "can_bulk_delete": False,
            "can_manage_depts": False,
        }
    role_names = get_user_role_names(user)
    if not role_names:
        return {
            "can_add": False,
            "can_delete": False,
            "can_read": False,
            "can_bulk_delete": False,
            "can_manage_depts": False,
        }
    roles = Role.query.filter(Role.name.in_(role_names)).all()
    return {
        "can_add": any(role.can_add for role in roles),
        "can_delete": any(role.can_delete for role in roles),
        "can_read": any(role.can_read for role in roles),
        "can_bulk_delete": any(role.can_bulk_delete for role in roles),
        "can_manage_depts": any(role.can_manage_depts for role in roles),
    }


def get_role_asset_permissions(user, asset_type):
    if not user:
        return {
            "can_add": False,
            "can_delete": False,
            "can_read": False,
            "can_bulk_delete": False,
        }
    role_names = get_user_role_names(user)
    if not role_names:
        return {
            "can_add": False,
            "can_delete": False,
            "can_read": False,
            "can_bulk_delete": False,
        }
    roles = Role.query.filter(Role.name.in_(role_names)).all()
    if not roles:
        return {
            "can_add": False,
            "can_delete": False,
            "can_read": False,
            "can_bulk_delete": False,
        }
    perms = (
        RolePermission.query.filter(RolePermission.role_id.in_([role.id for role in roles]))
        .filter_by(asset_type=asset_type)
        .all()
    )
    return {
        "can_add": any(perm.can_add for perm in perms),
        "can_delete": any(perm.can_delete for perm in perms),
        "can_read": any(perm.can_read for perm in perms),
        "can_bulk_delete": any(perm.can_bulk_delete for perm in perms),
    }


def require_static_permission(permission):
    def decorator(view):
        @wraps(view)
        def wrapped(asset_type, *args, **kwargs):
            if asset_type not in ASSET_DEFS:
                return redirect(url_for("index"))
            user = get_current_user()
            if not user:
                return redirect(url_for("login"))
            perms = get_role_asset_permissions(user, asset_type)
            if not perms.get(permission, False):
                if permission in {"can_delete", "can_bulk_delete"}:
                    log_audit(
                        "delete_attempt",
                        "asset",
                        entity_id=asset_type,
                        success=False,
                        details=f"Permission denied: {permission}",
                    )
                flash("You do not have permission to perform this action.", "error")
                return redirect(request.referrer or url_for("index"))
            return view(asset_type, *args, **kwargs)

        return wrapped

    return decorator


def require_custom_permission(permission):
    def decorator(view):
        @wraps(view)
        def wrapped(asset_key, *args, **kwargs):
            user = get_current_user()
            if not user:
                return redirect(url_for("login"))
            perms = get_role_asset_permissions(user, f"custom:{asset_key}")
            if not perms.get(permission, False):
                if permission in {"can_delete", "can_bulk_delete"}:
                    log_audit(
                        "delete_attempt",
                        "custom_asset",
                        entity_id=asset_key,
                        success=False,
                        details=f"Permission denied: {permission}",
                    )
                flash("You do not have permission to perform this action.", "error")
                return redirect(request.referrer or url_for("index"))
            return view(asset_key, *args, **kwargs)

        return wrapped

    return decorator


def require_permission(permission):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            user = get_current_user()
            if not user:
                return redirect(url_for("login"))
            perms = get_role_permissions(user)
            if not perms.get(permission, False):
                flash("You do not have permission to perform this action.", "error")
                return redirect(request.referrer or url_for("index"))
            return view(*args, **kwargs)

        return wrapped

    return decorator


def require_app_admin(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        user = get_current_user()
        if not user:
            return redirect(url_for("login"))
        if not user_has_app_admin(user):
            log_audit(
                "admin_attempt",
                "app_admin",
                success=False,
                details="App admin access denied",
            )
            flash("You do not have permission to perform this action.", "error")
            return redirect(request.referrer or url_for("index"))
        return view(*args, **kwargs)

    return wrapped


def require_bulk_delete(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        user = get_current_user()
        if not user:
            return redirect(url_for("login"))
        if not user_has_app_admin(user):
            log_audit(
                "bulk_delete_attempt",
                "app_admin",
                success=False,
                details="App admin access denied",
            )
            flash("You do not have permission to perform this action.", "error")
            return redirect(request.referrer or url_for("index"))
        perms = get_role_permissions(user)
        if not perms.get("can_bulk_delete", False):
            log_audit(
                "bulk_delete_attempt",
                "permission",
                success=False,
                details="Bulk delete permission denied",
            )
            flash("You do not have permission to perform this action.", "error")
            return redirect(request.referrer or url_for("index"))
        return view(*args, **kwargs)

    return wrapped


def api_auth_required(asset_type=None, permission=None):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            auth_header = request.headers.get("Authorization", "")
            if not auth_header.startswith("Bearer "):
                return jsonify({"error": "Missing token"}), 401
            token = auth_header.split(" ", 1)[1].strip()
            try:
                payload = _jwt_decode(token)
            except Exception:
                return jsonify({"error": "Invalid token"}), 401
            if payload.get("type") != "access":
                return jsonify({"error": "Invalid token type"}), 401
            user_id = payload.get("sub")
            user = User.query.get(user_id) if user_id else None
            if not user:
                return jsonify({"error": "User not found"}), 401
            request.api_user = user
            if permission:
                if asset_type:
                    perms = get_role_asset_permissions(user, asset_type)
                    if not perms.get(permission, False):
                        return jsonify({"error": "Forbidden"}), 403
                else:
                    perms = get_role_permissions(user)
                    if not perms.get(permission, False):
                        return jsonify({"error": "Forbidden"}), 403
            return view(*args, **kwargs)
        return wrapped
    return decorator


def is_free_filter(model):
    if hasattr(model, "status"):
        return model.query.filter(
            (model.status.is_(None))
            | (func.trim(model.status) == "")
            | (func.lower(model.status) == "in stock")
            | (func.lower(model.status) == "in_stock")
        ).filter(
            (model.assigned_to.is_(None))
            | (func.trim(model.assigned_to) == "")
            | (func.lower(model.assigned_to) == "free")
        )
    return model.query.filter(
        (model.assigned_to.is_(None))
        | (func.trim(model.assigned_to) == "")
        | (func.lower(model.assigned_to) == "free")
    )


def apply_status_filter_query(model, query_builder, status_filter):
    status_norm = normalize_status(status_filter)
    if status_norm in {"", "all"}:
        return query_builder
    if not hasattr(model, "status"):
        return query_builder
    if status_norm in {"in stock", "in_stock", "available"}:
        return query_builder.filter(
            (model.status.is_(None))
            | (func.trim(model.status) == "")
            | (func.lower(model.status) == "in stock")
            | (func.lower(model.status) == "in_stock")
        )
    if status_norm == "assigned":
        return query_builder.filter(func.lower(model.status) == "assigned")
    if status_norm in {"broken", "write off", "write_off"}:
        return query_builder.filter(
            func.lower(model.status).in_(["broken", "write off", "write_off"])
        )
    return query_builder


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = normalize_username(request.form.get("username", ""))
        password = request.form.get("password", "")
        remember = "remember_me" in request.form
        user = get_user_by_username_ci(username)
        local_ok = user and check_password_hash(user.password_hash, password)
        if local_ok:
            if not get_user_role_names(user):
                log_audit("login_denied", "auth", success=False, details=username)
                flash("Your account is not assigned to any role.", "error")
                return redirect(url_for("login"))
            session["user_id"] = user.id
            session.permanent = remember
            if not get_role_permissions(user).get("can_read", False) and user_has_app_admin(user):
                return redirect(url_for("list_users"))
            log_audit("login", "auth", details=username)
            return redirect(url_for("index"))

        try:
            ldap_ok = ldap_authenticate(username, password)
        except Exception as exc:
            app.logger.warning("LDAP authentication unavailable: %s", exc)
            ldap_ok = False

        if ldap_ok:
            if not user:
                user = ensure_ldap_user(username)
            if not user or not get_user_role_names(user):
                log_audit("login_denied", "auth", success=False, details=username)
                flash("Your account is not assigned to any role.", "error")
                return redirect(url_for("login"))
            session["user_id"] = user.id
            session.permanent = remember
            if not get_role_permissions(user).get("can_read", False) and user_has_app_admin(user):
                return redirect(url_for("list_users"))
            log_audit("login", "auth", details=username)
            return redirect(url_for("index"))

        log_audit("login_failed", "auth", success=False, details=username)
        flash("Invalid username or password.", "error")
        return redirect(url_for("login"))
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/api/auth/login", methods=["POST"])
def api_login():
    payload = request.get_json(silent=True) or {}
    username = normalize_username(payload.get("username", ""))
    password = payload.get("password", "")
    if not username or not password:
        return jsonify({"error": "Missing credentials"}), 400
    user = get_user_by_username_ci(username)
    if not user and "@" in username:
        user = User.query.filter(func.lower(User.email) == username.lower()).first()
    local_ok = user and check_password_hash(user.password_hash, password)
    if not local_ok:
        try:
            ldap_ok = ldap_authenticate(username, password)
        except Exception as exc:
            app.logger.warning("LDAP authentication unavailable: %s", exc)
            ldap_ok = False
        if ldap_ok:
            if not user:
                user = ensure_ldap_user(username)
        else:
            if "@" in username:
                short_name = username.split("@", 1)[0]
                try:
                    ldap_ok = ldap_authenticate(short_name, password)
                except Exception as exc:
                    app.logger.warning("LDAP authentication unavailable: %s", exc)
                    ldap_ok = False
                if ldap_ok and not user:
                    user = ensure_ldap_user(short_name)
    if not user:
        log_audit("login_failed", "api_auth", success=False, details=username)
        return jsonify({"error": "Invalid username or password"}), 401
    if not get_user_role_names(user):
        log_audit("login_denied", "api_auth", success=False, details=username)
        return jsonify({"error": "No roles assigned"}), 403
    access_token, refresh_token = _issue_tokens(user.id)
    log_audit("login", "api_auth", details=username)
    return jsonify(
        {
            "access_token": access_token,
            "refresh_token": refresh_token,
            "token_type": "Bearer",
            "expires_in": JWT_ACCESS_SECONDS,
        }
    )


@app.route("/api/auth/refresh", methods=["POST"])
def api_refresh():
    payload = request.get_json(silent=True) or {}
    refresh_token = payload.get("refresh_token", "")
    if not refresh_token:
        return jsonify({"error": "Missing refresh token"}), 400
    try:
        decoded = _jwt_decode(refresh_token)
    except Exception:
        return jsonify({"error": "Invalid refresh token"}), 401
    if decoded.get("type") != "refresh":
        return jsonify({"error": "Invalid token type"}), 401
    user_id = decoded.get("sub")
    user = User.query.get(user_id) if user_id else None
    if not user:
        return jsonify({"error": "User not found"}), 401
    access_token, new_refresh = _rotate_refresh_token(user.id, refresh_token)
    if not access_token:
        return jsonify({"error": "Refresh expired"}), 401
    return jsonify(
        {
            "access_token": access_token,
            "refresh_token": new_refresh,
            "token_type": "Bearer",
            "expires_in": JWT_ACCESS_SECONDS,
        }
    )


@app.route("/api/asset-types", methods=["GET"])
@api_auth_required(permission="can_read")
def api_asset_types():
    return jsonify(
        [
            {"key": key, "label": definition["label"]}
            for key, definition in ASSET_DEFS.items()
        ]
    )


@app.route("/api/users", methods=["GET"])
@api_auth_required(permission="can_read")
def api_users():
    usernames = {user.username for user in User.query.all() if user.username}
    try:
        usernames.update(get_ldap_users())
    except Exception:
        pass
    return jsonify(sorted(usernames, key=str.lower))


@app.route("/api/departments", methods=["GET"])
@api_auth_required(permission="can_read")
def api_departments():
    departments = Department.query.order_by(Department.name.asc()).all()
    return jsonify([dept.name for dept in departments])


@app.route("/api/assets/<asset_type>", methods=["GET"])
@api_auth_required(permission="can_read", asset_type=None)
def api_assets_list(asset_type):
    if asset_type not in ASSET_DEFS:
        return jsonify({"error": "Unknown asset type"}), 404
    user = request.api_user
    if not get_role_asset_permissions(user, asset_type).get("can_read", False):
        return jsonify({"error": "Forbidden"}), 403
    definition = ASSET_DEFS[asset_type]
    page = max(_parse_int(request.args.get("page", ""), 1), 1)
    per_page = max(min(_parse_int(request.args.get("per_page", ""), DEFAULT_PAGE_SIZE), 200), 1)
    query_builder = definition["model"].query
    query = normalize_search(request.args.get("q", ""))
    status_filter = (request.args.get("status") or "all").strip()
    query_builder = apply_status_filter_query(definition["model"], query_builder, status_filter)
    if query:
        filters = []
        for field_name, _label, field_type in definition["fields"]:
            if field_name == "connection" and definition["model"] is Mouse:
                if query in {"wired"}:
                    filters.append(definition["model"].wired.is_(True))
                elif query in {"wireless"}:
                    filters.append(definition["model"].wireless.is_(True))
                continue
            column = getattr(definition["model"], field_name)
            if field_type == "checkbox":
                if query in {"yes", "true", "1"}:
                    filters.append(column.is_(True))
                elif query in {"no", "false", "0"}:
                    filters.append(column.is_(False))
            else:
                filters.append(cast(column, String).ilike(f"%{query}%"))
        if filters:
            query_builder = query_builder.filter(or_(*filters))
    total_items = query_builder.count()
    items = (
        query_builder.order_by(definition["model"].id.asc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )
    rows = []
    for item in items:
        row = {"id": item.id}
        for field_name, _label, field_type in definition["fields"]:
            if field_name == "connection" and definition["model"] is Mouse:
                value = get_mouse_connection(item)
                field_type = "text"
            else:
                value = getattr(item, field_name, None)
            row[field_name] = format_static_field_value(field_name, field_type, value)
        rows.append(row)
    return jsonify(
        {
            "items": rows,
            "page": page,
            "per_page": per_page,
            "total": total_items,
        }
    )


@app.route("/api/assets/<asset_type>/<int:item_id>", methods=["GET"])
@api_auth_required(permission="can_read", asset_type=None)
def api_assets_get(asset_type, item_id):
    if asset_type not in ASSET_DEFS:
        return jsonify({"error": "Unknown asset type"}), 404
    user = request.api_user
    if not get_role_asset_permissions(user, asset_type).get("can_read", False):
        return jsonify({"error": "Forbidden"}), 403
    definition = ASSET_DEFS[asset_type]
    item = definition["model"].query.get_or_404(item_id)
    row = {"id": item.id}
    for field_name, _label, field_type in definition["fields"]:
        if field_name == "connection" and definition["model"] is Mouse:
            value = get_mouse_connection(item)
            field_type = "text"
        else:
            value = getattr(item, field_name, None)
        row[field_name] = format_static_field_value(field_name, field_type, value)
    return jsonify(row)


@app.route("/api/assets/<asset_type>", methods=["POST"])
@api_auth_required(permission="can_add", asset_type=None)
def api_assets_create(asset_type):
    if asset_type not in ASSET_DEFS:
        return jsonify({"error": "Unknown asset type"}), 404
    user = request.api_user
    if not get_role_asset_permissions(user, asset_type).get("can_add", False):
        return jsonify({"error": "Forbidden"}), 403
    definition = ASSET_DEFS[asset_type]
    payload = request.get_json(silent=True) or {}
    data, error = _normalize_asset_payload(definition, payload)
    if error:
        return jsonify({"error": error}), 400
    item = definition["model"](**data)
    db.session.add(item)
    db.session.commit()
    log_audit("create", "asset", entity_id=item.id, details=asset_audit_details(asset_type, item))
    assigned_to = getattr(item, "assigned_to", None)
    if assigned_to and str(assigned_to).strip().lower() not in {"", "free"}:
        log_assignment_change(asset_type, item.id, None, assigned_to, user)
        specs = build_assignment_specs(definition, item)
        send_assignment_email(assigned_to, definition["label"], specs)
    return jsonify({"id": item.id}), 201


@app.route("/api/assets/<asset_type>/<int:item_id>", methods=["PUT"])
@api_auth_required(permission="can_add", asset_type=None)
def api_assets_update(asset_type, item_id):
    if asset_type not in ASSET_DEFS:
        return jsonify({"error": "Unknown asset type"}), 404
    user = request.api_user
    if not get_role_asset_permissions(user, asset_type).get("can_add", False):
        return jsonify({"error": "Forbidden"}), 403
    definition = ASSET_DEFS[asset_type]
    item = definition["model"].query.get_or_404(item_id)
    payload = request.get_json(silent=True) or {}
    data, error = _normalize_asset_payload(definition, payload, existing=item)
    if error:
        return jsonify({"error": error}), 400
    old_values = {field_name: getattr(item, field_name, None) for field_name, _, _ in definition["fields"]}
    for field_name, _label, field_type in definition["fields"]:
        setattr(item, field_name, data.get(field_name))
    db.session.commit()
    log_audit("update", "asset", entity_id=item.id, details=format_changes(old_values, data))
    log_assignment_change(
        asset_type,
        item.id,
        old_values.get("assigned_to"),
        data.get("assigned_to"),
        user,
    )
    return jsonify({"status": "updated"})


@app.route("/api/assets/<asset_type>/<int:item_id>", methods=["DELETE"])
@api_auth_required(permission="can_delete", asset_type=None)
def api_assets_delete(asset_type, item_id):
    if asset_type not in ASSET_DEFS:
        return jsonify({"error": "Unknown asset type"}), 404
    user = request.api_user
    if not get_role_asset_permissions(user, asset_type).get("can_delete", False):
        return jsonify({"error": "Forbidden"}), 403
    definition = ASSET_DEFS[asset_type]
    item = definition["model"].query.get_or_404(item_id)
    details = asset_audit_details(asset_type, item)
    db.session.delete(item)
    db.session.commit()
    log_audit("delete", "asset", entity_id=item_id, details=details)
    return jsonify({"status": "deleted"})


@app.route("/")
@login_required
@require_permission("can_read")
def index():
    user = get_current_user()
    assets_list = []
    total_assets = 0
    total_assigned = 0
    total_available = 0
    for key, definition in ASSET_DEFS.items():
        perms = get_role_asset_permissions(user, key)
        if perms["can_read"]:
            assets_list.append(
                {
                    "key": key,
                    "label": definition["label"],
                    "count": definition["model"].query.count(),
                    "url": url_for("list_assets", asset_type=key),
                }
            )
        if definition.get("consumable"):
            for item in definition["model"].query.all():
                total = max(item.total_quantity or 0, 0)
                assigned = max(item.assigned_quantity or 0, 0)
                available = max(total - assigned, 0)
                total_assets += total
                total_assigned += assigned
                total_available += available
        else:
            available = is_free_filter(definition["model"]).count()
            if hasattr(definition["model"], "status"):
                status_col = func.lower(definition["model"].status)
                assigned = (
                    definition["model"]
                    .query.filter(
                        (definition["model"].assigned_to.is_(None) == False)
                        & (func.trim(definition["model"].assigned_to) != "")
                        & (func.lower(definition["model"].assigned_to) != "free")
                    )
                    .filter(
                        (definition["model"].status.is_(None))
                        | (func.trim(definition["model"].status) == "")
                        | (~status_col.in_(["broken", "write off", "write_off"]))
                    )
                    .count()
                )
            else:
                count = definition["model"].query.count()
                assigned = max(count - available, 0)
            total_assets += assigned + available
            total_assigned += assigned
            total_available += available
    for asset_type in get_custom_asset_types():
        perms = get_role_asset_permissions(user, f"custom:{asset_type.key}")
        if perms["can_read"]:
            assets_list.append(
                {
                    "key": asset_type.key,
                    "label": asset_type.label,
                    "count": AssetItem.query.filter_by(asset_type_id=asset_type.id).count(),
                    "url": url_for("list_custom_assets", asset_key=asset_type.key),
                }
            )
        total_assets += AssetItem.query.filter_by(asset_type_id=asset_type.id).count()
    dept_summary = build_dept_summary()
    return render_template(
        "index.html",
        user=user,
        assets=assets_list,
        total_assets=total_assets,
        total_assigned=total_assigned,
        total_available=total_available,
        dept_summary=dept_summary,
    )


@app.route("/reports")
@login_required
@require_permission("can_read")
def reports():
    dept_filter = (request.args.get("dept") or "all").strip()
    asset_filter = (request.args.get("asset") or "all").strip()
    status_filter = (request.args.get("status") or "all").strip()
    if normalize_status(status_filter) == "available":
        status_filter = "in_stock"
    cache_key = _report_cache_key(dept_filter, asset_filter, status_filter)
    rows = get_cached_report_rows(cache_key)
    if rows is None:
        rows = build_report_rows(
            dept_filter=dept_filter, asset_filter=asset_filter, status_filter=status_filter
        )
        set_cached_report_rows(cache_key, rows)
    return render_template(
        "report.html",
        dept_summary=build_dept_summary(),
        asset_summary=build_report_summary(rows),
        report_rows=rows,
        dept_filter=dept_filter,
        asset_filter=asset_filter,
        status_filter=status_filter,
        asset_options=[{"key": "all", "label": "All Assets"}] + [
            {"key": key, "label": definition["label"]} for key, definition in ASSET_DEFS.items()
        ] + [
            {"key": asset_type.key, "label": asset_type.label} for asset_type in get_custom_asset_types()
        ],
        dept_options=["all", "in_stock"] + [dept.name for dept in Department.query.order_by(Department.name.asc()).all()],
        generated_at=datetime.datetime.now(),
    )


@app.route("/reports/export")
@login_required
@require_permission("can_read")
def export_report():
    dept_filter = (request.args.get("dept") or "all").strip()
    asset_filter = (request.args.get("asset") or "all").strip()
    status_filter = (request.args.get("status") or "all").strip()
    if normalize_status(status_filter) == "available":
        status_filter = "in_stock"
    cache_key = _report_cache_key(dept_filter, asset_filter, status_filter)
    rows = get_cached_report_rows(cache_key)
    if rows is None:
        rows = build_report_rows(
            dept_filter=dept_filter, asset_filter=asset_filter, status_filter=status_filter
        )
        set_cached_report_rows(cache_key, rows)

    headers = [
        "Asset",
        "Asset Tag",
        "Vendor",
        "Model",
        "Processor",
        "RAM",
        "Hard Disk",
        "User",
        "Department",
        "Status",
        "Quantity",
    ]
    lines = [
        "<table>",
        "<thead><tr>",
        *[f"<th>{html.escape(header)}</th>" for header in headers],
        "</tr></thead>",
        "<tbody>",
    ]
    for row in rows:
        values = [
            row.get("asset", ""),
            row.get("asset_tag", ""),
            row.get("vendor", ""),
            row.get("model", ""),
            row.get("processor", ""),
            row.get("ram", ""),
            row.get("hard_disk", ""),
            row.get("user", ""),
            row.get("dept", ""),
            row.get("status", ""),
            str(row.get("quantity") or ""),
        ]
        cells = "".join(f"<td>{html.escape(str(value))}</td>" for value in values)
        lines.append(f"<tr>{cells}</tr>")
    lines.append("</tbody></table>")
    filename = f"asset-report-{datetime.datetime.now().strftime('%Y%m%d-%H%M%S')}.xls"
    resp = app.response_class("\n".join(lines), mimetype="application/vnd.ms-excel")
    resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return resp


@app.route("/assets/<asset_type>")
@login_required
@require_static_permission("can_read")
def list_assets(asset_type):
    user = get_current_user()
    asset_perms = get_role_asset_permissions(user, asset_type)
    definition = ASSET_DEFS[asset_type]
    page = max(_parse_int(request.args.get("page", ""), 1), 1)
    per_page = DEFAULT_PAGE_SIZE
    query_builder = definition["model"].query
    query = normalize_search(request.args.get("q", ""))
    status_filter = (request.args.get("status") or "all").strip()
    query_builder = apply_status_filter_query(definition["model"], query_builder, status_filter)
    if query:
        filters = []
        for field_name, _label, field_type in definition["fields"]:
            if field_name == "connection" and definition["model"] is Mouse:
                if query in {"wired"}:
                    filters.append(definition["model"].wired.is_(True))
                elif query in {"wireless"}:
                    filters.append(definition["model"].wireless.is_(True))
                continue
            column = getattr(definition["model"], field_name)
            if field_type == "checkbox":
                if query in {"yes", "true", "1"}:
                    filters.append(column.is_(True))
                elif query in {"no", "false", "0"}:
                    filters.append(column.is_(False))
            else:
                filters.append(cast(column, String).ilike(f"%{query}%"))
        if filters:
            query_builder = query_builder.filter(or_(*filters))
    total_items = query_builder.count()
    items = (
        query_builder.order_by(definition["model"].id.asc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )
    total_pages = max(math.ceil(total_items / per_page), 1)
    return render_template(
        "list.html",
        user=user,
        asset_type=asset_type,
        definition=definition,
        asset_title=format_asset_title(asset_type, definition["label"]),
        items=items,
        asset_perms=asset_perms,
        query=query,
        status_filter=status_filter,
        import_headers=get_static_import_headers(asset_type),
        page=page,
        total_pages=total_pages,
        total_items=total_items,
    )


@app.route("/assets/<asset_type>/view/<int:item_id>")
@login_required
@require_static_permission("can_read")
def view_asset(asset_type, item_id):
    definition = ASSET_DEFS[asset_type]
    item = definition["model"].query.get_or_404(item_id)
    history = (
        AssetAssignmentHistory.query.filter_by(asset_type=asset_type, asset_id=item.id)
        .order_by(AssetAssignmentHistory.created_at.desc())
        .all()
    )
    edit_history = (
        AuditLog.query.filter_by(entity_type="asset", action="update")
        .filter(AuditLog.entity_id == str(item.id))
        .order_by(AuditLog.created_at.desc())
        .all()
    )
    comments = (
        AssetComment.query.filter_by(asset_type=asset_type, asset_id=item.id)
        .order_by(AssetComment.created_at.desc())
        .all()
    )
    previous_users = []
    seen = set()
    for entry in history:
        if entry.to_user:
            key = normalize_assignee(entry.to_user)
            if key in seen or key in {"", "free"}:
                continue
            seen.add(key)
            previous_users.append(display_assignee(entry.to_user))
        if len(previous_users) >= 2:
            break
    fields = [
        (
            field_name,
            label,
            field_type,
            get_mouse_connection(item)
            if field_name == "connection" and definition["model"] is Mouse
            else getattr(item, field_name, None),
        )
        for field_name, label, field_type in definition["fields"]
        if field_name != "assigned_to"
    ]
    assigned_to = display_assignee(getattr(item, "assigned_to", None))
    return render_template(
        "asset_detail.html",
        asset_type=asset_type,
        asset_title=format_asset_title(asset_type, definition["label"]),
        item=item,
        fields=fields,
        assigned_to=assigned_to,
        history=history,
        edit_history=edit_history,
        comments=comments,
        previous_users=previous_users,
    )


@app.route("/assets/<asset_type>/view/<int:item_id>/comment", methods=["POST"])
@login_required
@require_static_permission("can_read")
def add_asset_comment(asset_type, item_id):
    definition = ASSET_DEFS.get(asset_type)
    if not definition:
        return redirect(url_for("index"))
    item = definition["model"].query.get_or_404(item_id)
    body = (request.form.get("comment") or "").strip()
    if not body:
        flash("Comment cannot be empty.", "error")
        return redirect(url_for("view_asset", asset_type=asset_type, item_id=item_id))
    user = get_current_user()
    comment = AssetComment(
        asset_type=asset_type,
        asset_id=item.id,
        body=body,
        user_id=user.id if user else None,
        username=user.username if user else None,
    )
    db.session.add(comment)
    db.session.commit()
    log_audit("comment", "asset", entity_id=item.id, details=f"type={asset_type}")
    flash("Comment added.", "success")
    return redirect(url_for("view_asset", asset_type=asset_type, item_id=item_id))


@app.route("/assets/<asset_type>/page")
@login_required
@require_static_permission("can_read")
def list_assets_page(asset_type):
    definition = ASSET_DEFS[asset_type]
    page = max(_parse_int(request.args.get("page", ""), 1), 1)
    per_page = DEFAULT_PAGE_SIZE
    query_builder = definition["model"].query
    query = normalize_search(request.args.get("q", ""))
    status_filter = (request.args.get("status") or "all").strip()
    query_builder = apply_status_filter_query(definition["model"], query_builder, status_filter)
    if query:
        filters = []
        for field_name, _label, field_type in definition["fields"]:
            column = getattr(definition["model"], field_name)
            if field_type == "checkbox":
                if query in {"yes", "true", "1"}:
                    filters.append(column.is_(True))
                elif query in {"no", "false", "0"}:
                    filters.append(column.is_(False))
            else:
                filters.append(cast(column, String).ilike(f"%{query}%"))
        if filters:
            query_builder = query_builder.filter(or_(*filters))
    total_items = query_builder.count()
    total_pages = max(math.ceil(total_items / per_page), 1)
    items = (
        query_builder.order_by(definition["model"].id.asc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )
    rows = []
    for item in items:
        row = {"id": item.id, "fields": {}}
        for field_name, _label, field_type in definition["fields"]:
            if field_name == "connection" and definition["model"] is Mouse:
                value = get_mouse_connection(item)
                field_type = "text"
            else:
                value = getattr(item, field_name, None)
            row["fields"][field_name] = format_static_field_value(
                field_name, field_type, value
            )
        rows.append(row)
    return jsonify(
        {
            "rows": rows,
            "page": page,
            "has_more": page < total_pages,
        }
    )


@app.route("/assets/<asset_type>/export")
@login_required
@require_static_permission("can_read")
def export_assets_excel(asset_type):
    definition = ASSET_DEFS[asset_type]
    headers = get_static_import_headers(asset_type)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for item in definition["model"].query.order_by(definition["model"].id.asc()).all():
        row = [item.id]
        for field_name, _label, field_type in definition["fields"]:
            if field_name == "connection" and definition["model"] is Mouse:
                value = get_mouse_connection(item)
                field_type = "text"
            else:
                value = getattr(item, field_name, None)
            if field_name == "assigned_to" and (not value or normalize_assignee(value) == "free"):
                value = ""
            if field_type == "checkbox":
                row.append("Yes" if value else "No")
            else:
                row.append(value if value is not None else "")
        sheet.append(row)
    filename = f"{asset_type}-assets-{datetime.datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/assets/<asset_type>/import", methods=["POST"])
@login_required
@require_static_permission("can_add")
def import_assets_excel(asset_type):
    if "file" not in request.files:
        flash("No file uploaded.", "error")
        return redirect(url_for("list_assets", asset_type=asset_type))
    file = request.files["file"]
    if not file or not file.filename:
        flash("No file selected.", "error")
        return redirect(url_for("list_assets", asset_type=asset_type))
    if not file.filename.lower().endswith(".xlsx"):
        flash("Upload an .xlsx file.", "error")
        return redirect(url_for("list_assets", asset_type=asset_type))
    workbook = load_workbook(file, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        flash("Excel file is empty.", "error")
        return redirect(url_for("list_assets", asset_type=asset_type))
    definition = ASSET_DEFS[asset_type]
    required_headers = {
        _normalize_header(label): label for _name, label, _field_type in definition["fields"]
    }
    provided_headers = {_normalize_header(cell) for cell in rows[0] if cell}
    missing_keys = sorted(set(required_headers.keys()) - provided_headers)
    if missing_keys:
        missing_labels = [required_headers[key] for key in missing_keys]
        flash(f"Missing required headers: {', '.join(missing_labels)}", "error")
        return redirect(url_for("list_assets", asset_type=asset_type))
    header_map = get_static_header_map(asset_type)
    headers = [header_map.get(_normalize_header(cell)) for cell in rows[0]]
    if not any(headers):
        flash("Header row does not match the expected format.", "error")
        return redirect(url_for("list_assets", asset_type=asset_type))
    created = 0
    errors = 0
    for row in rows[1:]:
        if row is None or all(cell in {None, ""} for cell in row):
            continue
        data = {}
        for idx, cell in enumerate(row):
            field_name = headers[idx] if idx < len(headers) else None
            if not field_name or field_name == "id":
                continue
            field_def = next((f for f in definition["fields"] if f[0] == field_name), None)
            field_type = field_def[2] if field_def else "text"
            if field_type == "checkbox":
                data[field_name] = _parse_bool(cell)
            elif field_type == "number":
                data[field_name] = _parse_int_value(cell, 0)
            else:
                data[field_name] = str(cell).strip() if cell is not None else ""
        if "assigned_to" in data and data["assigned_to"] == "":
            data["assigned_to"] = "free"
        if "status" in data:
            status_norm = normalize_status(data["status"])
            assigned_norm = normalize_assignee(data.get("assigned_to"))
            if status_norm in {"broken", "write off"}:
                data["status"] = "Broken" if status_norm == "broken" else "Write Off"
                data["assigned_to"] = "free"
                if "dept" in data:
                    data["dept"] = ""
            elif status_norm in {"in stock", "in_stock"}:
                data["status"] = "In Stock"
                data["assigned_to"] = "free"
                if "dept" in data:
                    data["dept"] = ""
            elif assigned_norm and assigned_norm != "free":
                data["status"] = "Assigned"
            else:
                data["status"] = "In Stock"
        if definition["model"] is Mouse:
            apply_mouse_connection_to_data(data)
        if "asset_tag" in data and data["asset_tag"]:
            existing = definition["model"].query.filter_by(asset_tag=data["asset_tag"]).first()
            if existing:
                errors += 1
                continue
        try:
            item = definition["model"](**data)
            db.session.add(item)
            created += 1
        except Exception:
            db.session.rollback()
            errors += 1
            continue
    db.session.commit()
    if created:
        log_audit("create", "asset_import", details=f"{asset_type} imported: {created}")
    if errors:
        flash(f"Imported {created} rows with {errors} skipped.", "error")
    else:
        flash(f"Imported {created} rows.", "success")
    return redirect(url_for("list_assets", asset_type=asset_type))


@app.route("/assets/<asset_type>/add", methods=["GET", "POST"])
@require_static_permission("can_add")
def add_asset(asset_type):
    definition = ASSET_DEFS[asset_type]
    model = definition["model"]
    user = get_current_user()

    if request.method == "POST":
        data = {}
        bulk_quantity = request.form.get("bulk_quantity", "1")
        for field_name, _, field_type in definition["fields"]:
            if field_name == "connection" and definition["model"] is Mouse:
                data[field_name] = request.form.get(field_name, "").strip()
            elif field_type == "checkbox":
                data[field_name] = field_name in request.form
            elif field_type == "number":
                value = request.form.get(field_name, "").strip()
                data[field_name] = _parse_int(value, 0)
            else:
                value = request.form.get(field_name, "").strip()
                data[field_name] = value
        if definition["model"] is Mouse:
            apply_mouse_connection_to_data(data)
        if "assigned_to" in data and data["assigned_to"] == "":
            data["assigned_to"] = "free"
        if "status" in data:
            status_norm = normalize_status(data["status"])
            assigned_norm = normalize_assignee(data.get("assigned_to"))
            if status_norm in {"broken", "write off"}:
                data["status"] = "Broken" if status_norm == "broken" else "Write Off"
                data["assigned_to"] = "free"
                if "dept" in data:
                    data["dept"] = ""
            elif status_norm in {"in stock", "in_stock"}:
                data["status"] = "In Stock"
                data["assigned_to"] = "free"
                if "dept" in data:
                    data["dept"] = ""
            elif assigned_norm and assigned_norm != "free":
                data["status"] = "Assigned"
            else:
                data["status"] = "In Stock"
        if "asset_tag" in data:
            data["asset_tag"] = data["asset_tag"].strip()
            if data["asset_tag"]:
                existing = (
                    model.query.filter(func.lower(model.asset_tag) == data["asset_tag"].lower())
                    .first()
                )
                if existing:
                    flash("Asset tag must be unique.", "error")
                    return render_template(
                        "add.html",
                        asset_type=asset_type,
                        definition=definition,
                        form_data=data,
                        bulk_quantity=bulk_quantity,
                        invalid_fields={"asset_tag"},
                    )
        if "total_quantity" in data and "assigned_quantity" in data:
            if data["assigned_quantity"] < 0 or data["total_quantity"] < 0:
                flash("Quantities cannot be negative.", "error")
                return render_template(
                    "add.html",
                    asset_type=asset_type,
                    definition=definition,
                    form_data=data,
                    bulk_quantity=bulk_quantity,
                )
            if data["assigned_quantity"] > data["total_quantity"]:
                flash("Assigned quantity cannot exceed total quantity.", "error")
                return render_template(
                    "add.html",
                    asset_type=asset_type,
                    definition=definition,
                    form_data=data,
                    bulk_quantity=bulk_quantity,
                )
        bulk_count = 1
        if definition.get("bulk_add"):
            bulk_count = max(_parse_int(bulk_quantity, 1), 1)
        if bulk_count > 1 and data.get("asset_tag"):
            flash("Bulk add cannot reuse a single asset tag. Leave asset tag blank or add one item at a time.", "error")
            return render_template(
                "add.html",
                asset_type=asset_type,
                definition=definition,
                form_data=data,
                bulk_quantity=bulk_quantity,
                invalid_fields={"asset_tag"},
            )
        created_items = []
        for _ in range(bulk_count):
            item = model(**data)
            db.session.add(item)
            created_items.append(item)
        db.session.commit()
        for item in created_items:
            log_audit("create", "asset", entity_id=item.id, details=asset_audit_details(asset_type, item))
            assigned_to = getattr(item, "assigned_to", None)
            if assigned_to and str(assigned_to).strip().lower() not in {"", "free"}:
                log_assignment_change(asset_type, item.id, None, assigned_to, user)
                specs = build_assignment_specs(definition, item)
                send_assignment_email(assigned_to, definition["label"], specs)
        if bulk_count > 1:
            flash(f"Added {bulk_count} {definition['label']} assets.", "success")
        else:
            flash(f"{definition['label']} asset added.", "success")
        return redirect(url_for("list_assets", asset_type=asset_type))

    return render_template(
        "add.html",
        asset_type=asset_type,
        definition=definition,
    )


@app.route("/assets/<asset_type>/delete/<int:item_id>", methods=["POST"])
@require_static_permission("can_delete")
def delete_asset(asset_type, item_id):
    definition = ASSET_DEFS[asset_type]
    item = definition["model"].query.get_or_404(item_id)
    details = asset_audit_details(asset_type, item)
    db.session.delete(item)
    db.session.commit()
    log_audit("delete", "asset", entity_id=item_id, details=details)
    return redirect(url_for("list_assets", asset_type=asset_type))


@app.route("/assets/<asset_type>/bulk-delete", methods=["POST"])
@require_static_permission("can_bulk_delete")
def bulk_delete_assets(asset_type):
    definition = ASSET_DEFS[asset_type]
    ids = [int(item_id) for item_id in request.form.getlist("selected_ids") if item_id.isdigit()]
    tag_input = request.form.get("bulk_tags", "")
    tag_list = parse_bulk_tags(tag_input)
    if tag_list:
        if not hasattr(definition["model"], "asset_tag"):
            flash("This asset type does not support asset tag deletes.", "error")
            return redirect(url_for("list_assets", asset_type=asset_type))
        tag_items = definition["model"].query.filter(
            definition["model"].asset_tag.in_(tag_list)
        ).all()
        ids.extend(item.id for item in tag_items)
    ids = sorted(set(ids))
    if ids:
        items = definition["model"].query.filter(definition["model"].id.in_(ids)).all()
        tags = [
            getattr(item, "asset_tag", None)
            for item in items
            if getattr(item, "asset_tag", None)
        ]
        definition["model"].query.filter(definition["model"].id.in_(ids)).delete(
            synchronize_session=False
        )
        db.session.commit()
        if tags:
            detail_text = f"type={asset_type} asset_tags={tags}"
        else:
            detail_text = f"type={asset_type} ids={ids}"
        log_audit("bulk_delete", "asset", entity_id=asset_type, details=detail_text)
        flash(f"Deleted {len(ids)} items.", "success")
    else:
        flash("No items selected.", "error")
    return redirect(url_for("list_assets", asset_type=asset_type))


@app.route("/assets/<asset_type>/edit/<int:item_id>", methods=["GET", "POST"])
@login_required
@require_static_permission("can_add")
def edit_asset(asset_type, item_id):
    definition = ASSET_DEFS[asset_type]
    item = definition["model"].query.get_or_404(item_id)
    user = get_current_user()
    old_values = {}
    for field_name, _, _ in definition["fields"]:
        if field_name == "connection" and definition["model"] is Mouse:
            old_values[field_name] = get_mouse_connection(item)
        else:
            old_values[field_name] = getattr(item, field_name, None)
    if request.method == "POST":
        connection_value = None
        for field_name, _, field_type in definition["fields"]:
            if field_name == "connection" and definition["model"] is Mouse:
                connection_value = request.form.get(field_name, "").strip()
                continue
            if field_type == "checkbox":
                setattr(item, field_name, field_name in request.form)
            elif field_type == "number":
                value = request.form.get(field_name, "").strip()
                setattr(item, field_name, _parse_int(value, 0))
            else:
                value = request.form.get(field_name, "").strip()
                if field_name == "assigned_to" and value == "":
                    value = "free"
                setattr(item, field_name, value)
        if definition["model"] is Mouse and connection_value is not None:
            conn_norm = normalize_connection(connection_value)
            item.wired = conn_norm == "wired"
            item.wireless = conn_norm == "wireless"
        if hasattr(item, "status"):
            status_norm = normalize_status(getattr(item, "status", ""))
            assigned_norm = normalize_assignee(getattr(item, "assigned_to", ""))
            if status_norm in {"broken", "write off"}:
                item.status = "Broken" if status_norm == "broken" else "Write Off"
                item.assigned_to = "free"
                if hasattr(item, "dept"):
                    item.dept = ""
            elif status_norm in {"in stock", "in_stock"}:
                item.status = "In Stock"
                item.assigned_to = "free"
                if hasattr(item, "dept"):
                    item.dept = ""
            elif assigned_norm and assigned_norm != "free":
                item.status = "Assigned"
            else:
                item.status = "In Stock"
        if hasattr(item, "asset_tag"):
            asset_tag_value = (getattr(item, "asset_tag", "") or "").strip()
            if asset_tag_value:
                existing = (
                    definition["model"]
                    .query.filter(func.lower(definition["model"].asset_tag) == asset_tag_value.lower())
                    .filter(definition["model"].id != item.id)
                    .first()
                )
                if existing:
                    flash("Asset tag must be unique.", "error")
                    return render_template(
                        "edit.html",
                        asset_type=asset_type,
                        definition=definition,
                        item=item,
                        invalid_fields={"asset_tag"},
                    )
        if hasattr(item, "total_quantity") and hasattr(item, "assigned_quantity"):
            if item.assigned_quantity < 0 or item.total_quantity < 0:
                flash("Quantities cannot be negative.", "error")
                return render_template(
                    "edit.html",
                    asset_type=asset_type,
                    definition=definition,
                    item=item,
                )
            if item.assigned_quantity > item.total_quantity:
                flash("Assigned quantity cannot exceed total quantity.", "error")
                return render_template(
                    "edit.html",
                    asset_type=asset_type,
                    definition=definition,
                    item=item,
                )
        db.session.commit()
        new_values = {}
        for field_name, _, _ in definition["fields"]:
            if field_name == "connection" and definition["model"] is Mouse:
                new_values[field_name] = get_mouse_connection(item)
            else:
                new_values[field_name] = getattr(item, field_name, None)
        change_details = format_changes(old_values, new_values)
        details = asset_audit_details(asset_type, item)
        if change_details:
            details = f"{details} changes={change_details}"
        log_audit("update", "asset", entity_id=item.id, details=details)
        old_assigned = str(old_values.get("assigned_to") or "").strip().lower()
        new_assigned_raw = str(new_values.get("assigned_to") or "").strip()
        new_assigned = new_assigned_raw.lower()
        if new_assigned != old_assigned:
            log_assignment_change(asset_type, item.id, old_values.get("assigned_to"), new_assigned_raw, user)
        if new_assigned and new_assigned not in {"", "free"} and new_assigned != old_assigned:
            specs = build_assignment_specs(definition, item)
            send_assignment_email(new_assigned_raw, definition["label"], specs)
        return redirect(url_for("list_assets", asset_type=asset_type))
    return render_template(
        "edit.html",
        asset_type=asset_type,
        definition=definition,
        item=item,
    )


@app.route("/assets/<asset_type>/copy/<int:item_id>", methods=["GET", "POST"])
@login_required
@require_static_permission("can_add")
def copy_asset(asset_type, item_id):
    definition = ASSET_DEFS[asset_type]
    item = definition["model"].query.get_or_404(item_id)
    data = {}
    for field_name, _, _ in definition["fields"]:
        if field_name == "connection" and definition["model"] is Mouse:
            data[field_name] = get_mouse_connection(item)
        else:
            data[field_name] = getattr(item, field_name, None)
    if "asset_tag" in data:
        data["asset_tag"] = ""
    return render_template(
        "add.html",
        asset_type=asset_type,
        definition=definition,
        form_data=data,
        bulk_quantity=1,
        form_action=url_for("add_asset", asset_type=asset_type),
    )


@app.route("/custom/<asset_key>")
@login_required
@require_custom_permission("can_read")
def list_custom_assets(asset_key):
    asset_type = get_custom_asset_by_key(asset_key)
    if not asset_type:
        return redirect(url_for("index"))
    user = get_current_user()
    asset_perms = get_role_asset_permissions(user, f"custom:{asset_key}")
    fields = AssetField.query.filter_by(asset_type_id=asset_type.id).all()
    assigned_fields, _ = get_custom_special_fields(fields)
    items = AssetItem.query.filter_by(asset_type_id=asset_type.id).all()
    query = normalize_search(request.args.get("q", ""))
    status_filter = (request.args.get("status") or "all").strip()
    if query:
        items = [
            item
            for item in items
            if matches_query(
                query,
                *[(item.data or {}).get(field.name) for field in fields],
            )
        ]
    if status_filter and normalize_status(status_filter) not in {"", "all"}:
        desired = normalize_status(status_filter)
        if desired == "available":
            desired = "in stock"
        filtered = []
        for item in items:
            assigned_to = ""
            for field_name in assigned_fields:
                assigned_to = (item.data or {}).get(field_name)
                if assigned_to is not None:
                    break
            status = get_custom_status(item.data or {}, assigned_to)
            if status == desired:
                filtered.append(item)
        items = filtered
    page = max(_parse_int(request.args.get("page", ""), 1), 1)
    per_page = DEFAULT_PAGE_SIZE
    total_items = len(items)
    total_pages = max(math.ceil(total_items / per_page), 1)
    start = (page - 1) * per_page
    end = start + per_page
    items = items[start:end]
    return render_template(
        "custom_list.html",
        asset_type=asset_type,
        asset_title=format_asset_title(asset_type.key, asset_type.label),
        fields=fields,
        fields_json=[{"name": field.name, "label": field.label, "field_type": field.field_type} for field in fields],
        items=items,
        asset_perms=asset_perms,
        assigned_fields=assigned_fields,
        query=query,
        status_filter=status_filter,
        import_headers=get_custom_import_headers(fields),
        page=page,
        total_pages=total_pages,
        total_items=total_items,
    )


@app.route("/custom/<asset_key>/view/<int:item_id>")
@login_required
@require_custom_permission("can_read")
def view_custom_asset(asset_key, item_id):
    asset_type = get_custom_asset_by_key(asset_key)
    if not asset_type:
        return redirect(url_for("index"))
    item = AssetItem.query.filter_by(asset_type_id=asset_type.id, id=item_id).first_or_404()
    fields = AssetField.query.filter_by(asset_type_id=asset_type.id).all()
    assigned_fields, _ = get_custom_special_fields(fields)
    history = (
        AssetAssignmentHistory.query.filter_by(
            asset_type=f"custom:{asset_key}", asset_id=item.id
        )
        .order_by(AssetAssignmentHistory.created_at.desc())
        .all()
    )
    edit_history = (
        AuditLog.query.filter_by(entity_type="custom_asset", action="update")
        .filter(AuditLog.entity_id == str(item.id))
        .order_by(AuditLog.created_at.desc())
        .all()
    )
    comments = (
        AssetComment.query.filter_by(asset_type=f"custom:{asset_key}", asset_id=item.id)
        .order_by(AssetComment.created_at.desc())
        .all()
    )
    previous_users = []
    seen = set()
    for entry in history:
        if entry.to_user:
            key = normalize_assignee(entry.to_user)
            if key in seen or key in {"", "free"}:
                continue
            seen.add(key)
            previous_users.append(display_assignee(entry.to_user))
        if len(previous_users) >= 2:
            break
    details = []
    for field in fields:
        if field.name in assigned_fields:
            continue
        value = (item.data or {}).get(field.name)
        details.append((field.label, format_custom_field_value(field, value, assigned_fields)))
    assigned_to = "-"
    for field_name in assigned_fields:
        value = (item.data or {}).get(field_name)
        if value is not None:
            assigned_to = display_assignee(value)
            break
    return render_template(
        "custom_asset_detail.html",
        asset_type=asset_type,
        asset_title=format_asset_title(asset_type.key, asset_type.label),
        item=item,
        fields=details,
        assigned_to=assigned_to,
        history=history,
        edit_history=edit_history,
        comments=comments,
        previous_users=previous_users,
    )


@app.route("/custom/<asset_key>/view/<int:item_id>/comment", methods=["POST"])
@login_required
@require_custom_permission("can_read")
def add_custom_asset_comment(asset_key, item_id):
    asset_type = get_custom_asset_by_key(asset_key)
    if not asset_type:
        return redirect(url_for("index"))
    item = AssetItem.query.filter_by(asset_type_id=asset_type.id, id=item_id).first_or_404()
    body = (request.form.get("comment") or "").strip()
    if not body:
        flash("Comment cannot be empty.", "error")
        return redirect(url_for("view_custom_asset", asset_key=asset_key, item_id=item_id))
    user = get_current_user()
    comment = AssetComment(
        asset_type=f"custom:{asset_key}",
        asset_id=item.id,
        body=body,
        user_id=user.id if user else None,
        username=user.username if user else None,
    )
    db.session.add(comment)
    db.session.commit()
    log_audit("comment", "custom_asset", entity_id=item.id, details=f"type={asset_key}")
    flash("Comment added.", "success")
    return redirect(url_for("view_custom_asset", asset_key=asset_key, item_id=item_id))


@app.route("/custom/<asset_key>/page")
@login_required
@require_custom_permission("can_read")
def list_custom_assets_page(asset_key):
    asset_type = get_custom_asset_by_key(asset_key)
    if not asset_type:
        return jsonify({"rows": [], "page": 1, "has_more": False})
    fields = AssetField.query.filter_by(asset_type_id=asset_type.id).all()
    assigned_fields, _ = get_custom_special_fields(fields)
    items = AssetItem.query.filter_by(asset_type_id=asset_type.id).all()
    query = normalize_search(request.args.get("q", ""))
    status_filter = (request.args.get("status") or "all").strip()
    if query:
        items = [
            item
            for item in items
            if matches_query(
                query,
                *[(item.data or {}).get(field.name) for field in fields],
            )
        ]
    if status_filter and normalize_status(status_filter) not in {"", "all"}:
        desired = normalize_status(status_filter)
        if desired == "available":
            desired = "in stock"
        filtered = []
        for item in items:
            assigned_to = ""
            for field_name in assigned_fields:
                assigned_to = (item.data or {}).get(field_name)
                if assigned_to is not None:
                    break
            status = get_custom_status(item.data or {}, assigned_to)
            if status == desired:
                filtered.append(item)
        items = filtered
    page = max(_parse_int(request.args.get("page", ""), 1), 1)
    per_page = DEFAULT_PAGE_SIZE
    total_items = len(items)
    total_pages = max(math.ceil(total_items / per_page), 1)
    start = (page - 1) * per_page
    end = start + per_page
    items = items[start:end]
    rows = []
    for item in items:
        row = {"id": item.id, "fields": {}}
        data = item.data or {}
        for field in fields:
            row["fields"][field.name] = format_custom_field_value(
                field, data.get(field.name), assigned_fields
            )
        rows.append(row)
    return jsonify(
        {
            "rows": rows,
            "page": page,
            "has_more": page < total_pages,
        }
    )


@app.route("/custom/<asset_key>/export")
@login_required
@require_custom_permission("can_read")
def export_custom_assets_excel(asset_key):
    asset_type = get_custom_asset_by_key(asset_key)
    if not asset_type:
        return redirect(url_for("index"))
    fields = AssetField.query.filter_by(asset_type_id=asset_type.id).all()
    headers = get_custom_import_headers(fields)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    items = AssetItem.query.filter_by(asset_type_id=asset_type.id).order_by(AssetItem.id.asc()).all()
    for item in items:
        row = [item.id]
        data = item.data or {}
        for field in fields:
            value = data.get(field.name)
            if field.field_type == "checkbox" and field.options:
                row.append(", ".join(value) if isinstance(value, list) else (value or ""))
            elif field.field_type == "checkbox":
                row.append("Yes" if value else "No")
            else:
                row.append(value if value is not None else "")
        sheet.append(row)
    filename = f"custom-{asset_key}-assets-{datetime.datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/custom/<asset_key>/import", methods=["POST"])
@login_required
@require_custom_permission("can_add")
def import_custom_assets_excel(asset_key):
    asset_type = get_custom_asset_by_key(asset_key)
    if not asset_type:
        return redirect(url_for("index"))
    if "file" not in request.files:
        flash("No file uploaded.", "error")
        return redirect(url_for("list_custom_assets", asset_key=asset_key))
    file = request.files["file"]
    if not file or not file.filename:
        flash("No file selected.", "error")
        return redirect(url_for("list_custom_assets", asset_key=asset_key))
    if not file.filename.lower().endswith(".xlsx"):
        flash("Upload an .xlsx file.", "error")
        return redirect(url_for("list_custom_assets", asset_key=asset_key))
    workbook = load_workbook(file, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        flash("Excel file is empty.", "error")
        return redirect(url_for("list_custom_assets", asset_key=asset_key))
    fields = AssetField.query.filter_by(asset_type_id=asset_type.id).all()
    required_headers = {_normalize_header(field.label): field.label for field in fields}
    provided_headers = {_normalize_header(cell) for cell in rows[0] if cell}
    missing_keys = sorted(set(required_headers.keys()) - provided_headers)
    if missing_keys:
        missing_labels = [required_headers[key] for key in missing_keys]
        flash(f"Missing required headers: {', '.join(missing_labels)}", "error")
        return redirect(url_for("list_custom_assets", asset_key=asset_key))
    header_map = get_custom_header_map(fields)
    headers = [header_map.get(_normalize_header(cell)) for cell in rows[0]]
    if not any(headers):
        flash("Header row does not match the expected format.", "error")
        return redirect(url_for("list_custom_assets", asset_key=asset_key))
    assigned_fields, _ = get_custom_special_fields(fields)
    asset_tag_fields = [field.name for field in fields if field.name.strip().lower() == "asset_tag"]
    existing_tags = set()
    if asset_tag_fields:
        for item in AssetItem.query.filter_by(asset_type_id=asset_type.id).all():
            data = item.data or {}
            for tag_field in asset_tag_fields:
                tag_value = data.get(tag_field)
                if tag_value:
                    existing_tags.add(str(tag_value).strip().lower())
    created = 0
    errors = 0
    for row in rows[1:]:
        if row is None or all(cell in {None, ""} for cell in row):
            continue
        data = {}
        for idx, cell in enumerate(row):
            field_name = headers[idx] if idx < len(headers) else None
            if not field_name or field_name == "id":
                continue
            field_def = next((f for f in fields if f.name == field_name), None)
            if not field_def:
                continue
            if field_def.field_type == "checkbox":
                if field_def.options:
                    if cell is None or cell == "":
                        data[field_def.name] = []
                    else:
                        data[field_def.name] = [part.strip() for part in str(cell).split(",") if part.strip()]
                else:
                    data[field_def.name] = _parse_bool(cell)
            elif field_def.field_type == "number":
                data[field_def.name] = _parse_int_value(cell, 0)
            else:
                data[field_def.name] = str(cell).strip() if cell is not None else ""
        for field_name in assigned_fields:
            if data.get(field_name, "") == "":
                data[field_name] = "free"
        if "status" in data:
            status_norm = normalize_status(data.get("status"))
            assigned_norm = normalize_assignee(data.get("assigned_to"))
            if status_norm in {"broken", "write off"}:
                data["status"] = "Broken" if status_norm == "broken" else "Write Off"
                data["assigned_to"] = "free"
                if "dept" in data:
                    data["dept"] = ""
            elif status_norm in {"in stock", "in_stock"}:
                data["status"] = "In Stock"
                data["assigned_to"] = "free"
                if "dept" in data:
                    data["dept"] = ""
            elif assigned_norm and assigned_norm != "free":
                data["status"] = "Assigned"
            else:
                data["status"] = "In Stock"
        if asset_tag_fields:
            duplicate = False
            for tag_field in asset_tag_fields:
                tag_value = data.get(tag_field)
                if tag_value and str(tag_value).strip().lower() in existing_tags:
                    errors += 1
                    duplicate = True
                    break
            if duplicate:
                continue
        try:
            item = AssetItem(asset_type_id=asset_type.id, data=data)
            db.session.add(item)
            created += 1
        except Exception:
            db.session.rollback()
            errors += 1
            continue
    db.session.commit()
    if created:
        log_audit("create", "custom_asset_import", details=f"{asset_key} imported: {created}")
    if errors:
        flash(f"Imported {created} rows with {errors} skipped.", "error")
    else:
        flash(f"Imported {created} rows.", "success")
    return redirect(url_for("list_custom_assets", asset_key=asset_key))


@app.route("/custom/<asset_key>/add", methods=["GET", "POST"])
@login_required
@require_custom_permission("can_add")
def add_custom_asset(asset_key):
    asset_type = get_custom_asset_by_key(asset_key)
    if not asset_type:
        return redirect(url_for("index"))
    fields = AssetField.query.filter_by(asset_type_id=asset_type.id).all()
    assigned_fields, _ = get_custom_special_fields(fields)
    user = get_current_user()
    if request.method == "POST":
        data = {}
        for field in fields:
            if field.field_type == "checkbox":
                if field.options:
                    data[field.name] = request.form.getlist(field.name)
                else:
                    data[field.name] = field.name in request.form
            else:
                data[field.name] = request.form.get(field.name, "").strip()
        for field_name in assigned_fields:
            if data.get(field_name, "") == "":
                data[field_name] = "free"
        if "status" in data:
            status_norm = normalize_status(data.get("status"))
            assigned_norm = normalize_assignee(data.get("assigned_to"))
            if status_norm in {"broken", "write off"}:
                data["status"] = "Broken" if status_norm == "broken" else "Write Off"
                data["assigned_to"] = "free"
                if "dept" in data:
                    data["dept"] = ""
            elif status_norm in {"in stock", "in_stock"}:
                data["status"] = "In Stock"
                data["assigned_to"] = "free"
                if "dept" in data:
                    data["dept"] = ""
            elif assigned_norm and assigned_norm != "free":
                data["status"] = "Assigned"
            else:
                data["status"] = "In Stock"
        item = AssetItem(asset_type_id=asset_type.id, data=data)
        db.session.add(item)
        db.session.commit()
        log_audit("create", "custom_asset", entity_id=item.id, details=asset_key)
        assigned_user = None
        for field_name in assigned_fields:
            assigned_user = data.get(field_name)
            if assigned_user:
                break
        if assigned_user and str(assigned_user).strip().lower() not in {"", "free"}:
            log_assignment_change(f"custom:{asset_key}", item.id, None, assigned_user, user)
            specs = []
            for field in fields:
                if field.name in assigned_fields:
                    continue
                value = data.get(field.name)
                if field.field_type == "checkbox":
                    if field.options:
                        value = ", ".join(value) if value else "-"
                    else:
                        value = "Yes" if value else "No"
                if value not in (None, ""):
                    specs.append((field.label, value))
            send_assignment_email(assigned_user, asset_type.label, specs)
        return redirect(url_for("list_custom_assets", asset_key=asset_key))
    return render_template(
        "custom_add.html",
        asset_type=asset_type,
        fields=fields,
        assigned_fields=assigned_fields,
    )


@app.route("/custom/<asset_key>/delete/<int:item_id>", methods=["POST"])
@login_required
@require_custom_permission("can_delete")
def delete_custom_asset(asset_key, item_id):
    asset_type = get_custom_asset_by_key(asset_key)
    if not asset_type:
        return redirect(url_for("index"))
    item = AssetItem.query.filter_by(asset_type_id=asset_type.id, id=item_id).first_or_404()
    db.session.delete(item)
    db.session.commit()
    log_audit("delete", "custom_asset", entity_id=item_id, details=asset_key)
    return redirect(url_for("list_custom_assets", asset_key=asset_key))


@app.route("/custom/<asset_key>/bulk-delete", methods=["POST"])
@login_required
@require_custom_permission("can_bulk_delete")
def bulk_delete_custom_assets(asset_key):
    asset_type = get_custom_asset_by_key(asset_key)
    if not asset_type:
        return redirect(url_for("index"))
    ids = [int(item_id) for item_id in request.form.getlist("selected_ids") if item_id.isdigit()]
    tag_input = request.form.get("bulk_tags", "")
    tag_list = parse_bulk_tags(tag_input)
    if tag_list:
        candidates = AssetItem.query.filter_by(asset_type_id=asset_type.id).all()
        for item in candidates:
            data = item.data or {}
            if data.get("asset_tag") in tag_list:
                ids.append(item.id)
    ids = sorted(set(ids))
    if ids:
        items = AssetItem.query.filter(
            AssetItem.asset_type_id == asset_type.id, AssetItem.id.in_(ids)
        ).all()
        tags = []
        for item in items:
            data = item.data or {}
            tag = data.get("asset_tag")
            if tag:
                tags.append(tag)
        AssetItem.query.filter(
            AssetItem.asset_type_id == asset_type.id, AssetItem.id.in_(ids)
        ).delete(synchronize_session=False)
        db.session.commit()
        if tags:
            detail_text = f"type={asset_key} asset_tags={tags}"
        else:
            detail_text = f"type={asset_key} ids={ids}"
        log_audit(
            "bulk_delete",
            "custom_asset",
            entity_id=asset_key,
            details=detail_text,
        )
        flash(f"Deleted {len(ids)} items.", "success")
    else:
        flash("No items selected.", "error")
    return redirect(url_for("list_custom_assets", asset_key=asset_key))


@app.route("/custom/<asset_key>/edit/<int:item_id>", methods=["GET", "POST"])
@login_required
@require_custom_permission("can_add")
def edit_custom_asset(asset_key, item_id):
    asset_type = get_custom_asset_by_key(asset_key)
    if not asset_type:
        return redirect(url_for("index"))
    fields = AssetField.query.filter_by(asset_type_id=asset_type.id).all()
    assigned_fields, _ = get_custom_special_fields(fields)
    item = AssetItem.query.filter_by(asset_type_id=asset_type.id, id=item_id).first_or_404()
    user = get_current_user()
    old_values = dict(item.data or {})
    if request.method == "POST":
        data = {}
        for field in fields:
            if field.field_type == "checkbox":
                if field.options:
                    data[field.name] = request.form.getlist(field.name)
                else:
                    data[field.name] = field.name in request.form
            else:
                data[field.name] = request.form.get(field.name, "").strip()
        for field_name in assigned_fields:
            if data.get(field_name, "") == "":
                data[field_name] = "free"
        if "status" in data:
            status_norm = normalize_status(data.get("status"))
            assigned_norm = normalize_assignee(data.get("assigned_to"))
            if status_norm in {"broken", "write off"}:
                data["status"] = "Broken" if status_norm == "broken" else "Write Off"
                data["assigned_to"] = "free"
                if "dept" in data:
                    data["dept"] = ""
            elif status_norm in {"in stock", "in_stock"}:
                data["status"] = "In Stock"
                data["assigned_to"] = "free"
                if "dept" in data:
                    data["dept"] = ""
            elif assigned_norm and assigned_norm != "free":
                data["status"] = "Assigned"
            else:
                data["status"] = "In Stock"
        old_assigned_value = None
        for field_name in assigned_fields:
            old_assigned_value = (item.data or {}).get(field_name)
            if old_assigned_value:
                break
        item.data = data
        db.session.commit()
        change_details = format_changes(old_values, data)
        details = f"type={asset_key}"
        if change_details:
            details = f"{details} changes={change_details}"
        log_audit("update", "custom_asset", entity_id=item.id, details=details)
        new_assigned_value = None
        for field_name in assigned_fields:
            new_assigned_value = data.get(field_name)
            if new_assigned_value:
                break
        if new_assigned_value != old_assigned_value:
            log_assignment_change(
                f"custom:{asset_key}",
                item.id,
                old_assigned_value,
                new_assigned_value,
                user,
            )
        if new_assigned_value:
            old_text = str(old_assigned_value or "").strip().lower()
            new_text = str(new_assigned_value or "").strip().lower()
            if new_text and new_text not in {"", "free"} and new_text != old_text:
                specs = []
                for field in fields:
                    if field.name in assigned_fields:
                        continue
                    value = data.get(field.name)
                    if field.field_type == "checkbox":
                        if field.options:
                            value = ", ".join(value) if value else "-"
                        else:
                            value = "Yes" if value else "No"
                    if value not in (None, ""):
                        specs.append((field.label, value))
                send_assignment_email(new_assigned_value, asset_type.label, specs)
        return redirect(url_for("list_custom_assets", asset_key=asset_key))
    return render_template(
        "custom_edit.html",
        asset_type=asset_type,
        fields=fields,
        item=item,
        assigned_fields=assigned_fields,
    )


@app.route("/asset-types")
@login_required
@require_app_admin
def list_asset_types():
    asset_types = get_custom_asset_types()
    builtin_types = get_builtin_asset_types()
    return render_template("asset_types.html", asset_types=asset_types, builtin_types=builtin_types)


@app.route("/asset-types/builtin/<asset_key>/edit", methods=["GET", "POST"])
@login_required
@require_app_admin
def edit_builtin_asset_type(asset_key):
    if asset_key not in ASSET_DEFS:
        flash("Unknown asset type.", "error")
        return redirect(url_for("list_asset_types"))
    definition = ASSET_DEFS[asset_key]
    setting = BuiltinAssetTypeSetting.query.filter_by(key=asset_key).first()
    if not setting:
        setting = BuiltinAssetTypeSetting(key=asset_key, label=definition["label"])
        db.session.add(setting)
        db.session.commit()
    if request.method == "POST":
        label = request.form.get("label", "").strip()
        if not label:
            flash("Label is required.", "error")
            return redirect(url_for("edit_builtin_asset_type", asset_key=asset_key))
        setting.label = label
        for field_name, _, _ in definition.get("fields", []):
            field_label = request.form.get(f"field_label_{field_name}", "").strip()
            if not field_label:
                field_label = field_name.replace("_", " ").title()
            field_options = request.form.get(f"field_options_{field_name}", "").strip()
            field_setting = BuiltinAssetFieldSetting.query.filter_by(
                asset_key=asset_key, field_name=field_name
            ).first()
            if not field_setting:
                field_setting = BuiltinAssetFieldSetting(
                    asset_key=asset_key, field_name=field_name, label=field_label, options=field_options
                )
                db.session.add(field_setting)
            else:
                field_setting.label = field_label
                field_setting.options = field_options
        db.session.commit()
        apply_builtin_overrides()
        log_audit("update", "asset_type", entity_id=asset_key, details=f"type=builtin label={label}")
        flash("Built-in asset type updated.", "success")
        return redirect(url_for("list_asset_types"))
    field_rows = []
    for field_name, label, _ in definition.get("fields", []):
        field_setting = BuiltinAssetFieldSetting.query.filter_by(
            asset_key=asset_key, field_name=field_name
        ).first()
        field_rows.append(
            {
                "name": field_name,
                "label": field_setting.label if field_setting else label,
                "options": field_setting.options if field_setting else "",
            }
        )
    return render_template(
        "asset_type_builtin_edit.html",
        asset_key=asset_key,
        label=setting.label,
        fields=field_rows,
    )


@app.route("/asset-types/add", methods=["GET", "POST"])
@login_required
@require_app_admin
def add_asset_type():
    if request.method == "POST":
        label = request.form.get("label", "").strip()
        key_input = request.form.get("key", "").strip()
        key = slugify_key(key_input or label)
        if not label:
            flash("Asset type label is required.", "error")
            return redirect(url_for("add_asset_type"))
        if not key:
            flash("Asset type key is required.", "error")
            return redirect(url_for("add_asset_type"))
        if key in ASSET_DEFS:
            flash("Asset key conflicts with a built-in asset type.", "error")
            return redirect(url_for("add_asset_type"))
        if AssetType.query.filter_by(key=key).first():
            flash("Asset key already exists.", "error")
            return redirect(url_for("add_asset_type"))
        field_rows = []
        for idx in range(1, 11):
            name = request.form.get(f"field_name_{idx}", "").strip()
            label_field = request.form.get(f"field_label_{idx}", "").strip()
            field_type = request.form.get(f"field_type_{idx}", "text").strip()
            if not name and not label_field:
                continue
            if not name:
                name = slugify_key(label_field)
            options = []
            if field_type in {"checkbox", "select"}:
                for opt_idx in range(1, 6):
                    opt_value = request.form.get(
                        f"field_options_{idx}_{opt_idx}", ""
                    ).strip()
                    if opt_value:
                        options.append(opt_value)
            field_rows.append(
                (name, label_field or name.replace("_", " ").title(), field_type, options)
            )
        seen = set()
        for name, _, _, _ in field_rows:
            if name in seen:
                flash("Field names must be unique.", "error")
                return redirect(url_for("add_asset_type"))
            seen.add(name)
        default_fields = [
            ("assigned_to", "User", "text", []),
            ("dept", "Dept", "text", []),
            ("status", "Status", "select", list(STATUS_OPTIONS)),
        ]
        for name, label_field, field_type, options in default_fields:
            if name not in seen:
                field_rows.append((name, label_field, field_type, options))
                seen.add(name)
        if not field_rows:
            flash("At least one field is required.", "error")
            return redirect(url_for("add_asset_type"))
        asset_type = AssetType(key=key, label=label)
        db.session.add(asset_type)
        db.session.flush()
        for name, label_field, field_type, options in field_rows:
            db.session.add(
                AssetField(
                    asset_type_id=asset_type.id,
                    name=name,
                    label=label_field,
                    field_type=field_type,
                    options=options,
                )
            )
        db.session.commit()
        ensure_role_permissions()
        log_audit("create", "asset_type", entity_id=asset_type.id, details=key)
        return redirect(url_for("list_asset_types"))
    return render_template("asset_type_add.html")


@app.route("/asset-types/edit/<int:asset_type_id>", methods=["GET", "POST"])
@login_required
@require_app_admin
def edit_asset_type(asset_type_id):
    asset_type = AssetType.query.get_or_404(asset_type_id)
    fields = AssetField.query.filter_by(asset_type_id=asset_type.id).all()
    old_values = {"label": asset_type.label}
    if request.method == "POST":
        label = request.form.get("label", "").strip()
        if not label:
            flash("Asset type label is required.", "error")
            return redirect(url_for("edit_asset_type", asset_type_id=asset_type.id))
        asset_type.label = label
        for field in fields:
            field.label = request.form.get(f"field_label_{field.id}", field.label).strip()
            field_type = request.form.get(f"field_type_{field.id}", field.field_type)
            field.field_type = field_type
            options = []
            if field_type in {"checkbox", "select"}:
                for opt_idx in range(1, 6):
                    opt_value = request.form.get(
                        f"field_options_{field.id}_{opt_idx}", ""
                    ).strip()
                    if opt_value:
                        options.append(opt_value)
            field.options = options
        new_fields = []
        for idx in range(1, 6):
            name = request.form.get(f"new_field_name_{idx}", "").strip()
            label_field = request.form.get(f"new_field_label_{idx}", "").strip()
            field_type = request.form.get(f"new_field_type_{idx}", "text").strip()
            if not name and not label_field:
                continue
            if not name:
                name = slugify_key(label_field)
            options = []
            if field_type in {"checkbox", "select"}:
                for opt_idx in range(1, 6):
                    opt_value = request.form.get(
                        f"new_field_options_{idx}_{opt_idx}", ""
                    ).strip()
                    if opt_value:
                        options.append(opt_value)
            new_fields.append(
                (name, label_field or name.replace("_", " ").title(), field_type, options)
            )
        existing_names = {field.name for field in fields}
        for name, _, _, _ in new_fields:
            if name in existing_names:
                flash("New field name conflicts with existing field.", "error")
                return redirect(url_for("edit_asset_type", asset_type_id=asset_type.id))
            existing_names.add(name)
        for name, label_field, field_type, options in new_fields:
            db.session.add(
                AssetField(
                    asset_type_id=asset_type.id,
                    name=name,
                    label=label_field,
                    field_type=field_type,
                    options=options,
                )
            )
        db.session.commit()
        new_values = {"label": asset_type.label}
        change_details = format_changes(old_values, new_values)
        details = asset_type.key
        if change_details:
            details = f"{details} changes={change_details}"
        log_audit("update", "asset_type", entity_id=asset_type.id, details=details)
        return redirect(url_for("list_asset_types"))
    return render_template(
        "asset_type_edit.html", asset_type=asset_type, fields=fields
    )


@app.route("/asset-types/delete/<int:asset_type_id>", methods=["POST"])
@login_required
@require_app_admin
def delete_asset_type(asset_type_id):
    asset_type = AssetType.query.get_or_404(asset_type_id)
    if AssetItem.query.filter_by(asset_type_id=asset_type.id).first():
        flash("Asset type has items and cannot be deleted.", "error")
        return redirect(url_for("list_asset_types"))
    RolePermission.query.filter_by(asset_type=f"custom:{asset_type.key}").delete()
    db.session.delete(asset_type)
    db.session.commit()
    log_audit("delete", "asset_type", entity_id=asset_type.id, details=asset_type.key)
    return redirect(url_for("list_asset_types"))


@app.route("/free")
@login_required
@require_permission("can_read")
def free_inventory():
    user = get_current_user()
    sections = []
    for asset_type, definition in ASSET_DEFS.items():
        perms = get_role_asset_permissions(user, asset_type)
        if not perms["can_read"]:
            continue
        is_consumable = definition.get("consumable", False)
        has_asset_tag = hasattr(definition["model"], "asset_tag") and not is_consumable
        if is_consumable:
            items = definition["model"].query.all()
        else:
            items = is_free_filter(definition["model"]).all()
        entries = []
        grouped_entries = {}
        total_assigned = 0
        total_available = 0
        for item in items:
            if is_consumable:
                available = max((item.total_quantity or 0) - (item.assigned_quantity or 0), 0)
                if available <= 0:
                    continue
                total_available += available
                total_assigned += max(item.assigned_quantity or 0, 0)
                model_label = " ".join(
                    part
                    for part in [item.size, item.speed, item.vendor]
                    if part
                )
                entries.append(
                    {
                        "id": item.id,
                        "model": model_label or "N/A",
                        "available": available,
                    }
                )
            else:
                total_available += 1
                model_value = getattr(item, "model", "N/A")
                if has_asset_tag:
                        entries.append(
                            {
                                "id": item.id,
                                "asset_tag": getattr(item, "asset_tag", None),
                                "model": model_value,
                                "vendor": getattr(item, "vendor", None),
                                "processor": getattr(item, "processor", None),
                                "ram": getattr(item, "ram", None),
                                "hard_disk": getattr(item, "hard_disk", None),
                                "assigned_to": item.assigned_to or "free",
                            }
                        )
                else:
                    grouped_entries[model_value] = grouped_entries.get(model_value, 0) + 1
        if not is_consumable:
            total_assigned = 0
            for item in definition["model"].query.all():
                if get_item_status(item) == "assigned":
                    total_assigned += 1
        if not is_consumable and not has_asset_tag:
            entries = [
                {"model": model_name, "quantity": count}
                for model_name, count in sorted(grouped_entries.items(), key=lambda item: str(item[0]))
            ]
        sections.append(
            {
                "key": asset_type,
                "label": definition["label"],
                "items": entries,
                "note": "",
                "consumable": is_consumable,
                "show_asset_tag": has_asset_tag,
                "grouped": (not is_consumable and not has_asset_tag),
                "available_count": total_available,
                "assigned_count": total_assigned,
            }
        )
    for asset_type in get_custom_asset_types():
        perms = get_role_asset_permissions(user, f"custom:{asset_type.key}")
        if not perms["can_read"]:
            continue
        fields = AssetField.query.filter_by(asset_type_id=asset_type.id).all()
        assigned_fields, model_fields = get_custom_special_fields(fields)
        asset_tag_fields = [field.name for field in fields if field.name.strip().lower() == "asset_tag"]
        has_asset_tag = bool(asset_tag_fields)
        items = AssetItem.query.filter_by(asset_type_id=asset_type.id).all()
        entries = []
        grouped_entries = {}
        note = ""
        total_assigned = 0
        total_available = 0
        if not assigned_fields:
            note = "No assigned_to field configured for this asset type."
        else:
            for item in items:
                assigned_to = None
                for field_name in assigned_fields:
                    assigned_to = (item.data or {}).get(field_name)
                    if assigned_to is not None:
                        break
                if assigned_to is None:
                    assigned_to = ""
                status = get_custom_status(item.data or {}, assigned_to)
                if status == "in stock":
                    total_available += 1
                    model = "N/A"
                    for field_name in model_fields:
                        model_value = (item.data or {}).get(field_name)
                        if model_value:
                            model = model_value
                            break
                    if model == "N/A":
                        for field in fields:
                            if field.field_type == "text" and field.name not in assigned_fields:
                                candidate = (item.data or {}).get(field.name)
                                if candidate:
                                    model = candidate
                                    break
                    if has_asset_tag:
                        tag_value = None
                        for tag_field in asset_tag_fields:
                            tag_value = (item.data or {}).get(tag_field)
                            if tag_value:
                                break
                        entries.append(
                            {
                                "id": item.id,
                                "asset_tag": tag_value,
                                "model": model,
                                "assigned_to": assigned_to or "free",
                            }
                        )
                    else:
                        grouped_entries[model] = grouped_entries.get(model, 0) + 1
                elif status == "assigned":
                    total_assigned += 1
        if assigned_fields and not total_assigned:
            total_assigned = len(items) - total_available
        if assigned_fields and not has_asset_tag:
            entries = [
                {"model": model_name, "quantity": count}
                for model_name, count in sorted(grouped_entries.items(), key=lambda item: str(item[0]))
            ]
        sections.append(
            {
                "key": f"custom-{asset_type.key}",
                "label": asset_type.label,
                "items": entries,
                "note": note,
                "available_count": total_available,
                "assigned_count": total_assigned,
                "show_asset_tag": has_asset_tag,
                "grouped": bool(assigned_fields and not has_asset_tag),
            }
        )
    return render_template("free.html", user=user, sections=sections)


@app.context_processor
def inject_user():
    user = get_current_user()
    ldap_users = []
    dept_options = []
    if user:
        ldap_users = get_ldap_user_display_list()
        dept_options = get_dept_options_cached()
    return {
        "current_user": user,
        "permissions": get_role_permissions(user),
        "ldap_users": ldap_users,
        "app_admin": user_has_app_admin(user),
        "branding": get_branding(),
        "branding_logo_url": get_branding_logo_url(),
        "branding_name": get_branding_name(),
        "dept_options": dept_options,
        "display_assignee": display_assignee,
        "update_status": get_update_status(),
        "dockerhub_repo": DOCKERHUB_REPO,
        "app_version": APP_VERSION or "dev",
    }


@app.route("/users")
@login_required
@require_app_admin
def list_users():
    users = User.query.order_by(User.username.asc()).all()
    roles = Role.query.order_by(Role.name.asc()).all()
    role_map = {role.id: role.name for role in roles}
    memberships = GroupMember.query.all()
    group_ids = {membership.group_id for membership in memberships}
    groups = Group.query.filter(Group.id.in_(group_ids)).all() if group_ids else []
    group_map = {group.id: group.name for group in groups}
    user_groups = {}
    for membership in memberships:
        name = group_map.get(membership.group_id)
        if not name:
            continue
        user_groups.setdefault(membership.user_id, []).append(name)
    user_roles = {}
    for link in UserRole.query.all():
        role_name = role_map.get(link.role_id)
        if not role_name:
            continue
        user_roles.setdefault(link.user_id, []).append(role_name)
    query = normalize_search(request.args.get("q", ""))
    if query:
        users = [
            user
            for user in users
            if matches_query(
                query,
                user.username,
                user.email,
                user.role,
                ", ".join(user_groups.get(user.id, [])),
                ", ".join(user_roles.get(user.id, [])),
            )
        ]
    return render_template(
        "users.html",
        users=users,
        roles=roles,
        user_groups=user_groups,
        user_roles=user_roles,
        query=query,
    )


@app.route("/user-assets")
@login_required
@require_permission("can_read")
def user_assets():
    query = (request.args.get("user") or "").strip()
    normalized = normalize_assignee(query)
    sections = []
    total_items = 0
    total_quantity = 0
    if normalized:
        for asset_key, definition in ASSET_DEFS.items():
            model = definition["model"]
            label = definition["label"]
            is_consumable = definition.get("consumable", False)
            rows = []
            assigned_count = 0
            display_fields = []
            preferred_order = [
                "asset_tag",
                "vendor",
                "model",
                "processor",
                "ram",
                "hard_disk",
                "screen_size",
                "size",
                "dept",
            ]
            for field_name, field_label, field_type in definition["fields"]:
                if field_name in {"assigned_to", "status"}:
                    continue
                if is_consumable and field_name == "total_quantity":
                    continue
                if is_consumable and field_name == "assigned_quantity":
                    field_label = "Quantity"
                display_fields.append((field_name, field_label, field_type))
            order_map = {name: idx for idx, name in enumerate(preferred_order)}
            display_fields.sort(key=lambda item: order_map.get(item[0], 100 + len(order_map)))
            if is_consumable:
                matches = model.query.filter(func.lower(model.assigned_to) == normalized).all()
                for item in matches:
                    qty = max(item.assigned_quantity or 0, 0)
                    if qty <= 0:
                        continue
                    assigned_count += qty
                    total_quantity += qty
                    row = {}
                    for field_name, _label, field_type in display_fields:
                        value = getattr(item, field_name, None)
                        if field_name == "assigned_quantity":
                            value = qty
                        row[field_name] = format_static_field_value(
                            field_name, field_type, value
                        )
                    rows.append(row)
            else:
                matches = model.query.filter(func.lower(model.assigned_to) == normalized).all()
                assigned_count = len(matches)
                total_items += assigned_count
                for item in matches:
                    row = {}
                    for field_name, _label, field_type in display_fields:
                        if field_name == "connection" and model is Mouse:
                            value = get_mouse_connection(item)
                            field_type = "text"
                        else:
                            value = getattr(item, field_name, None)
                        row[field_name] = format_static_field_value(
                            field_name, field_type, value
                        )
                    rows.append(row)
            if rows:
                sections.append(
                    {
                        "label": label,
                        "columns": display_fields,
                        "rows": rows,
                        "count": assigned_count,
                        "consumable": is_consumable,
                    }
                )
        for asset_type in get_custom_asset_types():
            fields = AssetField.query.filter_by(asset_type_id=asset_type.id).all()
            assigned_fields, model_fields = get_custom_special_fields(fields)
            if not assigned_fields:
                continue
            display_fields = [
                field
                for field in fields
                if not is_assigned_field(field)
                and field.name.strip().lower() != "status"
            ]
            rows = []
            assigned_count = 0
            for item in AssetItem.query.filter_by(asset_type_id=asset_type.id).all():
                assigned_to = None
                for field_name in assigned_fields:
                    assigned_to = (item.data or {}).get(field_name)
                    if assigned_to is not None:
                        break
                if normalize_assignee(assigned_to) != normalized:
                    continue
                row = {}
                for field in display_fields:
                    value = (item.data or {}).get(field.name)
                    row[field.name] = format_custom_field_value(
                        field, value, assigned_fields
                    )
                rows.append(row)
            assigned_count = len(rows)
            total_items += assigned_count
            if rows:
                sections.append(
                    {
                        "label": asset_type.label,
                        "columns": [(field.name, field.label, field.field_type) for field in display_fields],
                        "rows": rows,
                        "count": assigned_count,
                        "consumable": False,
                        "custom": True,
                    }
                )
    return render_template(
        "user_assets.html",
        query=query,
        display_query=display_assignee(query),
        sections=sections,
        total_items=total_items,
        total_quantity=total_quantity,
    )


@app.route("/ldap", methods=["GET", "POST"])
@login_required
@require_app_admin
def ldap_settings():
    config_row = LdapConfig.query.first()
    if request.method == "POST":
        action = request.form.get("action", "save")
        form_config = _ldap_config_from_form(request.form, existing=config_row)
        if action == "test":
            ok, message = _ldap_test_connection(form_config)
            flash(message, "success" if ok else "error")
            display_config = dict(form_config)
            display_config["bind_password"] = ""
            return render_template("ldap_settings.html", config=display_config)
        if config_row:
            for key, value in form_config.items():
                setattr(config_row, key, value)
        else:
            config_row = LdapConfig(**form_config)
            db.session.add(config_row)
        db.session.commit()
        flash("LDAP settings saved.", "success")
        return redirect(url_for("ldap_settings"))
    return render_template("ldap_settings.html", config=_ldap_form_values(config_row))


@app.route("/branding", methods=["GET", "POST"])
@login_required
@require_app_admin
def branding_settings():
    branding = BrandingConfig.query.first()
    if request.method == "POST":
        company_name = request.form.get("company_name", "").strip() or None
        remove_logo = request.form.get("remove_logo") == "1"
        if not branding:
            branding = BrandingConfig(company_name=company_name)
            db.session.add(branding)
        branding.company_name = company_name
        if remove_logo and branding.logo_filename:
            branding.logo_filename = None
        file = request.files.get("logo")
        if file and file.filename:
            ext = os.path.splitext(file.filename)[1].lower()
            if ext not in {".png", ".jpg", ".jpeg", ".gif"}:
                flash("Logo must be PNG, JPG, or GIF.", "error")
                return redirect(url_for("branding_settings"))
            folder = "/data/branding"
            os.makedirs(folder, exist_ok=True)
            filename = f"logo{ext}"
            filepath = os.path.join(folder, filename)
            file.save(filepath)
            branding.logo_filename = filename
        db.session.commit()
        flash("Branding updated.", "success")
        return redirect(url_for("branding_settings"))
    return render_template("branding.html", branding=branding)


@app.route("/branding/logo")
def branding_logo():
    branding = BrandingConfig.query.first()
    if not branding or not branding.logo_filename:
        return ("", 404)
    path = os.path.join("/data/branding", branding.logo_filename)
    if not os.path.exists(path):
        return ("", 404)
    return send_file(path)


@app.route("/departments", methods=["GET", "POST"])
@login_required
@require_permission("can_manage_depts")
def manage_departments():
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        if not name:
            flash("Department name is required.", "error")
            return redirect(url_for("manage_departments"))
        if Department.query.filter(func.lower(Department.name) == name.lower()).first():
            flash("Department already exists.", "error")
            return redirect(url_for("manage_departments"))
        db.session.add(Department(name=name))
        db.session.commit()
        flash("Department added.", "success")
        return redirect(url_for("manage_departments"))
    departments = Department.query.order_by(Department.name.asc()).all()
    return render_template("departments.html", departments=departments)


@app.route("/departments/<int:dept_id>/delete", methods=["POST"])
@login_required
@require_permission("can_manage_depts")
def delete_department(dept_id):
    dept = Department.query.get_or_404(dept_id)
    db.session.delete(dept)
    db.session.commit()
    flash("Department deleted.", "success")
    return redirect(url_for("manage_departments"))


@app.route("/logs")
@login_required
@require_app_admin
def view_logs():
    lines = _parse_int(request.args.get("lines", ""), 200)
    lines = max(min(lines, 2000), 50)
    log_text = read_log_tail(lines=lines)
    return render_template("logs.html", log_text=log_text, lines=lines)


@app.route("/logs/tail")
@login_required
@require_app_admin
def logs_tail():
    lines = _parse_int(request.args.get("lines", ""), 200)
    lines = max(min(lines, 2000), 50)
    log_text = read_log_tail(lines=lines)
    return app.response_class(log_text or "", mimetype="text/plain")


@app.route("/backups")
@login_required
@require_app_admin
def backup_settings():
    return render_template("backup.html")


@app.route("/backups/config")
@login_required
@require_app_admin
def backup_config_download():
    config_dump = {
        "generated_at": datetime.datetime.utcnow().isoformat(),
        "branding_config": serialize_model_list(BrandingConfig),
        "departments": serialize_model_list(Department),
        "roles": serialize_model_list(Role),
        "role_permissions": serialize_model_list(RolePermission),
        "user_roles": serialize_model_list(UserRole),
        "asset_types": serialize_model_list(AssetType),
        "asset_fields": serialize_model_list(AssetField),
        "builtin_asset_types": serialize_model_list(BuiltinAssetTypeSetting),
        "builtin_asset_fields": serialize_model_list(BuiltinAssetFieldSetting),
        "ldap_config": serialize_model_list(LdapConfig),
        "smtp_config": serialize_model_list(SMTPConfig),
        "smtp_recipients": serialize_model_list(SMTPRecipient),
        "groups": serialize_model_list(Group),
        "group_roles": serialize_model_list(GroupRole),
        "group_members": serialize_model_list(GroupMember),
    }
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("config.json", json.dumps(config_dump, default=str, indent=2))
        branding = BrandingConfig.query.first()
        if branding and branding.logo_filename:
            logo_path = os.path.join("/data/branding", branding.logo_filename)
            if os.path.exists(logo_path):
                archive.write(logo_path, arcname=f"branding/{branding.logo_filename}")
    output.seek(0)
    filename = f"asset-config-backup-{datetime.datetime.now().strftime('%Y%m%d-%H%M%S')}.zip"
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/zip")


@app.route("/backups/config/restore", methods=["POST"])
@login_required
@require_app_admin
def restore_config_backup():
    if "file" not in request.files:
        flash("No backup file uploaded.", "error")
        return redirect(url_for("backup_settings"))
    file = request.files["file"]
    if not file or not file.filename:
        flash("No backup file selected.", "error")
        return redirect(url_for("backup_settings"))
    if not file.filename.lower().endswith(".zip"):
        flash("Upload a .zip backup file.", "error")
        return redirect(url_for("backup_settings"))
    try:
        archive = zipfile.ZipFile(file, "r")
    except zipfile.BadZipFile:
        flash("Invalid backup file.", "error")
        return redirect(url_for("backup_settings"))
    if "config.json" not in archive.namelist():
        flash("Backup does not include config.json.", "error")
        return redirect(url_for("backup_settings"))
    try:
        config_data = json.loads(archive.read("config.json"))
    except Exception:
        flash("Backup config.json could not be read.", "error")
        return redirect(url_for("backup_settings"))

    db.session.query(RolePermission).delete()
    db.session.query(UserRole).delete()
    db.session.query(GroupRole).delete()
    db.session.query(GroupMember).delete()
    db.session.query(AssetField).delete()
    db.session.query(AssetType).delete()
    db.session.query(BuiltinAssetFieldSetting).delete()
    db.session.query(BuiltinAssetTypeSetting).delete()
    db.session.query(LdapConfig).delete()
    db.session.query(SMTPRecipient).delete()
    db.session.query(SMTPConfig).delete()
    db.session.query(Department).delete()
    db.session.query(BrandingConfig).delete()
    db.session.query(Group).delete()
    db.session.query(Role).delete()
    db.session.commit()

    _restore_model_rows(Role, config_data.get("roles", []))
    _restore_model_rows(RolePermission, config_data.get("role_permissions", []))
    _restore_model_rows(UserRole, config_data.get("user_roles", []))
    _restore_model_rows(Group, config_data.get("groups", []))
    _restore_model_rows(GroupRole, config_data.get("group_roles", []))
    _restore_model_rows(GroupMember, config_data.get("group_members", []))
    _restore_model_rows(Department, config_data.get("departments", []))
    _restore_model_rows(AssetType, config_data.get("asset_types", []))
    _restore_model_rows(AssetField, config_data.get("asset_fields", []))
    _restore_model_rows(BuiltinAssetTypeSetting, config_data.get("builtin_asset_types", []))
    _restore_model_rows(BuiltinAssetFieldSetting, config_data.get("builtin_asset_fields", []))
    _restore_model_rows(LdapConfig, config_data.get("ldap_config", []))
    _restore_model_rows(SMTPConfig, config_data.get("smtp_config", []))
    _restore_model_rows(SMTPRecipient, config_data.get("smtp_recipients", []))
    _restore_model_rows(BrandingConfig, config_data.get("branding_config", []))
    db.session.commit()

    os.makedirs("/data/branding", exist_ok=True)
    for name in archive.namelist():
        if not name.startswith("branding/"):
            continue
        filename = os.path.basename(name)
        if not filename:
            continue
        with archive.open(name) as source, open(os.path.join("/data/branding", filename), "wb") as target:
            target.write(source.read())

    _reset_sequences(
        [
            "role",
            "role_permission",
            "user_role",
            "group",
            "group_role",
            "group_member",
            "department",
            "asset_type",
            "asset_field",
            "builtin_asset_type_setting",
            "builtin_asset_field_setting",
            "ldap_config",
            "smtp_config",
            "smtp_recipient",
            "branding_config",
        ]
    )
    _DEPT_CACHE["timestamp"] = 0.0
    log_audit("update", "config_backup", details="Configuration restored")
    flash("Configuration restored.", "success")
    return redirect(url_for("backup_settings"))


@app.route("/ldap/sync", methods=["POST"])
@login_required
@require_app_admin
def ldap_sync_users():
    if not ldap_enabled():
        log_audit("sync_failed", "ldap_users", success=False, details="Not configured")
        flash("LDAP is not configured. Save LDAP settings first.", "error")
        return redirect(url_for("list_users"))
    try:
        records = get_ldap_user_records(force_refresh=True)
    except LDAPInvalidFilterError as exc:
        log_audit("sync_failed", "ldap_users", success=False, details=str(exc))
        flash(
            "LDAP sync failed. Your user/list filter looks invalid. Please review the LDAP "
            "User Filter and User List Filter fields.",
            "error",
        )
        return redirect(url_for("list_users"))
    except LDAPExceptionError as exc:
        log_audit("sync_failed", "ldap_users", success=False, details=str(exc))
        flash("LDAP sync failed. Please verify server settings and credentials.", "error")
        return redirect(url_for("list_users"))
    except Exception as exc:
        log_audit("sync_failed", "ldap_users", success=False, details=str(exc))
        flash("Unable to connect to LDAP server.", "error")
        return redirect(url_for("list_users"))
    if not records:
        log_audit("sync_failed", "ldap_users", success=False, details="No users found")
        flash("No LDAP users found or LDAP bind failed.", "error")
        return redirect(url_for("list_users"))
    created = 0
    updated = 0
    for record in records:
        username_norm = normalize_username(record.get("username", ""))
        if not username_norm:
            continue
        email = (record.get("email") or "").strip() or None
        existing = get_user_by_username_ci(username_norm)
        if existing:
            if email and existing.email != email:
                existing.email = email
                updated += 1
            continue
        try:
            new_user = ensure_ldap_user(username_norm)
        except LDAPInvalidFilterError as exc:
            log_audit("sync_failed", "ldap_users", success=False, details=str(exc))
            flash(
                "LDAP sync failed. Your User Filter looks invalid. "
                "Please fix the User Filter in LDAP settings.",
                "error",
            )
            return redirect(url_for("list_users"))
        except LDAPExceptionError as exc:
            log_audit("sync_failed", "ldap_users", success=False, details=str(exc))
            flash("LDAP sync failed. Please verify server settings and credentials.", "error")
            return redirect(url_for("list_users"))
        if email and new_user.email != email:
            new_user.email = email
        created += 1
    if updated:
        db.session.commit()
    log_audit("sync", "ldap_users", details=f"Added {created} users")
    message = f"LDAP sync complete. Added {created} users."
    if updated:
        message = f"{message} Updated {updated} emails."
    flash(message, "success")
    return redirect(url_for("list_users"))


@app.route("/ldap/sync-groups", methods=["POST"])
@login_required
@require_app_admin
def ldap_sync_groups():
    if not ldap_enabled():
        log_audit("sync_failed", "ldap_groups", success=False, details="Not configured")
        flash("LDAP is not configured. Save LDAP settings first.", "error")
        return redirect(url_for("list_groups"))
    try:
        groups = get_ldap_groups(force_refresh=True)
    except LDAPInvalidFilterError as exc:
        log_audit("sync_failed", "ldap_groups", success=False, details=str(exc))
        flash(
            "LDAP sync failed. Your group filter looks invalid. Please review the LDAP "
            "Group Filter field.",
            "error",
        )
        return redirect(url_for("list_groups"))
    except LDAPExceptionError as exc:
        log_audit("sync_failed", "ldap_groups", success=False, details=str(exc))
        flash("LDAP sync failed. Please verify server settings and credentials.", "error")
        return redirect(url_for("list_groups"))
    except Exception as exc:
        log_audit("sync_failed", "ldap_groups", success=False, details=str(exc))
        flash("Unable to connect to LDAP server.", "error")
        return redirect(url_for("list_groups"))
    if not groups:
        log_audit("sync_failed", "ldap_groups", success=False, details="No groups found")
        flash("No LDAP groups found or LDAP bind failed.", "error")
        return redirect(url_for("list_groups"))
    created = 0
    updated = 0
    for group in groups:
        name = group.get("name")
        if not name:
            continue
        existing = Group.query.filter_by(name=name).first()
        if not existing:
            existing = Group(name=name, role="unassigned")
            db.session.add(existing)
            db.session.flush()
            created += 1
        else:
            updated += 1
        GroupMember.query.filter_by(group_id=existing.id).delete()
        for member_dn in group.get("members", []):
            match = re.search(r"CN=([^,]+)", str(member_dn), re.IGNORECASE)
            username = match.group(1) if match else None
            if not username:
                continue
            username_norm = normalize_username(username)
            user = get_user_by_username_ci(username_norm)
            if not user:
                user = ensure_ldap_user(username_norm)
            db.session.add(GroupMember(group_id=existing.id, user_id=user.id))
    db.session.commit()
    log_audit(
        "sync",
        "ldap_groups",
        details=f"Added {created} groups, updated {updated} groups",
    )
    flash(
        f"LDAP sync complete. Added {created} groups, updated {updated} groups.",
        "success",
    )
    return redirect(url_for("list_groups"))


@app.route("/groups")
@login_required
@require_app_admin
def list_groups():
    groups = Group.query.order_by(Group.name.asc()).all()
    roles = Role.query.order_by(Role.name.asc()).all()
    role_map = {role.id: role.name for role in roles}
    members = GroupMember.query.all()
    member_counts = {}
    for membership in members:
        member_counts[membership.group_id] = member_counts.get(membership.group_id, 0) + 1
    group_roles = {}
    for link in GroupRole.query.all():
        role_name = role_map.get(link.role_id)
        if not role_name:
            continue
        group_roles.setdefault(link.group_id, []).append(role_name)
    query = normalize_search(request.args.get("q", ""))
    if query:
        groups = [
            group
            for group in groups
            if matches_query(
                query,
                group.name,
                ", ".join(group_roles.get(group.id, [])),
            )
        ]
    return render_template(
        "groups.html",
        groups=groups,
        roles=roles,
        member_counts=member_counts,
        group_roles=group_roles,
        query=query,
    )


@app.route("/groups/add", methods=["GET", "POST"])
@login_required
@require_app_admin
def add_group():
    roles = Role.query.order_by(Role.name.asc()).all()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        role_names = request.form.getlist("roles")
        if not name:
            flash("Group name is required.", "error")
            return redirect(url_for("add_group"))
        if Group.query.filter_by(name=name).first():
            flash("Group already exists.", "error")
            return redirect(url_for("add_group"))
        selected_roles = Role.query.filter(Role.name.in_(role_names)).all()
        if not selected_roles:
            selected_roles = Role.query.filter_by(name="unassigned").all()
        primary_role = selected_roles[0].name
        group = Group(name=name, role=primary_role)
        db.session.add(group)
        db.session.flush()
        for role in selected_roles:
            db.session.add(GroupRole(group_id=group.id, role_id=role.id))
        db.session.commit()
        log_audit("create", "group", entity_id=group.id, details=name)
        return redirect(url_for("list_groups"))
    return render_template("group_add.html", roles=roles)


@app.route("/groups/edit/<int:group_id>", methods=["GET", "POST"])
@login_required
@require_app_admin
def edit_group(group_id):
    group = Group.query.get_or_404(group_id)
    roles = Role.query.order_by(Role.name.asc()).all()
    users = User.query.order_by(User.username.asc()).all()
    old_member_ids = [member.user_id for member in GroupMember.query.filter_by(group_id=group.id).all()]
    old_role_ids = [role.role_id for role in GroupRole.query.filter_by(group_id=group.id).all()]
    old_values = {
        "name": group.name,
        "roles": sorted(old_role_ids),
        "members": sorted(old_member_ids),
    }
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        role_names = request.form.getlist("roles")
        if not name:
            flash("Group name is required.", "error")
            return redirect(url_for("edit_group", group_id=group.id))
        existing = Group.query.filter_by(name=name).first()
        if existing and existing.id != group.id:
            flash("Group name already exists.", "error")
            return redirect(url_for("edit_group", group_id=group.id))
        selected_roles = Role.query.filter(Role.name.in_(role_names)).all()
        if not selected_roles:
            selected_roles = Role.query.filter_by(name="unassigned").all()
        group.name = name
        group.role = selected_roles[0].name
        GroupRole.query.filter_by(group_id=group.id).delete()
        for role in selected_roles:
            db.session.add(GroupRole(group_id=group.id, role_id=role.id))
        selected_ids = {int(uid) for uid in request.form.getlist("members") if uid.isdigit()}
        GroupMember.query.filter_by(group_id=group.id).delete()
        for user in users:
            if user.id in selected_ids:
                db.session.add(GroupMember(group_id=group.id, user_id=user.id))
        db.session.commit()
        new_role_ids = [role.id for role in selected_roles]
        new_values = {
            "name": group.name,
            "roles": sorted(new_role_ids),
            "members": sorted(selected_ids),
        }
        change_details = format_changes(old_values, new_values)
        details = group.name
        if change_details:
            details = f"{details} changes={change_details}"
        log_audit("update", "group", entity_id=group.id, details=details)
        return redirect(url_for("list_groups"))
    existing_members = {
        member.user_id
        for member in GroupMember.query.filter_by(group_id=group.id).all()
    }
    existing_roles = {
        role.role_id for role in GroupRole.query.filter_by(group_id=group.id).all()
    }
    return render_template(
        "group_edit.html",
        group=group,
        roles=roles,
        users=users,
        existing_members=existing_members,
        existing_roles=existing_roles,
    )


@app.route("/groups/delete/<int:group_id>", methods=["POST"])
@login_required
@require_app_admin
def delete_group(group_id):
    group = Group.query.get_or_404(group_id)
    GroupMember.query.filter_by(group_id=group.id).delete()
    db.session.delete(group)
    db.session.commit()
    log_audit("delete", "group", entity_id=group.id, details=group.name)
    return redirect(url_for("list_groups"))


@app.route("/groups/bulk-delete", methods=["POST"])
@login_required
@require_bulk_delete
def bulk_delete_groups():
    ids = [int(item_id) for item_id in request.form.getlist("selected_ids") if item_id.isdigit()]
    if ids:
        groups = Group.query.filter(Group.id.in_(ids)).all()
        for group in groups:
            GroupMember.query.filter_by(group_id=group.id).delete()
            GroupRole.query.filter_by(group_id=group.id).delete()
            db.session.delete(group)
        db.session.commit()
        log_audit("bulk_delete", "group", details=f"Deleted ids: {ids}")
    return redirect(url_for("list_groups"))


@app.route("/users/add", methods=["GET", "POST"])
@login_required
@require_app_admin
def add_user():
    roles = Role.query.order_by(Role.name.asc()).all()
    if request.method == "POST":
        username = normalize_username(request.form.get("username", ""))
        password = request.form.get("password", "")
        email = request.form.get("email", "").strip() or None
        role_names = request.form.getlist("roles")
        if not username or not password or not role_names:
            flash("All fields are required.", "error")
            return redirect(url_for("add_user"))
        if get_user_by_username_ci(username):
            flash("Username already exists.", "error")
            return redirect(url_for("add_user"))
        selected_roles = Role.query.filter(Role.name.in_(role_names)).all()
        if not selected_roles:
            flash("Selected roles do not exist.", "error")
            return redirect(url_for("add_user"))
        primary_role = selected_roles[0].name
        user = User(
            username=username,
            password_hash=generate_password_hash(password),
            role=primary_role,
            email=email,
        )
        db.session.add(user)
        db.session.flush()
        for role in selected_roles:
            db.session.add(UserRole(user_id=user.id, role_id=role.id))
        db.session.commit()
        log_audit("create", "user", entity_id=user.id, details=username)
        return redirect(url_for("list_users"))
    return render_template("user_add.html", roles=roles)


@app.route("/users/edit/<int:user_id>", methods=["GET", "POST"])
@login_required
@require_app_admin
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    roles = Role.query.order_by(Role.name.asc()).all()
    old_roles = [role.role_id for role in UserRole.query.filter_by(user_id=user.id).all()]
    old_values = {
        "username": user.username,
        "roles": sorted(old_roles),
    }
    if request.method == "POST":
        username = normalize_username(request.form.get("username", ""))
        password = request.form.get("password", "")
        email = request.form.get("email", "").strip() or None
        role_names = request.form.getlist("roles")
        if not username or not role_names:
            flash("Username and role are required.", "error")
            return redirect(url_for("edit_user", user_id=user.id))
        existing = get_user_by_username_ci(username)
        if existing and existing.id != user.id:
            flash("Username already exists.", "error")
            return redirect(url_for("edit_user", user_id=user.id))
        selected_roles = Role.query.filter(Role.name.in_(role_names)).all()
        if not selected_roles:
            flash("Selected roles do not exist.", "error")
            return redirect(url_for("edit_user", user_id=user.id))
        user.username = username
        user.role = selected_roles[0].name
        user.email = email
        if password:
            user.password_hash = generate_password_hash(password)
        UserRole.query.filter_by(user_id=user.id).delete()
        for role in selected_roles:
            db.session.add(UserRole(user_id=user.id, role_id=role.id))
        db.session.commit()
        new_roles = [role.id for role in selected_roles]
        new_values = {"username": user.username, "roles": sorted(new_roles)}
        change_details = format_changes(old_values, new_values)
        details = user.username
        if change_details:
            details = f"{details} changes={change_details}"
        log_audit("update", "user", entity_id=user.id, details=details)
        return redirect(url_for("list_users"))
    existing_roles = {
        role.role_id for role in UserRole.query.filter_by(user_id=user.id).all()
    }
    return render_template(
        "user_edit.html", user=user, roles=roles, existing_roles=existing_roles
    )


@app.route("/users/delete/<int:user_id>", methods=["POST"])
@login_required
@require_app_admin
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.username == "admin":
        flash("Default admin user cannot be deleted.", "error")
        return redirect(url_for("list_users"))
    db.session.delete(user)
    db.session.commit()
    log_audit("delete", "user", entity_id=user.id, details=user.username)
    return redirect(url_for("list_users"))


@app.route("/users/bulk-delete", methods=["POST"])
@login_required
@require_bulk_delete
def bulk_delete_users():
    ids = [int(item_id) for item_id in request.form.getlist("selected_ids") if item_id.isdigit()]
    if ids:
        users = User.query.filter(User.id.in_(ids)).all()
        current = get_current_user()
        for user in users:
            if user.username == "admin":
                continue
            if current and user.id == current.id:
                continue
            db.session.delete(user)
        db.session.commit()
        log_audit("bulk_delete", "user", details=f"Deleted ids: {ids}")
    return redirect(url_for("list_users"))


@app.route("/roles")
@login_required
@require_app_admin
def list_roles():
    roles = Role.query.order_by(Role.name.asc()).all()
    perms = RolePermission.query.all()
    perms_map = {}
    for perm in perms:
        perms_map.setdefault(perm.role_id, {})[perm.asset_type] = perm
    return render_template(
        "roles.html", roles=roles, assets=get_asset_display_list(), perms_map=perms_map
    )


@app.route("/audit")
@login_required
@require_app_admin
def audit_log():
    query = normalize_search(request.args.get("q", ""))
    logs = AuditLog.query.order_by(AuditLog.created_at.desc()).all()
    if query:
        logs = [
            entry
            for entry in logs
            if matches_query(
                query,
                entry.username,
                entry.action,
                entry.entity_type,
                entry.entity_id,
                entry.ip_address,
                entry.details,
            )
        ]
    return render_template("audit.html", logs=logs, query=query)


@app.route("/smtp", methods=["GET", "POST"])
@login_required
@require_app_admin
def smtp_settings():
    config = SMTPConfig.query.first()
    if request.method == "POST":
        host = request.form.get("host", "").strip()
        port = _parse_int(request.form.get("port", ""), 0) or None
        encryption = request.form.get("encryption", "none").strip().lower()
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        skip_auth = "skip_auth" in request.form
        sender_email = request.form.get("sender_email", "").strip()
        enabled = "enabled" in request.form
        monthly_report_enabled = "monthly_report_enabled" in request.form
        monthly_report_day = _parse_int(request.form.get("monthly_report_day", ""), 1)
        low_stock_enabled = "low_stock_enabled" in request.form
        low_stock_threshold = _parse_int(request.form.get("low_stock_threshold", ""), 5)
        low_stock_frequency_days = _parse_int(request.form.get("low_stock_frequency_days", ""), 1)
        if config:
            config.host = host
            config.port = port
            config.encryption = encryption
            config.username = username
            if password:
                config.password = password
            config.skip_auth = skip_auth
            config.sender_email = sender_email
            config.enabled = enabled
            config.monthly_report_enabled = monthly_report_enabled
            config.monthly_report_day = monthly_report_day
            config.low_stock_enabled = low_stock_enabled
            config.low_stock_threshold = low_stock_threshold
            config.low_stock_frequency_days = low_stock_frequency_days
        else:
            config = SMTPConfig(
                host=host,
                port=port,
                encryption=encryption,
                username=username,
                password=password,
                skip_auth=skip_auth,
                sender_email=sender_email,
                enabled=enabled,
                monthly_report_enabled=monthly_report_enabled,
                monthly_report_day=monthly_report_day,
                low_stock_enabled=low_stock_enabled,
                low_stock_threshold=low_stock_threshold,
                low_stock_frequency_days=low_stock_frequency_days,
            )
            db.session.add(config)
        db.session.commit()
        flash("SMTP settings saved.", "success")
        return redirect(url_for("smtp_settings"))
    recipients = SMTPRecipient.query.order_by(SMTPRecipient.email.asc()).all()
    return render_template("smtp.html", config=config, recipients=recipients)


@app.route("/smtp/send-monthly", methods=["POST"])
@login_required
@require_app_admin
def smtp_send_monthly():
    ok = send_monthly_report(force=True)
    flash("Monthly report sent." if ok else "Monthly report not sent.", "success" if ok else "error")
    return redirect(url_for("smtp_settings"))


@app.route("/smtp/send-low-stock", methods=["POST"])
@login_required
@require_app_admin
def smtp_send_low_stock():
    ok = send_low_stock_report(force=True)
    flash("Low stock report sent." if ok else "Low stock report not sent.", "success" if ok else "error")
    return redirect(url_for("smtp_settings"))


@app.route("/smtp/recipients/add", methods=["POST"])
@login_required
@require_app_admin
def smtp_add_recipient():
    email = request.form.get("email", "").strip()
    if not email:
        flash("Recipient email is required.", "error")
        return redirect(url_for("smtp_settings"))
    if SMTPRecipient.query.filter_by(email=email).first():
        flash("Recipient already exists.", "error")
        return redirect(url_for("smtp_settings"))
    recipient = SMTPRecipient(
        email=email,
        notify_create="notify_create" in request.form,
        notify_update="notify_update" in request.form,
        notify_delete="notify_delete" in request.form,
        notify_bulk_delete="notify_bulk_delete" in request.form,
        notify_monthly="notify_monthly" in request.form,
        notify_low_stock="notify_low_stock" in request.form,
    )
    db.session.add(recipient)
    db.session.commit()
    flash("Recipient added.", "success")
    return redirect(url_for("smtp_settings"))


@app.route("/smtp/recipients/<int:recipient_id>/update", methods=["POST"])
@login_required
@require_app_admin
def smtp_update_recipient(recipient_id):
    recipient = SMTPRecipient.query.get_or_404(recipient_id)
    recipient.notify_create = "notify_create" in request.form
    recipient.notify_update = "notify_update" in request.form
    recipient.notify_delete = "notify_delete" in request.form
    recipient.notify_bulk_delete = "notify_bulk_delete" in request.form
    recipient.notify_monthly = "notify_monthly" in request.form
    recipient.notify_low_stock = "notify_low_stock" in request.form
    db.session.commit()
    flash("Recipient updated.", "success")
    return redirect(url_for("smtp_settings"))


@app.route("/smtp/recipients/<int:recipient_id>/delete", methods=["POST"])
@login_required
@require_app_admin
def smtp_delete_recipient(recipient_id):
    recipient = SMTPRecipient.query.get_or_404(recipient_id)
    db.session.delete(recipient)
    db.session.commit()
    flash("Recipient deleted.", "success")
    return redirect(url_for("smtp_settings"))


@app.route("/roles/add", methods=["GET", "POST"])
@login_required
@require_app_admin
def add_role():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("Role name is required.", "error")
            return redirect(url_for("add_role"))
        if Role.query.filter_by(name=name).first():
            flash("Role already exists.", "error")
            return redirect(url_for("add_role"))
        role = Role(name=name, can_add=False, can_delete=False, can_read=False)
        db.session.add(role)
        db.session.flush()
        role.can_manage_depts = "can_manage_depts" in request.form
        any_add = False
        any_delete = False
        any_read = False
        any_bulk_delete = False
        for asset in get_asset_display_list():
            asset_key = asset["key"]
            can_read = f"perm_{asset_key}_read" in request.form
            can_add = f"perm_{asset_key}_add" in request.form
            can_delete = f"perm_{asset_key}_delete" in request.form
            can_bulk_delete = f"perm_{asset_key}_bulk_delete" in request.form
            any_add = any_add or can_add
            any_delete = any_delete or can_delete
            any_read = any_read or can_read
            any_bulk_delete = any_bulk_delete or can_bulk_delete
            db.session.add(
                RolePermission(
                    role_id=role.id,
                    asset_type=asset_key,
                    can_add=can_add,
                    can_delete=can_delete,
                    can_read=can_read,
                    can_bulk_delete=can_bulk_delete,
                )
            )
        role.can_add = any_add
        role.can_delete = any_delete
        role.can_read = any_read
        role.can_bulk_delete = any_bulk_delete
        db.session.commit()
        log_audit("create", "role", entity_id=role.id, details=role.name)
        return redirect(url_for("list_roles"))
    return render_template("role_add.html", assets=get_asset_display_list())


@app.route("/roles/edit/<int:role_id>", methods=["GET", "POST"])
@login_required
@require_app_admin
def edit_role(role_id):
    role = Role.query.get_or_404(role_id)
    assets = get_asset_display_list()
    old_values = {"name": role.name}
    old_perms = RolePermission.query.filter_by(role_id=role.id).all()
    old_perm_map = {
        perm.asset_type: {
            "read": perm.can_read,
            "add": perm.can_add,
            "delete": perm.can_delete,
            "bulk": perm.can_bulk_delete,
        }
        for perm in old_perms
    }
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("Role name is required.", "error")
            return redirect(url_for("edit_role", role_id=role.id))
        existing = Role.query.filter_by(name=name).first()
        if existing and existing.id != role.id:
            flash("Role name already exists.", "error")
            return redirect(url_for("edit_role", role_id=role.id))
        role.name = name
        role.can_manage_depts = "can_manage_depts" in request.form
        any_add = False
        any_delete = False
        any_read = False
        any_bulk_delete = False
        for asset in assets:
            asset_key = asset["key"]
            can_read = f"perm_{asset_key}_read" in request.form
            can_add = f"perm_{asset_key}_add" in request.form
            can_delete = f"perm_{asset_key}_delete" in request.form
            can_bulk_delete = f"perm_{asset_key}_bulk_delete" in request.form
            any_add = any_add or can_add
            any_delete = any_delete or can_delete
            any_read = any_read or can_read
            any_bulk_delete = any_bulk_delete or can_bulk_delete
            perm = RolePermission.query.filter_by(
                role_id=role.id, asset_type=asset_key
            ).first()
            if not perm:
                perm = RolePermission(
                    role_id=role.id,
                    asset_type=asset_key,
                    can_add=can_add,
                    can_delete=can_delete,
                    can_read=can_read,
                    can_bulk_delete=can_bulk_delete,
                )
                db.session.add(perm)
            else:
                perm.can_add = can_add
                perm.can_delete = can_delete
                perm.can_read = can_read
                perm.can_bulk_delete = can_bulk_delete
        role.can_add = any_add
        role.can_delete = any_delete
        role.can_read = any_read
        role.can_bulk_delete = any_bulk_delete
        db.session.commit()
        new_values = {"name": role.name}
        new_perms = RolePermission.query.filter_by(role_id=role.id).all()
        new_perm_map = {
            perm.asset_type: {
                "read": perm.can_read,
                "add": perm.can_add,
                "delete": perm.can_delete,
                "bulk": perm.can_bulk_delete,
            }
            for perm in new_perms
        }
        change_details = format_changes(old_values, new_values)
        perm_changes = []
        for asset_key in sorted(set(old_perm_map.keys()) | set(new_perm_map.keys())):
            if old_perm_map.get(asset_key) != new_perm_map.get(asset_key):
                perm_changes.append(f"{asset_key}: {old_perm_map.get(asset_key)} -> {new_perm_map.get(asset_key)}")
        details = role.name
        if change_details:
            details = f"{details} changes={change_details}"
        if perm_changes:
            details = f"{details} perms={'; '.join(perm_changes)}"
        log_audit("update", "role", entity_id=role.id, details=details)
        return redirect(url_for("list_roles"))
    perms = RolePermission.query.filter_by(role_id=role.id).all()
    perms_map = {perm.asset_type: perm for perm in perms}
    return render_template(
        "role_edit.html", role=role, assets=assets, perms_map=perms_map
    )


@app.route("/roles/delete/<int:role_id>", methods=["POST"])
@login_required
@require_app_admin
def delete_role(role_id):
    role = Role.query.get_or_404(role_id)
    if role.name in {"admin", "operator", "reader"}:
        flash("Default roles cannot be deleted.", "error")
        return redirect(url_for("list_roles"))
    if User.query.filter_by(role=role.name).first():
        flash("Role is assigned to users and cannot be deleted.", "error")
        return redirect(url_for("list_roles"))
    db.session.delete(role)
    db.session.commit()
    log_audit("delete", "role", entity_id=role.id, details=role.name)
    return redirect(url_for("list_roles"))


@app.route("/roles/bulk-delete", methods=["POST"])
@login_required
@require_bulk_delete
def bulk_delete_roles():
    ids = [int(item_id) for item_id in request.form.getlist("selected_ids") if item_id.isdigit()]
    if ids:
        roles = Role.query.filter(Role.id.in_(ids)).all()
        for role in roles:
            if role.name in {"admin", "operator", "reader", "app_admin", "unassigned"}:
                continue
            if UserRole.query.filter_by(role_id=role.id).first():
                continue
            if GroupRole.query.filter_by(role_id=role.id).first():
                continue
            if User.query.filter_by(role=role.name).first():
                continue
            db.session.delete(role)
        db.session.commit()
        log_audit("bulk_delete", "role", details=f"Deleted ids: {ids}")
    return redirect(url_for("list_roles"))


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
